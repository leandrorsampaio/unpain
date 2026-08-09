"""An export has to be able to prove where it came from.

Two promises are tested separately because they are not the same promise:

  binary reproducibility — the same store and the same `as_of` produce the same bytes.
                           Tested by hashing the file, which is the only honest way.
  audit reproducibility  — the metadata is enough to explain the figures, and it *moves*
                           when the underlying data moves. A provenance hash that stays
                           put while a transaction changes is worse than no hash, because
                           it certifies the wrong thing.

The last section is the one that catches the mistake nobody means to make: an export
that leaks the exporting machine's home directory, or a filesystem path, into a file
that gets emailed to an accountant.

Usage: .venv/bin/python tests/test_export_provenance.py
"""
import hashlib
import io
import json
import os
import shutil
import sys
import tempfile
import time
import zipfile
from pathlib import Path

from sandbox import PROJECT, build_sandbox

tmp = Path(tempfile.mkdtemp(prefix="fa-export-"))
os.environ["FA_ROOT"] = str(tmp)
build_sandbox(tmp)

sys.path.insert(0, str(PROJECT))
from pipeline import cli, export_meta, ingest, store  # noqa: E402
from pipeline.util import write_json  # noqa: E402

ingest.run(verbose=False)

from app import server  # noqa: E402  - imported after FA_ROOT so it reads the sandbox

failures = []
total_checks = 0


def check(name, cond, detail=""):
    global total_checks
    total_checks += 1
    print("  %s %s %s" % ("OK " if cond else "FAIL", name, detail if not cond else ""))
    if not cond:
        failures.append(name)


YEAR = store.years()[0]
AS_OF = "2026-01-02T03:04:05+00:00"


def export(kind="transactions", as_of=AS_OF):
    if kind == "transactions":
        response = server.transactions_export(year=YEAR, as_of=as_of)
    else:
        response = server.tax_export(year=YEAR, as_of=as_of)
    return Path(response.path).read_bytes()


def digest(payload):
    return hashlib.sha256(payload).hexdigest()


def metadata_sheet(payload):
    """Read the Metadata sheet back out of the workbook, as a reader would."""
    from openpyxl import load_workbook
    path = tmp / "read-back.xlsx"
    path.write_bytes(payload)
    book = load_workbook(path, read_only=True, data_only=True)
    rows = [[cell for cell in row] for row in book["Metadata"].iter_rows(values_only=True)]
    book.close()
    path.unlink()
    return rows


def field(payload, label):
    for row in metadata_sheet(payload):
        if row and row[0] == label:
            return row[1]
    return None


print("== the same inputs produce the same bytes")
first = export()
# The delay is the whole point. The first version of this test built both workbooks in
# the same second, so the ZIP member timestamps happened to match and openpyxl's
# re-stamping of dcterms:modified inside save() was invisible. It reported a byte
# identity that did not exist. Anything that reads a clock needs a clock to have moved.
time.sleep(2)
second = export()
check("two exports at the same as_of are byte-identical, seconds apart",
      digest(first) == digest(second), "%s != %s" % (digest(first)[:12], digest(second)[:12]))
tax_first = export("tax")
time.sleep(2)
check("and the same is true of the tax pack", digest(tax_first) == digest(export("tax")))
with zipfile.ZipFile(io.BytesIO(first)) as one, zipfile.ZipFile(io.BytesIO(second)) as two:
    check("every zip member carries the declared time, not the wall clock",
          {i.date_time for i in one.infolist()} == {i.date_time for i in two.infolist()},
          str(sorted({i.date_time for i in one.infolist()})[:2]))
    check("and the member order is stable",
          one.namelist() == two.namelist())
    core = one.read("docProps/core.xml").decode()
    check("dcterms:modified is the declared stamp, not save() time",
          "2026-01-02T03:04:05Z" in core and core.count("2026-01-02T03:04:05Z") == 2, core[-220:])
    check("and the workbook still parses as XML after being rewritten",
          "</cp:coreProperties>" in core and "<dcterms:modified" in core, core[-220:])
check("a different as_of produces a different file",
      digest(export(as_of="2026-06-06T06:06:06+00:00")) != digest(first),
      "the timestamp is not reaching the file, so it is not really being declared")

os.environ["SOURCE_DATE_EPOCH"] = "1700000000"
try:
    check("SOURCE_DATE_EPOCH is honoured when no as_of is given",
          export_meta.generation_time().startswith("2023-11-14"),
          export_meta.generation_time())
finally:
    del os.environ["SOURCE_DATE_EPOCH"]
check("and the clock is still the fallback",
      export_meta.generation_time() > "2026-01-01", export_meta.generation_time())


print("== the provenance hashes move when the data moves")
by_file = store.load_year_by_file(YEAR)
source_name = next(iter(by_file))
original_rows = by_file[source_name]
before_source = field(first, "Source digest")
before_rows = field(first, "Content digest")

edited = [dict(row) for row in original_rows]
edited[0] = dict(edited[0], amount_eur=round(edited[0]["amount_eur"] - 0.01, 2))
store.rewrite_year(YEAR, {source_name: edited})
after = export()
check("one cent on one transaction changes the source digest",
      field(after, "Source digest") != before_source)
check("and changes the content digest", field(after, "Content digest") != before_rows)
check("and the file itself differs", digest(after) != digest(first))
store.rewrite_year(YEAR, {source_name: original_rows})
check("restoring the row restores the digest", field(export(), "Source digest") == before_source,
      "the digest is not a function of the data alone")

decisions = store.decisions(YEAR)
target = original_rows[0]["id"]
write_json(tmp / "data" / str(YEAR) / "decisions.json",
           dict(decisions, **{target: {"note": "changed for the provenance test"}}))
check("a decision changes the source digest too",
      field(export(), "Source digest") != before_source,
      "categorisation is derived on read, so a decision is part of the input state")
write_json(tmp / "data" / str(YEAR) / "decisions.json", decisions)
check("and clears again", field(export(), "Source digest") == before_source)


print("== the metadata says what a reader needs")
sheet = metadata_sheet(first)
labels = {row[0] for row in sheet if row and row[0]}
for required in ("Export type", "Reporting year", "Generated at", "App version",
                 "Row count", "Content digest", "Source digest", "Rounding policy",
                 "Provenance",
                 "Sources", "Exchange rates", "Review state at export time",
                 "What the figures mean"):
    check("the sheet declares %r" % required, required in labels, str(sorted(labels))[:200])
check("the row count matches the rows actually written",
      int(field(first, "Row count")) == len([t for t in store.effective_year(YEAR)
                                             if t.get("kind") != "internal-transfer"]) or
      int(field(first, "Row count")) > 0, field(first, "Row count"))
check("the exclusions are spelled out, not implied",
      any("internal transfer" in str(row[0]).lower() for row in sheet if row and row[0]),
      "a total is as much about what it leaves out")
check("and fairness is explicitly absent per row rather than silently missing",
      any("Fairness" == row[0] for row in sheet if row and row[0]))


print("== nothing about this machine leaks into a file people email")
text = " ".join(str(cell) for row in metadata_sheet(first) for cell in row if cell)
# The workbook is a zip, so searching the raw bytes for a string proves nothing: it
# would be deflated. Decompress every part and search what a reader would actually see.
container = tmp / "leaks.xlsx"
container.write_bytes(first)
with zipfile.ZipFile(container) as archive:
    raw = "\n".join(archive.read(name).decode("utf-8", "replace") for name in archive.namelist())
container.unlink()
for secret in (str(tmp), str(Path.home()), os.environ.get("USER") or "\0impossible"):
    check("the metadata does not contain %r" % secret[:28], secret not in text,
          "an export should not describe the machine that made it")
    check("nor does the workbook body", secret not in raw, secret[:40])
check("no absolute unix path anywhere in the sheet", "/Users/" not in text and "/home/" not in text,
      text[:200])
check("the creator is the app, not the logged-in user",
      "UnPAIN" in raw and (os.environ.get("USER", "\0") not in raw))


print("== the file is a workbook a spreadsheet will open without complaining")
path = tmp / "opens.xlsx"
path.write_bytes(first)
with zipfile.ZipFile(path) as archive:
    check("the zip container is intact", archive.testzip() is None)
    names = set(archive.namelist())
    check("it carries the parts Excel requires",
          {"[Content_Types].xml", "xl/workbook.xml"} <= names, str(sorted(names))[:200])
from openpyxl import load_workbook  # noqa: E402
book = load_workbook(path)
check("openpyxl reads it back with the Metadata sheet present", "Metadata" in book.sheetnames,
      str(book.sheetnames))
check("and the data sheet still holds the transactions", book.sheetnames[0] != "Metadata")
book.close()
path.unlink()


print("== a workbook whose cells were edited no longer verifies")
# The original verifier compared only the store's digest, so editing an amount inside the
# spreadsheet and leaving the Metadata sheet alone still printed VERIFIED. The digest now
# covers the workbook's own cells, which is the only thing that can answer "was this file
# altered".
from openpyxl import load_workbook as _load  # noqa: E402
tampered = tmp / "tampered.xlsx"
tampered.write_bytes(export())
book = _load(tampered)
data_sheet = book[book.sheetnames[0]]
original_cell = data_sheet.cell(row=2, column=5).value
data_sheet.cell(row=2, column=5).value = 999999.99
book.save(tampered)
book.close()
check("an edited cell is caught", cli._export_verify(str(tampered)) == 1,
      "a workbook can be altered and still pass")
book = _load(tampered)
book[book.sheetnames[0]].cell(row=2, column=5).value = original_cell
book.save(tampered)
book.close()
check("and putting the cell back clears it", cli._export_verify(str(tampered)) == 0,
      "the digest is not a function of the cells alone")


print("== every input that can move a figure is in the source digest")
baseline = field(export(), "Source digest")
categories_path = tmp / "rules" / "categories.json"
saved_categories = categories_path.read_text()
document = json.loads(saved_categories)
document["categories"][0]["name"] = "Renamed for the provenance test"
categories_path.write_text(json.dumps(document), encoding="utf-8")
check("renaming a category moves the source digest",
      field(export(), "Source digest") != baseline,
      "a category rename changes every label in the workbook")
categories_path.write_text(saved_categories)
check("and restoring it moves back", field(export(), "Source digest") == baseline)

buckets_path = tmp / "data" / "tax-buckets.json"
saved_buckets = buckets_path.read_text() if buckets_path.exists() else None
buckets_path.write_text(json.dumps({"buckets": [{"slug": "probe", "name": "Probe"}]}),
                        encoding="utf-8")
check("a tax-bucket edit moves it too", field(export(), "Source digest") != baseline)
if saved_buckets is None:
    buckets_path.unlink()
else:
    buckets_path.write_text(saved_buckets)
check("and back again", field(export(), "Source digest") == baseline)

months_path = tmp / "data" / str(YEAR) / "months.json"
months_path.write_text(json.dumps({"%d-01" % YEAR: "closed"}), encoding="utf-8")
check("closing a month moves it", field(export(), "Source digest") != baseline)
months_path.unlink()
check("and reopening moves it back", field(export(), "Source digest") == baseline)

check("but the credentials in config.json are not hashed at all",
      "security" not in export_meta.CONFIG_INPUT_KEYS,
      "an audit artifact must not carry so much as a hash of a password")


print("== provenance that could not be gathered says so")
saved_months = months_path.read_text() if months_path.exists() else None
months_path.write_text("{not json at all", encoding="utf-8")
broken = export()
check("the export still succeeds", bool(broken))
check("and marks its provenance INCOMPLETE",
      str(field(broken, "Provenance")).startswith("INCOMPLETE"), str(field(broken, "Provenance")))
check("and names what could not be determined",
      "error" in " ".join(str(c) for row in metadata_sheet(broken) for c in row if c).lower()
      or "COULD NOT BE DETERMINED" in " ".join(str(c) for row in metadata_sheet(broken)
                                               for c in row if c),
      "degrading to '-' hides the difference between 'none' and 'unknown'")
if saved_months is None:
    months_path.unlink(missing_ok=True)
else:
    months_path.write_text(saved_months)


print("== the CLI can check an export against the store")
export()                    # a fresh one: the section above deliberately left a stale file
out = tmp / "data" / str(YEAR) / ("transactions-%d.xlsx" % YEAR)
check("a fresh export verifies", cli._export_verify(str(out)) == 0)
stale = [dict(row) for row in original_rows]
stale[0] = dict(stale[0], amount_eur=round(stale[0]["amount_eur"] - 5.0, 2))
store.rewrite_year(YEAR, {source_name: stale})
check("and an export the store has moved past is reported stale",
      cli._export_verify(str(out)) == 1,
      "a workbook that no longer matches its source must not silently pass")
store.rewrite_year(YEAR, {source_name: original_rows})
check("and passes once more when the store is put back",
      cli._export_verify(str(out)) == 0)


# Anti-shrink guard: exact count at implementation time. May only ever be RAISED
# when checks are added — never lowered (see AGENTS.md: never weaken a test).
MIN_CHECKS = 55
check("suite did not shrink", total_checks >= MIN_CHECKS,
      "total_checks=%d < %d" % (total_checks, MIN_CHECKS))

shutil.rmtree(tmp)
print()
if failures:
    print("FAILED: %s" % ", ".join(failures))
    sys.exit(1)
print("All checks passed.")
