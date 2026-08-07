"""End-to-end pipeline regression test on synthetic fixtures.

Runs in an isolated temp root (FA_ROOT), so it never touches real data.
Usage: .venv/bin/python tests/test_pipeline.py
"""
import csv
import json
import os
import shutil
import subprocess
import sys
import tempfile
from datetime import date
from pathlib import Path

from sandbox import FIXTURES, PROJECT, build_sandbox

tmp = Path(tempfile.mkdtemp(prefix="fa-test-"))
os.environ["FA_ROOT"] = str(tmp)
build_sandbox(tmp)

sys.path.insert(0, str(PROJECT))
from fastapi import HTTPException  # noqa: E402
from pipeline import anchors, closings, coverage, doctor, formats, fx, ingest, settle, store, transfers  # noqa: E402
from pipeline.util import cents, load_accounts, read_json, write_json  # noqa: E402
from app import server  # noqa: E402

failures = []
total_checks = 0


def check(name, cond, detail=""):
    global total_checks
    total_checks += 1
    print("  %s %s %s" % ("OK " if cond else "FAIL", name, detail if not cond else ""))
    if not cond:
        failures.append(name)


def tree_snapshot(path):
    return {str(item.relative_to(path)): (item.stat().st_mtime_ns, item.read_bytes())
            for item in sorted(path.rglob("*")) if item.is_file()}


# --- ingest ---
ingest.run(verbose=False)
txns = store.effective_year(2026)
check("12 transactions ingested", len(txns) == 12, "got %d" % len(txns))
check("duplicate same-day REWE kept twice", sum(1 for t in txns if "REWE" in (t["counterparty"] or "")) == 2)

# --- doctor starts clean, is read-only, and exits successfully through the CLI ---
clean_before = tree_snapshot(tmp / "data")
clean_doctor = doctor.run()
clean_after = tree_snapshot(tmp / "data")
check("doctor reports clean fixture", clean_doctor["findings"] == [], str(clean_doctor["findings"]))
check("doctor clean scan writes nothing", clean_before == clean_after)
clean_cli = subprocess.run([sys.executable, "-m", "pipeline.cli", "doctor", "2026"],
                           cwd=str(PROJECT), env=os.environ.copy(), capture_output=True, text=True)
check("doctor CLI exits zero on clean fixture", clean_cli.returncode == 0, clean_cli.stdout + clean_cli.stderr)

# --- statement coverage: effective account counts, active range and low-activity flags ---
statement_coverage = coverage.coverage(2026, today=date(2026, 6, 30))
coverage_accounts = {account["id"]: account for account in statement_coverage["accounts"]}
check("coverage preserves account order",
      [account["id"] for account in statement_coverage["accounts"]]
      == ["bank1-person1", "card1-person1", "bank2-person2", "joint-account",
          "cash-person1", "cash-person2"])
check("coverage fixture counts per month",
      coverage_accounts["bank1-person1"]["months"][5] == 5
      and coverage_accounts["bank2-person2"]["months"][5] == 6
      and coverage_accounts["cash-person1"]["months"][5] == 1
      and all(sum(coverage_accounts[account]["months"]) == 0
              for account in ("card1-person1", "joint-account", "cash-person2")))
check("coverage active range follows fixture month", statement_coverage["active_range"] == [6, 6])
check("cash defaults low activity while normal accounts do not",
      coverage_accounts["cash-person2"]["low_activity"] is True
      and coverage_accounts["card1-person1"]["low_activity"] is False)

coverage_raw = store.load_year_by_file(2026)
future_txn = dict(store.load_year_raw(2026)[0])
future_txn.update({"id": "coverage-future", "date": "2026-12-01",
                   "source": {"file": "coverage-future.csv", "format": "test"}})
store.append_transactions(2026, "coverage-future", [future_txn])
check("coverage current-year active range caps at today",
      coverage.coverage(2026, today=date(2026, 9, 15))["active_range"] == [6, 9])
(tmp / "data" / "2026" / "transactions" / "coverage-future.jsonl").unlink()
store.rewrite_year(2026, coverage_raw)

reassigned_txn = next(t for t in store.load_year_raw(2026) if t["account"] == "bank1-person1")
store.save_decisions(2026, {reassigned_txn["id"]: {"account": "joint-account"}})
reassigned = {account["id"]: account for account in coverage.coverage(2026)["accounts"]}
check("coverage counts decision-reassigned account",
      reassigned["bank1-person1"]["months"][5] == 4
      and reassigned["joint-account"]["months"][5] == 1)
store.save_decisions(2026, {})

# the bank account is now a resettable RAW correction (moved out of the main decision modal)
acct_txn = next(t for t in store.load_year_raw(2026) if t["account"] == "bank1-person1")
server.transaction_edit(server.TxnEdit(year=2026, id=acct_txn["id"], date=acct_txn["date"],
                                       counterparty=acct_txn["counterparty"] or "x",
                                       amount_eur=acct_txn["amount_eur"], account="joint-account"))
check("raw account edit moves the entry",
      next(t for t in store.effective_year(2026) if t["id"] == acct_txn["id"])["account"] == "joint-account"
      and next(t for t in store.load_year_raw(2026) if t["id"] == acct_txn["id"]).get("manual_edit"))
server.transaction_edit_reset(server.DecisionClear(year=2026, id=acct_txn["id"]))
check("raw account edit resets to imported account",
      next(t for t in store.effective_year(2026) if t["id"] == acct_txn["id"])["account"] == "bank1-person1")
try:
    server.transaction_edit(server.TxnEdit(year=2026, id=acct_txn["id"], date=acct_txn["date"],
                                           counterparty=acct_txn["counterparty"] or "x",
                                           amount_eur=acct_txn["amount_eur"], account="no-such-account"))
    check("raw account edit rejects unknown account", False)
except HTTPException as exc:
    check("raw account edit rejects unknown account", exc.status_code == 400)

server.account_update(server.AccountUpdate(id="cash-person2", low_activity=False))
cash_meta = next(account for account in server.meta()["accounts"] if account["id"] == "cash-person2")
cash_coverage = next(account for account in coverage.coverage(2026)["accounts"]
                     if account["id"] == "cash-person2")
check("low activity false round-trips and overrides cash default",
      cash_meta["low_activity"] is False and cash_coverage["low_activity"] is False)
server.account_update(server.AccountUpdate(id="cash-person2", low_activity=True))
check("low activity true round-trips through update",
      next(account for account in server.meta()["accounts"]
           if account["id"] == "cash-person2")["low_activity"] is True)

# --- idempotency: re-ingesting the same file adds nothing ---
shutil.copy(FIXTURES / "bank2-person2__2026-06.csv", tmp / "inbox" / "bank2-person2__2026-06.csv")
ingest.run(verbose=False)
check("re-ingest is a no-op", len(store.effective_year(2026)) == 12)

# --- internal transfers: the equalization pair ---
internal = [t for t in store.effective_year(2026) if t["kind"] == "internal-transfer"]
check("pair and marker transfers detected", len(internal) == 3, "got %d" % len(internal))
check("pair matching works without account IBANs", sum(t.get("transfer_reason", "").startswith("pair:") for t in internal) == 2)
check("marker matching works without account IBANs", sum(t.get("transfer_reason") == "marker" for t in internal) == 1)

# --- a transfer decision can restore a false positive and survives re-detection ---
original_by_file = store.load_year_by_file(2026)
pair_txn = next(t for t in store.load_year_raw(2026) if t.get("transfer_reason", "").startswith("pair:"))
transfer_decs = store.decisions(2026)
transfer_decs[pair_txn["id"]] = {"kind": "normal"}
store.save_decisions(2026, transfer_decs)
transfers.mark_internal(2026)
restored = next(t for t in store.effective_year(2026) if t["id"] == pair_txn["id"])
check("kind normal overrides detected transfer", restored["kind"] == "normal" and restored["sharing"] != "out-of-scope")
forced_transfer = store.rules_engine.effective(
    next(t for t in store.load_year_raw(2026) if t.get("kind") == "normal"),
    {"kind": "internal-transfer"}, store.rules_engine.load_rules(), owner="person1")
check("kind internal-transfer decision is out of scope",
      forced_transfer["kind"] == "internal-transfer" and forced_transfer["sharing"] == "out-of-scope")
store.save_decisions(2026, {})
store.rewrite_year(2026, original_by_file)

# --- rules: groceries auto-booked, Amazon forced to review ---
txns = store.effective_year(2026)
matched = {t["counterparty"]: t for t in txns}
check("REWE rule-matched", any(t["status"] == "rule-matched" and t["category"] == "core-living/groceries"
                               for t in txns if "REWE" in (t["counterparty"] or "")))
check("Amazon in review queue", any(t["status"] == "needs_review"
                                    for t in txns if "AMAZON" in (t["counterparty"] or "")))

# --- extraction safety: force_review must override even a matching merchant rule ---
forced = dict(next(t for t in store.load_year_raw(2026) if "REWE" in (t["counterparty"] or "")))
forced["force_review"] = True
check("forced extraction row ignores merchant rule",
      store.rules_engine.effective(forced, None, store.rules_engine.load_rules(), owner="person2")["status"] == "needs_review")

# --- "Send to review": a decision force_review overrides an otherwise-matching rule ---
rewe = next(t for t in store.load_year_raw(2026) if "REWE" in (t["counterparty"] or ""))
check("decision force_review overrides matching rule",
      store.rules_engine.effective(dict(rewe), {"force_review": True}, store.rules_engine.load_rules())["status"] == "needs_review")
server.decision(server.Decision(year=2026, id=rewe["id"], fields={"force_review": True}))
check("force_review decision persists to needs_review",
      next(t for t in store.effective_year(2026) if t["id"] == rewe["id"])["status"] == "needs_review")
try:
    server.decision(server.Decision(year=2026, id=rewe["id"], fields={"force_review": "yes"}))
    check("decision rejects non-bool force_review", False)
except HTTPException as exc:
    check("decision rejects non-bool force_review", exc.status_code == 400)
# send-to-review KEEPS the chosen category while still surfacing the item in the queue
server.decision(server.Decision(year=2026, id=rewe["id"],
                                fields={"force_review": True, "category": "core-living/groceries", "sharing": "shared"}))
kept = next(t for t in store.effective_year(2026) if t["id"] == rewe["id"])
check("send-to-review keeps category while in review",
      kept["status"] == "needs_review" and kept["category"] == "core-living/groceries")
# confirming (force_review:false) takes it back out of the queue, category intact
server.decision(server.Decision(year=2026, id=rewe["id"],
                                fields={"category": "core-living/groceries", "sharing": "shared", "force_review": False}))
confirmed = next(t for t in store.effective_year(2026) if t["id"] == rewe["id"])
check("confirming clears force_review",
      confirmed["status"] == "confirmed" and confirmed["category"] == "core-living/groceries")
store.save_decisions(2026, {})  # reset so later checks start from a clean decision set

# --- rule scoping: person rules only hit their accounts, and beat family rules ---
rules_path = tmp / "rules" / "merchant-rules.json"
rules_data = json.loads(rules_path.read_text())
rules_data["rules"].insert(0, {"id": "scope-test", "scope": "person2",
                               "match": {"field": "counterparty", "contains": "UNKNOWN VENDOR"},
                               "category": "sports/equipment", "sharing": "shared"})
rules_data["rules"].append({"id": "scope-test-2", "scope": "person2",
                            "match": {"field": "counterparty", "contains": "REWE"},
                            "category": "recreation/hobbies", "sharing": "shared"})
rules_path.write_text(json.dumps(rules_data))
scoped = store.effective_year(2026)
vendor_t = next(t for t in scoped if "UNKNOWN VENDOR" in (t["counterparty"] or ""))
check("person2 rule ignores person1 account", vendor_t["status"] == "needs_review")
rewe_t = next(t for t in scoped if "REWE" in (t["counterparty"] or ""))
check("person2 rule beats family rule on her account", rewe_t["category"] == "recreation/hobbies"
      and rewe_t["matched_rule"] == "scope-test-2")
rules_data["rules"] = [r for r in rules_data["rules"] if not r["id"].startswith("scope-test")]
rules_path.write_text(json.dumps(rules_data))

# --- decisions: categorize salaries, settlement goes income-proportional ---
decs = store.decisions(2026)
for t in txns:
    p = (t.get("purpose") or "")
    if "GEHALT" in p or "SALARY" in p:
        owner = "person2" if t["account"].endswith("person2") else "person1"
        decs[t["id"]] = {"category": "to-receive/salary", "sharing": "shared", "income_owner": owner}
store.save_decisions(2026, decs)
s = settle.settlement(2026)
check("ratio from actual salaries (60/40)", abs(s["ratio"]["person1"] - 0.6) < 1e-9, str(s["ratio"]))
check("total shared expenses 393.49", abs(s["total_shared_expenses"] - 393.49) < 0.005, str(s["total_shared_expenses"]))
check("transfer person1->person2 70.14", s["transfer"] == {"from": "person1", "to": "person2", "amount": 70.14}, str(s["transfer"]))

# --- out-of-scope income is invisible to summaries and settlement ratios ---
income_before_oos = settle.year_summary(2026)["income"]
person2_salary = next(t for t in txns if "GEHALT" in (t.get("purpose") or ""))
decs[person2_salary["id"]]["sharing"] = "out-of-scope"
store.save_decisions(2026, decs)
oos_settlement = settle.settlement(2026)
income_after_oos = settle.year_summary(2026)["income"]
check("out-of-scope income excluded from ratio", oos_settlement["ratio"] == {"person1": 1.0, "person2": 0.0}, str(oos_settlement["ratio"]))
check("out-of-scope income excluded from summary", abs(income_before_oos - income_after_oos - 3200.0) < 0.005, str(income_after_oos))
decs[person2_salary["id"]]["sharing"] = "shared"
store.save_decisions(2026, decs)

# --- splits must sum exactly ---
amazon = next(t for t in txns if "AMAZON" in (t["counterparty"] or ""))
decs[amazon["id"]] = {"splits": [
    {"amount": -40.0, "category": "living-upgrades/deco", "sharing": "shared"},
    {"amount": -49.0, "category": "recreation/hobbies", "sharing": "personal:person1"}]}
store.save_decisions(2026, decs)
summary = settle.year_summary(2026)
check("split lands in both categories", summary["by_category"].get("living-upgrades/deco") == -40.0
      and summary["by_category"].get("recreation/hobbies") == -49.0)
check("personal split excluded from shared", abs(settle.settlement(2026)["total_shared_expenses"] - (393.49 - 49.0)) < 0.005)

bad = dict(decs)
bad[amazon["id"]] = {"splits": [{"amount": -40.0, "category": "living-upgrades/deco", "sharing": "shared"}]}
store.save_decisions(2026, bad)
bad_effective = next(t for t in store.effective_year(2026) if t["id"] == amazon["id"])
check("bad stored split degrades to review", bad_effective["status"] == "needs_review"
      and bad_effective.get("error") == "invalid split" and not bad_effective.get("splits"))
check("bad stored split does not break year summary", isinstance(settle.year_summary(2026)["expenses"], float))
store.save_decisions(2026, decs)
before_bad_post = dict(store.decisions(2026))
try:
    server.decision(server.Decision(year=2026, id=amazon["id"], fields={"splits": [
        {"amount": -40.0, "category": "living-upgrades/deco", "sharing": "shared"}]}))
    check("bad split POST rejected", False)
except HTTPException as exc:
    check("bad split POST rejected", exc.status_code == 400, str(exc.detail))
check("bad split POST leaves decisions unchanged", store.decisions(2026) == before_bad_post)
bulk_before = store.decisions(2026)
try:
    server.decisions_bulk(server.DecisionsBulk(year=2026, items=[
        {"id": amazon["id"], "fields": {"note": "must not persist"}},
        {"id": vendor_t["id"], "fields": {"category": "missing/category"}},
    ]))
    check("invalid bulk decision rejected", False)
except HTTPException as exc:
    check("invalid bulk decision rejected", exc.status_code == 400)
check("bulk decisions are atomic", store.decisions(2026) == bulk_before)
for label, fields in (
        ("unknown field", {"surprise": True}),
        ("unknown category", {"category": "missing/category"}),
        ("unknown sharing owner", {"sharing": "personal:nobody"}),
        ("unknown income owner", {"income_owner": "nobody"}),
        ("unknown account", {"account": "missing-account"}),
        ("unknown kind", {"kind": "maybe-transfer"})):
    try:
        server.decision(server.Decision(year=2026, id=amazon["id"], fields=fields))
        check("decision rejects " + label, False)
    except HTTPException as exc:
        check("decision rejects " + label, exc.status_code == 400)
null_category = dict(decs)
null_category[amazon["id"]] = {"category": None, "sharing": "shared"}
store.save_decisions(2026, null_category)
null_effective = next(t for t in store.effective_year(2026) if t["id"] == amazon["id"])
check("null category remains in review", null_effective["status"] == "needs_review" and null_effective["category"] is None)
auto_category = dict(decs)
auto_category[amazon["id"]] = {"category": "auto:items", "sharing": "shared"}
store.save_decisions(2026, auto_category)
auto_effective = next(t for t in store.effective_year(2026) if t["id"] == amazon["id"])
check("manual auto-items category resolves by amount",
      auto_effective["category"] == "core-living/items-over-50" and auto_effective["status"] == "confirmed")
store.save_decisions(2026, decs)

# --- year_cost excluded monthly, included annually ---
vendor = next(t for t in store.effective_year(2026) if "UNKNOWN VENDOR" in (t["counterparty"] or ""))
decs[vendor["id"]] = {"category": "sports/equipment", "sharing": "shared", "year_cost": True}
store.save_decisions(2026, decs)
ys = settle.year_summary(2026)
june = ys["months"][5]
# personal splits still count as family expenses (only equalization ignores them)
check("year_cost excluded from month", june["year_costs_excluded"] == -120.0 and
      abs(june["expenses"] - (-393.49 + 120.0)) < 0.005, json.dumps(june))
check("year_cost included annually", abs(ys["expenses"] - -393.49) < 0.005, str(ys["expenses"]))
monthly_without_year_cost = settle.settlement(2026, 6)["total_shared_expenses"]
decs[vendor["id"]]["year_cost"] = False
store.save_decisions(2026, decs)
monthly_with_regular_cost = settle.settlement(2026, 6)["total_shared_expenses"]
check("year_cost excluded from monthly settlement",
      abs(monthly_with_regular_cost - monthly_without_year_cost - 120.0) < 0.005)
decs[vendor["id"]]["year_cost"] = True
store.save_decisions(2026, decs)

# --- tax workflow: category mappings are candidates until explicitly reviewed ---
tax_path = tmp / "rules" / "tax-buckets.json"
tax_data = json.loads(tax_path.read_text())
medical = next(b for b in tax_data["buckets"] if b["slug"] == "aussergewoehnliche-belastungen")
medical["category_map"] = ["health/doctors"]
tax_path.write_text(json.dumps(tax_data))
decs[vendor["id"]] = {"category": "health/doctors", "sharing": "shared"}
store.save_decisions(2026, decs)
candidate = next(t for t in store.effective_year(2026) if t["id"] == vendor["id"])
candidate_report = next(b for b in settle.tax_report(2026) if b["bucket"] == "aussergewoehnliche-belastungen")
check("mapped tax category is a candidate", candidate["tax_bucket_source"] == "category-map"
      and not candidate["tax_confirmed"] and candidate_report["candidate_count"] == 1)

decs[vendor["id"]].update({
    "tax_confirmed": True, "tax_owner": "person2",
    "attachments": [{"file": "2026/doctor.pdf", "description": "Invoice"}],
})
store.save_decisions(2026, decs)
confirmed_report = next(b for b in settle.tax_report(2026) if b["bucket"] == "aussergewoehnliche-belastungen")
confirmed_item = confirmed_report["owners"]["person2"]["items"][0]
check("confirmed tax item uses explicit claimant", confirmed_item["confirmed"]
      and confirmed_item["tax_owner"] == "person2" and confirmed_report["confirmed_count"] == 1)
check("tax evidence readiness is derived", confirmed_item["ready"]
      and confirmed_item["has_receipt"] and confirmed_item["payment_proof"])

# --- cash.csv is authoritative; duplicate deletion/editing is safe ---
cash_payload = dict(date="2026-08-10", account="cash-person1", amount=-10.0,
                    currency="EUR", description="Duplicate coffee", category="sports/equipment")
first_cash = server.cash(server.CashEntry(**cash_payload))
second_cash = server.cash(server.CashEntry(**cash_payload))
duplicate_cash = [t for t in store.effective_year(2026) if t.get("counterparty") == "Duplicate coffee"]
check("two identical cash entries added", len(duplicate_cash) == 2)
second_id = second_cash["id"]
server.decision(server.Decision(year=2026, id=second_id, fields={
    "category": "sports/equipment", "sharing": "shared", "note": "Keep this note",
    "attachments": [{"file": "2026/coffee.pdf", "description": "Receipt"}],
}))
server.cash_delete(server.CashDelete(year=2026, id=first_cash["id"]))
ingest.regenerate_cash()
duplicate_cash = [t for t in store.effective_year(2026) if t.get("counterparty") == "Duplicate coffee"]
check("deleting first duplicate stays deleted after ingest", len(duplicate_cash) == 1)
survivor = duplicate_cash[0]
check("surviving duplicate decision rekeyed", survivor.get("note") == "Keep this note"
      and len(survivor.get("attachments") or []) == 1)

edited = server.cash_edit(server.CashEdit(
    year=2026, id=survivor["id"], date="2026-08-11", account="cash-person1",
    amount=-11.0, currency="EUR", description="Edited coffee", category="sports/equipment"))
edited_txn = next(t for t in store.effective_year(2026) if t["id"] == edited["id"])
check("cash edit preserves note and attachments", edited_txn.get("note") == "Keep this note"
      and len(edited_txn.get("attachments") or []) == 1)
try:
    server.cash_edit(server.CashEdit(
        year=2026, id=edited["id"], date="2026-08-12", account="missing-account",
        amount=-12.0, currency="EUR", description="Invalid edit", category="sports/equipment"))
    check("invalid cash edit rejected", False)
except HTTPException as exc:
    check("invalid cash edit rejected", exc.status_code == 400)
check("invalid cash edit leaves original", any(t["id"] == edited["id"] for t in store.load_year_raw(2026)))

# The CLI ingest path must rebuild cash.jsonl from the edited CSV rather than
# incrementally appending a now-deleted duplicate back into canonical data.
cli_payload = dict(date="2026-08-20", account="cash-person1", amount=-7.0,
                   currency="EUR", description="CLI duplicate", category="sports/equipment")
server.cash(server.CashEntry(**cli_payload))
server.cash(server.CashEntry(**cli_payload))
cash_path = tmp / "inbox" / "cash.csv"
with open(cash_path, newline="", encoding="utf-8") as f:
    cash_rows = list(csv.reader(f))
for i, row in enumerate(cash_rows):
    if "CLI duplicate" in row:
        del cash_rows[i]
        break
with open(cash_path, "w", newline="", encoding="utf-8") as f:
    csv.writer(f).writerows(cash_rows)
ingest.run(verbose=False)
check("CLI ingest respects cash CSV deletion",
      sum(t.get("counterparty") == "CLI duplicate" for t in store.load_year_raw(2026)) == 1)

months = store.months_state(2026)
months["2026-08"] = "closed"
store.save_months_state(2026, months)
blocked = []
for operation in (
        lambda: server.cash(server.CashEntry(**cash_payload)),
        lambda: server.cash_delete(server.CashDelete(year=2026, id=edited["id"])),
        lambda: server.cash_edit(server.CashEdit(
            year=2026, id=edited["id"], date="2026-09-01", account="cash-person1",
            amount=-11.0, currency="EUR", description="Blocked edit", category="sports/equipment"))):
    try:
        operation()
        blocked.append(False)
    except HTTPException as exc:
        blocked.append(exc.status_code == 409)
check("closed month rejects cash add/edit/delete", all(blocked), str(blocked))
months["2026-08"] = "open"
store.save_months_state(2026, months)

# Effective account use blocks deletion; orphan references fail clearly.
decs_now = store.decisions(2026)
decs_now[edited["id"]]["account"] = "cash-person2"
store.save_decisions(2026, decs_now)
check("account usage counts decision reassignment", server.account_usage()["counts"].get("cash-person2", 0) >= 1)
decs_now[edited["id"]]["account"] = "missing-account"
store.save_decisions(2026, decs_now)
try:
    settle.settlement(2026)
    check("missing effective account fails clearly", False)
except ValueError as exc:
    check("missing effective account fails clearly", "missing-account" in str(exc))
decs_now[edited["id"]].pop("account")
store.save_decisions(2026, decs_now)

# Parsing must fail loudly on changed/bad date formats and sign debit/credit columns.
bad_dates = tmp / "bad-dates.csv"
bad_dates.write_text("date,account,amount,description\nnot-a-date,cash-person1,-5.00,Bad row\n")
try:
    formats.parse(bad_dates, formats.detect(bad_dates))
    check("all-bad dates rejected", False)
except ValueError as exc:
    check("all-bad dates rejected", "invalid dates" in str(exc))
small_export = tmp / "small-export.csv"
small_export.write_text(
    "Date,Amount,Name\n"
    "2026-01-01,-1.00,One\n"
    "2026-01-02,-2.00,Two\n"
    "2026-01-03,-3.00,Three\n"
    "Kontostand,100.00,Trailer\n")
small_cfg = {"signature": ["Date", "Amount"], "delimiter": ",", "decimal": "dot",
             "date_format": "%Y-%m-%d", "columns": {"date": "Date", "amount": "Amount",
             "counterparty": "Name"}}
small_rows, small_stats = formats.parse(small_export, small_cfg, with_stats=True)
check("small export accepts one trailer row", len(small_rows) == 3 and small_stats["skipped"] == 1)
debit_credit = tmp / "debit-credit.csv"
debit_credit.write_text("Date,Debit,Credit,Name\n2026-01-01,10.00,,Purchase\n2026-01-02,,5.00,Refund\n")
dc_cfg = {"signature": ["Date", "Debit", "Credit"], "delimiter": ",", "decimal": "dot",
          "date_format": "%Y-%m-%d", "columns": {"date": "Date", "amount_debit": "Debit",
          "amount_credit": "Credit", "counterparty": "Name"}}
dc_rows = formats.parse(debit_credit, dc_cfg)
check("debit and credit signs are deterministic", [row["amount"] for row in dc_rows] == [-10.0, 5.0])

# Deutsche Bank credit cards ship two columns both named 'Betrag': the foreign
# amount, then the EUR amount the bank already converted. Resolving them by name
# leaves it to duplicate-key resolution which one wins, and picking the foreign
# one books e.g. 27.97 BRL as 27.97 EUR. The config addresses both by position.
card_export = tmp / "db-card.csv"
card_export.write_text(
    "Kreditkartentransaktionen\n"
    "Kreditkarte;Kundennummer;Kartennummer;Karteninhaber\n"
    "Test Card;000 0000000;0000********0000;TEST HOLDER\n"
    "\n"
    "Abrechnungsdatum: 25.3.2026\n"
    "Belegdatum;Eingangstag;Verwendungszweck;Fremdwährung;Betrag;Kurs;Betrag;Währung\n"
    "1.3.2026;3.3.2026;LOCAL SHOP, 12345 CITY, DEU;EUR;-10,00;1,00000;-10,00;EUR\n"
    "2.3.2026;4.3.2026;FOREIGN SHOP, 00000000 CITY, BRA;BRL;-27,97;6,18805;-4,52;EUR\n"
    "Saldo:;;;;;;-14,52;EUR\n", encoding="utf-8")
card_cfg = formats.detect(card_export)
card_rows, card_stats = formats.parse(card_export, card_cfg, with_stats=True)
check("deutsche bank credit card format is detected", card_cfg["name"] == "deutsche-bank-kreditkarte")
check("credit card books the EUR column, never the foreign amount",
      [row["amount"] for row in card_rows] == [-10.0, -4.52]
      and {row["currency"] for row in card_rows} == {"EUR"})
check("credit card Saldo trailer is consumed, not counted as an invalid row",
      card_stats["skipped"] == 0)
# A statement that prints its own period total lets the bank check our arithmetic.
# It is not a running balance, so it cannot be an anchor, but it still catches a
# dropped or mis-read row before anything is written.
total_cfg = {"signature": ["Datum", "Betrag"], "delimiter": ";", "decimal": "comma",
             "date_format": "%d.%m.%Y", "columns": {"date": "Datum", "amount": 1},
             "statement_total": {"match": "Saldo", "amount_column": 1}}
good_total = tmp / "with-total.csv"
good_total.write_text("Datum;Betrag\n01.03.2026;-10,00\n02.03.2026;-15,50\nSaldo:;-25,50\n",
                      encoding="utf-8")
check("a statement whose rows match its stated total parses",
      [row["amount"] for row in formats.parse(good_total, total_cfg)] == [-10.0, -15.5])
short_total = tmp / "missing-row.csv"
short_total.write_text("Datum;Betrag\n01.03.2026;-10,00\nSaldo:;-25,50\n", encoding="utf-8")
try:
    formats.parse(short_total, total_cfg)
    check("a row missing against the stated total is refused", False)
except ValueError as exc:
    check("a row missing against the stated total is refused", "add up to" in str(exc))
no_total = tmp / "no-total.csv"
no_total.write_text("Datum;Betrag\n01.03.2026;-10,00\n02.03.2026;-15,50\n", encoding="utf-8")
check("a statement without a stated total is still accepted",
      len(formats.parse(no_total, total_cfg)) == 2)

# A format config whose decimal style contradicts the file is the defect that
# reached real books: '26,00' read as dot-decimal becomes 2600.00, and every total
# downstream stays perfectly self-consistent while being a hundred times wrong.
wrong_decimal = tmp / "wrong-decimal.csv"
wrong_decimal.write_text(
    "date,title,amount\n"
    "2026-03-01,One,\"26,00\"\n2026-03-02,Two,\"207,78\"\n2026-03-03,Three,\"1.234,56\"\n",
    encoding="utf-8")
dot_cfg = {"signature": ["date", "title", "amount"], "delimiter": ",", "decimal": "dot",
           "date_format": "%Y-%m-%d", "columns": {"date": "date", "amount": "amount",
                                                  "counterparty": "title"}}
try:
    formats.parse(wrong_decimal, dot_cfg)
    check("a comma-decimal file read as dot-decimal is refused", False)
except ValueError as exc:
    check("a comma-decimal file read as dot-decimal is refused", "decimal comma" in str(exc))
check("the same file parses once the config agrees",
      [row["amount"] for row in formats.parse(wrong_decimal, dict(dot_cfg, decimal="comma"))]
      == [26.0, 207.78, 1234.56])

dot_file = tmp / "dot-decimal.csv"
dot_file.write_text("date,title,amount\n2026-03-01,One,26.00\n2026-03-02,Two,207.78\n"
                    "2026-03-03,Three,1234.56\n", encoding="utf-8")
try:
    formats.parse(dot_file, dict(dot_cfg, decimal="comma"))
    check("a dot-decimal file read as comma-decimal is refused", False)
except ValueError as exc:
    check("a dot-decimal file read as comma-decimal is refused", "decimal point" in str(exc))
check("a matching dot-decimal config still parses",
      [row["amount"] for row in formats.parse(dot_file, dot_cfg)] == [26.0, 207.78, 1234.56])

# Whole-number amounts say nothing about decimal style and must not accuse a config.
ambiguous = tmp / "ambiguous.csv"
ambiguous.write_text("date,title,amount\n2026-03-01,One,26\n2026-03-02,Two,207\n"
                     "2026-03-03,Three,1234\n", encoding="utf-8")
check("amounts with no decimal separator are not judged",
      len(formats.parse(ambiguous, dot_cfg)) == 3
      and len(formats.parse(ambiguous, dict(dot_cfg, decimal="comma"))) == 3)

# Pair matching sees only two opposite amounts a few days apart, so an unrelated
# receipt and payment of the same size look exactly like a transfer between your own
# accounts. Marking both silently deleted a category a person had already assigned,
# and the money left every total without a trace.
pair_a = next(t for t in store.load_year_raw(2026)
              if t.get("kind") == "internal-transfer" and t["amount_eur"] > 0)
store.save_decisions(2026, {})
transfers.mark_internal(2026)
check("a paired transaction starts out excluded",
      next(t for t in store.effective_year(2026) if t["id"] == pair_a["id"])["kind"]
      == "internal-transfer")
store.save_decisions(2026, {pair_a["id"]: {"category": "to-receive/reimbursement",
                                           "sharing": "shared"}})
transfers.mark_internal(2026)
released = next(t for t in store.effective_year(2026) if t["id"] == pair_a["id"])
check("assigning a category releases it from the transfer pairing",
      released["kind"] == "normal" and not released.get("transfer_reason"))
check("and it is counted again",
      any(e["txn"]["id"] == pair_a["id"] for e in settle.entries(store.effective_year(2026))))

# The pair is the whole evidence for excluding either leg. Release one and the other
# is excluding money on the strength of a pairing that no longer exists — invisibly,
# because an excluded transaction shows up nowhere. Seven real savings-plan purchases
# sat like this after their partners turned out to be reimbursements.
partner_id = pair_a.get("transfer_partner")
check("a pair records which transaction it is paired with", bool(partner_id))
partner_now = next(t for t in store.effective_year(2026) if t["id"] == partner_id)
check("releasing one leg releases its orphaned partner",
      partner_now["kind"] == "normal" and not partner_now.get("transfer_reason"))

# Marking a transaction out of scope agrees it should not count, which is what a
# genuine transfer looks like — it must not be mistaken for a contradiction. Tested
# on a still-paired transaction, since releasing one leaves its partner unpaired.
store.save_decisions(2026, {})
transfers.mark_internal(2026)
pair_b = next(t for t in store.load_year_raw(2026)
              if t.get("kind") == "internal-transfer" and t["id"] != pair_a["id"])
store.save_decisions(2026, {pair_b["id"]: {"sharing": "out-of-scope"}})
transfers.mark_internal(2026)
check("sharing alone does not release a transfer",
      next(t for t in store.effective_year(2026) if t["id"] == pair_b["id"])["kind"]
      == "internal-transfer")

# An explicit kind decision still wins over everything, category included.
store.save_decisions(2026, {pair_b["id"]: {"category": "to-receive/reimbursement",
                                           "kind": "internal-transfer"}})
transfers.mark_internal(2026)
check("an explicit internal-transfer decision beats a category",
      next(t for t in store.effective_year(2026) if t["id"] == pair_b["id"])["kind"]
      == "internal-transfer")
store.save_decisions(2026, {})
transfers.mark_internal(2026)

# --- integrity checks that watch the import boundary, where every real bug this
# store has seen actually originated (a parser or config producing wrong rows,
# which the arithmetic then faithfully totalled).


def _checks_named(name, year=None):
    return [f for f in doctor.run(year)["findings"] if f["check"] == name]


# An account labelled with the wrong currency reads a manual balance or an anchor in
# that currency, turning R$1.500 into 1.500 EUR.
currency_doc = read_json(tmp / "data" / "accounts.json")
original_accounts = json.loads(json.dumps(currency_doc))
for account in currency_doc["accounts"]:
    if account["id"] == "bank1-person1":
        account["currency"] = "BRL"
write_json(tmp / "data" / "accounts.json", currency_doc)
check("an account whose currency contradicts its transactions is flagged",
      any("bank1-person1" in f["ids"] for f in _checks_named("account-currency")))
write_json(tmp / "data" / "accounts.json", original_accounts)
check("matching account currencies are not flagged", not _checks_named("account-currency"))

# An anchor keeps the currency it was written with, so changing an account's
# currency strands the anchors it already had. verify() then declines the span
# rather than comparing across currencies — correct, but silent: the account simply
# stops being verified and nothing says so.
anchors.add_manual("bank1-person1", "2026-05-31", 1000.0)
check("a fresh anchor matches its account and is not flagged",
      not _checks_named("anchor-currency"))
drifted_doc = read_json(tmp / "data" / "accounts.json")
for account in drifted_doc["accounts"]:
    if account["id"] == "bank1-person1":
        account["currency"] = "BRL"
write_json(tmp / "data" / "accounts.json", drifted_doc)
drift_findings = _checks_named("anchor-currency")
check("an anchor left behind by a currency change is flagged",
      any("bank1-person1" in f["ids"] for f in drift_findings),
      str([f["message"] for f in drift_findings]))
check("the message says the account is no longer being verified",
      any("no longer being verified" in f["message"] for f in drift_findings))
write_json(tmp / "data" / "accounts.json", original_accounts)
anchors.remove("bank1-person1", "2026-05-31")
check("clearing the stale anchor clears the finding", not _checks_named("anchor-currency"))

# A placeholder rate file converts every foreign amount wrongly and looks fine.
fx_backup = fx.CACHE.read_bytes() if fx.CACHE.exists() else None
fx._rates = None
fx.CACHE.parent.mkdir(parents=True, exist_ok=True)
brl_txn = dict(store.load_year_raw(2026)[0])
brl_txn.update({"id": "fx-probe", "currency": "BRL", "amount_original": -100.0, "fx_rate": 6.25,
                "source": {"file": "fx-probe.csv", "format": "test"}})
store.append_transactions(2026, "fx-probe", [brl_txn])
fx.CACHE.write_text("Date,USD,BRL\n" + "".join(
    "2026-01-%02d,1.1000,6.2500\n" % day for day in range(1, 29)), encoding="utf-8")
fx_findings = _checks_named("fx-cache")
check("a flat placeholder rate series is flagged",
      any("identical" in f["message"] for f in fx_findings), str([f["message"] for f in fx_findings]))
check("a two-currency rate file is flagged as a stand-in",
      any("only" in f["message"] and "currencies" in f["message"] for f in fx_findings))
fx.CACHE.unlink()
fx._rates = None
check("a missing rate cache is reported, not downloaded",
      any("no ECB rate cache" in f["message"] for f in _checks_named("fx-cache"))
      and not fx.CACHE.exists())
(tmp / "data" / "2026" / "transactions" / "fx-probe.jsonl").unlink()
if fx_backup is not None:
    fx.CACHE.write_bytes(fx_backup)
fx._rates = None

# The out-of-scope category is a label; only the sharing flag removes a transaction
# from the totals. Carrying the label alone leaves it counted, as an expense.
drift_txn = next(t for t in store.load_year_raw(2026) if not t.get("splits"))
store.save_decisions(2026, {drift_txn["id"]: {"category": "out-of-scope/out-of-scope",
                                              "sharing": "shared"}})
check("out-of-scope category without the flag is flagged",
      any(drift_txn["id"] in f["ids"] for f in _checks_named("out-of-scope-drift", 2026)))
store.save_decisions(2026, {drift_txn["id"]: {"category": "out-of-scope/out-of-scope",
                                              "sharing": "out-of-scope"}})
check("out-of-scope category with the flag is not flagged",
      not _checks_named("out-of-scope-drift", 2026))
store.save_decisions(2026, {})

# A merchant rule drifts exactly the same way, and a check that reads only decisions
# cannot see it. Fifteen real dividends were counted as expenses for this reason: the
# rule said out-of-scope category, shared sharing, and nothing in the app disagreed.
rules_path = tmp / "rules" / "merchant-rules.json"
rules_backup = read_json(rules_path)
rule_txn = next(t for t in store.load_year_raw(2026)
                if not t.get("splits") and t.get("kind") != "internal-transfer"
                and (t.get("counterparty") or "").strip())
write_json(rules_path, {"rules": rules_backup["rules"] + [
    {"id": "drift-probe", "scope": "family",
     "match": {"field": "counterparty", "contains": rule_txn["counterparty"]},
     "category": "out-of-scope/out-of-scope", "sharing": "shared"}]})
check("a merchant rule with the same drift is flagged too",
      any(rule_txn["id"] in f["ids"] for f in _checks_named("out-of-scope-drift", 2026)))
write_json(rules_path, rules_backup)
check("and the finding clears when the rule is corrected",
      not _checks_named("out-of-scope-drift", 2026))

# The same drift inside a split. Skipping split parents left this completely
# unpoliced: settle counts a part carrying the out-of-scope label without the flag
# exactly as it counts an unsplit one, so the expense figure moves and nothing says so.
split_txn = next(t for t in store.load_year_raw(2026) if t.get("kind") != "internal-transfer")
half = round(split_txn["amount_eur"] / 2, 2)
rest = round(split_txn["amount_eur"] - half, 2)
store.save_decisions(2026, {split_txn["id"]: {"splits": [
    {"amount": half, "category": "out-of-scope/out-of-scope", "sharing": "shared"},
    {"amount": rest, "category": "core-living/groceries", "sharing": "shared"}]}})
drift_effective = next(t for t in store.effective_year(2026) if t["id"] == split_txn["id"])
check("the drifted split part is genuinely counted",
      any(e["category"] == "out-of-scope/out-of-scope"
          for e in settle.entries([drift_effective])))
check("out-of-scope drift inside a split part is flagged",
      any(split_txn["id"] in f["ids"] for f in _checks_named("out-of-scope-drift", 2026)))
store.save_decisions(2026, {split_txn["id"]: {"splits": [
    {"amount": half, "category": "out-of-scope/out-of-scope", "sharing": "out-of-scope"},
    {"amount": rest, "category": "core-living/groceries", "sharing": "shared"}]}})
check("and it clears once that part carries the flag",
      not _checks_named("out-of-scope-drift", 2026))
store.save_decisions(2026, {})

# An excluded transaction appears in no total and no list, so a mark whose partner is
# gone can only be found by asking. Doctor is the only thing that asks.
orphan_files = store.load_year_by_file(2026)
orphan_txn = None
for rows in orphan_files.values():
    for row in rows:
        if row.get("kind") == "internal-transfer" and row.get("transfer_partner"):
            row["transfer_partner"] = "no-such-transaction#1"
            orphan_txn = row
            break
    if orphan_txn:
        break
store.rewrite_year(2026, orphan_files)
check("doctor reports a transfer mark whose partner is gone",
      any(orphan_txn["id"] in f["ids"] for f in _checks_named("orphan-transfer-mark", 2026)))
transfers.mark_internal(2026)
check("re-running detection releases it",
      not _checks_named("orphan-transfer-mark", 2026))

# Releasing a leg by hand takes effect the moment the decision is saved, while the
# stored row still says internal-transfer until detection next runs. Judging the pair
# by the stored value reports it as intact while one half is already being counted and
# the other is still excluded — the money is wrong in exactly one direction and
# nothing says so.
store.save_decisions(2026, {})
transfers.mark_internal(2026)
live = next(t for t in store.load_year_raw(2026)
            if t.get("kind") == "internal-transfer" and t.get("transfer_partner"))
store.save_decisions(2026, {live["transfer_partner"]: {"kind": "normal"}})
sides = {t["id"]: t["kind"] for t in store.effective_year(2026)
         if t["id"] in (live["id"], live["transfer_partner"])}
check("a decision alone leaves the stored pair untouched until reconciliation runs",
      sorted(sides.values()) == ["internal-transfer", "normal"])
check("which doctor reports, so the window is never silent",
      any(live["id"] in f["ids"] for f in _checks_named("orphan-transfer-mark", 2026)))
# Reconciliation is what closes it, and every production write path runs it.
transfers.mark_internal(2026)
sides = {t["id"]: t["kind"] for t in store.effective_year(2026)
         if t["id"] in (live["id"], live["transfer_partner"])}
check("reconciling releases both legs in one pass, never just one",
      sorted(sides.values()) == ["normal", "normal"])
check("so no orphan is left for doctor to report",
      not _checks_named("orphan-transfer-mark", 2026))
transfers.mark_internal(2026)
sides_again = {t["id"]: t["kind"] for t in store.effective_year(2026)
               if t["id"] in (live["id"], live["transfer_partner"])}
check("and a second pass changes nothing", sides_again == sides)
store.save_decisions(2026, {})
transfers.mark_internal(2026)

# Closing a month is meant to settle it, but the lock only rejects decisions. A
# merchant rule, transfer detection and a re-ingest all rewrite a closed month without
# asking and without leaving a trace — one rule edit moved a closed year's expenses by
# 63.34 EUR in production and nothing noticed. The figures are recorded at closing so a
# later change is visible. It is reported, not prevented: such a change is often a
# correction, and refusing those preserves a known error to protect a wrong figure.
closing_months = store.months_state(2026)
closing_txn = next(t for t in store.load_year_raw(2026) if t.get("kind") != "internal-transfer")
closing_month = int(closing_txn["date"][5:7])
closing_key = "2026-%02d" % closing_month
store.save_months_state(2026, {})
closings.save(2026, {})

server.close_month(server.MonthState(year=2026, month=closing_month, state="closed"))
check("closing a month records what it contained",
      closings.load(2026).get(closing_key, {}).get("transactions") is not None)
check("and nothing has drifted yet",
      not _checks_named("closed-month-drift", 2026))

# Move it through a merchant rule: the path a decisions-only view is blind to, and the
# one that actually moved a closed month in production.
closing_rules = read_json(tmp / "rules" / "merchant-rules.json")
write_json(tmp / "rules" / "merchant-rules.json", {"rules": closing_rules["rules"] + [
    {"id": "closing-drift-probe", "scope": "family",
     "match": {"field": "counterparty", "contains": closing_txn["counterparty"]},
     "category": "out-of-scope/out-of-scope", "sharing": "out-of-scope"}]})
store._EFFECTIVE_CACHE.clear()
check("a closed month whose figures moved is reported",
      any(closing_key in f["message"] for f in _checks_named("closed-month-drift", 2026)))
write_json(tmp / "rules" / "merchant-rules.json", closing_rules)
store._EFFECTIVE_CACHE.clear()
check("and the report clears once the figures match again",
      not _checks_named("closed-month-drift", 2026))

server.close_month(server.MonthState(year=2026, month=closing_month, state="open"))
check("reopening withdraws the baseline along with the claim",
      closing_key not in closings.load(2026))
store.save_months_state(2026, {closing_key: "closed"})
check("a closed month with no baseline is reported as unwatched, never as unchanged",
      bool(_checks_named("closed-month-unwatched", 2026)))
check("adopting a baseline starts watching it",
      closings.baseline(2026) == [closing_key]
      and not _checks_named("closed-month-unwatched", 2026)
      and not _checks_named("closed-month-drift", 2026))
store.save_months_state(2026, closing_months)
closings.save(2026, {})

# An upload that legitimately produced no rows writes no transaction file, so it must
# not be reported as a missing source — that finding could never be cleared.
stale_uploads = [
    {"id": "empty-upload", "source_stem": "bank1-person1__empty", "total": 0, "added": 0,
     "account": "bank1-person1", "kind": "table", "status": "processed"},
    {"id": "vanished-upload", "source_stem": "bank1-person1__vanished", "total": 7, "added": 7,
     "account": "bank1-person1", "kind": "table", "status": "processed"},
]
write_json(tmp / "data" / "uploads.json", {"uploads": stale_uploads})
stale_found = [f for f in doctor.run()["findings"] if f["check"] == "stale-upload-ref"]
check("an empty statement is not reported as a missing source",
      not any("empty-upload" in (f.get("ids") or []) for f in stale_found))
check("an upload whose rows really vanished is still reported",
      any("vanished-upload" in (f.get("ids") or []) for f in stale_found))
write_json(tmp / "data" / "uploads.json", {"uploads": []})

# A split part may be left uncategorized only by marking it out of scope. The split
# editor depends on this contract: without it there is no way to exclude one leg of a
# split, which is what drove an "out of scope" category into existence as a workaround.
split_host = next(t for t in store.effective_year(2026) if not t.get("splits"))
half = round(split_host["amount_eur"] / 2, 2)
server.decision(server.Decision(year=2026, id=split_host["id"], fields={"splits": [
    {"amount": half, "category": "core-living/groceries", "sharing": "shared"},
    {"amount": round(split_host["amount_eur"] - half, 2), "category": None, "sharing": "out-of-scope"},
], "category": "core-living/groceries", "sharing": "shared"}))
split_saved = next(t for t in store.effective_year(2026) if t["id"] == split_host["id"])
check("a split part can be out of scope with no category",
      any(p.get("sharing") == "out-of-scope" and not p.get("category") for p in split_saved["splits"]))
check("the out-of-scope split part stays outside the math",
      not any(r["in_expense_math"] for r in server._export_rows(2026)
              if r["transaction_id"] == split_host["id"] and r["sharing"] == "out-of-scope"))
try:
    server.decision(server.Decision(year=2026, id=split_host["id"], fields={"splits": [
        {"amount": half, "category": "core-living/groceries", "sharing": "shared"},
        {"amount": round(split_host["amount_eur"] - half, 2), "category": None, "sharing": "shared"},
    ], "category": "core-living/groceries", "sharing": "shared"}))
    check("a shared split part still needs a category", False)
except HTTPException as exc:
    check("a shared split part still needs a category", exc.status_code == 400)
server.decision_clear(server.DecisionClear(year=2026, id=split_host["id"]))

# The spreadsheet export must reproduce the app's own figures, or it is a
# plausible-looking file that quietly disagrees with every screen.
export_rows = server._export_rows(2026)
counted = [r for r in export_rows if r["in_expense_math"]]
export_lines = []
for txn in store.effective_year(2026):
    if txn.get("kind") == "internal-transfer":
        continue
    if txn.get("splits"):
        export_lines += [p["amount"] for p in txn["splits"]
                         if (p.get("sharing") or txn.get("sharing")) != "out-of-scope"]
    elif txn.get("sharing") != "out-of-scope":
        export_lines.append(txn["amount_eur"])
check("export counts the same lines as the app",
      len(counted) == len(export_lines),
      "export=%d app=%d" % (len(counted), len(export_lines)))
check("export sums to the same total as the app",
      cents(sum(r["amount_eur"] for r in counted)) == cents(sum(export_lines)))
check("export emits every transaction plus every split part",
      len(export_rows) == len(store.effective_year(2026))
      + sum(len(t.get("splits") or []) for t in store.effective_year(2026)))
check("export never counts a split parent and its parts twice",
      all(not (r["row_type"] == "transaction" and r["in_expense_math"]
               and any(o["transaction_id"] == r["transaction_id"] and o["row_type"] == "split-part"
                       for o in export_rows))
          for r in export_rows))
check("export flags internal transfers as outside the math",
      all(not r["in_expense_math"] for r in export_rows if r["kind"] == "internal-transfer"))
check("export flags out-of-scope rows as outside the math",
      all(not r["in_expense_math"] for r in export_rows if r["sharing"] == "out-of-scope"))

# A split part states only what differs from its parent. The workbook is what an
# external audit reconciles against, so resolving that inheritance differently from
# the totals makes the app disagree with its own audit trail: parts appeared
# uncategorised, not year-cost and untaxed while the app counted them as all three.
inherit_txn = next(t for t in store.load_year_raw(2026) if t.get("kind") != "internal-transfer")
inherit_half = round(inherit_txn["amount_eur"] / 2, 2)
inherit_rest = round(inherit_txn["amount_eur"] - inherit_half, 2)
inherit_bucket = read_json(tmp / "rules" / "tax-buckets.json")["buckets"][0]["slug"]
store.save_decisions(2026, {inherit_txn["id"]: {
    "category": "core-living/groceries", "year_cost": True, "tax_bucket": inherit_bucket,
    "splits": [{"amount": inherit_half, "sharing": "shared"},
               {"amount": inherit_rest, "sharing": "shared"}]}})
inherit_effective = next(t for t in store.effective_year(2026) if t["id"] == inherit_txn["id"])
inherit_entries = settle.entries([inherit_effective])
inherit_exported = [r for r in server._export_rows(2026)
                    if r["transaction_id"] == inherit_txn["id"] and r["row_type"] == "split-part"]
check("a split part inherits the parent's fields in the totals",
      all(e["category"] == "core-living/groceries" and e["year_cost"]
          and e["tax_bucket"] == inherit_bucket for e in inherit_entries))
check("and the workbook resolves that inheritance identically",
      len(inherit_exported) == len(inherit_entries)
      and all((r["category_slug"] or None) == e["category"]
              and bool(r["year_cost"]) == bool(e["year_cost"])
              and (r["tax_bucket"] or None) == (e["tax_bucket"] or None)
              for r, e in zip(inherit_exported, inherit_entries)))
store.save_decisions(2026, {})
export_rows = server._export_rows(2026)
# Build the actual workbook too: the row builder being right says nothing about the
# endpoint that writes the file.
export_response = server.transactions_export(2026)
export_path = Path(export_response.path)
check("export endpoint writes a workbook", export_path.is_file() and export_path.stat().st_size > 0)
from openpyxl import load_workbook  # noqa: E402
export_book = load_workbook(export_path)
export_sheet = export_book[export_book.sheetnames[0]]
check("workbook has a header row plus every export row",
      export_sheet.max_row == len(export_rows) + 1,
      "sheet=%d rows=%d" % (export_sheet.max_row, len(export_rows)))
check("workbook columns match the declared schema",
      [cell.value for cell in export_sheet[1]] == [name for name, _ in server.EXPORT_COLUMNS])
check("workbook ships the legend sheet", "Legend" in export_book.sheetnames)
check("workbook ships the summary sheet", "Summary" in export_book.sheetnames)

# The Summary sheet is only worth anything for an audit if its figures can be
# rebuilt from the rows beside them. Recompute its SUMIFS here and compare with
# what the dashboard reports.
export_summary = settle.year_summary(2026)


def _export_sumifs(income, month=None, exclude_year_costs=False):
    total = 0.0
    for row in export_rows:
        if not row["in_expense_math"] or row["is_income"] != income:
            continue
        if month is not None and row["month"] != month:
            continue
        if exclude_year_costs and row["year_cost"]:
            continue
        total += row["amount_eur"] or 0
    return total


check("export income is decided by category, not by sign",
      all(r["is_income"] == (r["category_slug"] in settle.income_categories())
          for r in export_rows if r["category_slug"]))
check("summary year income rebuilds from the rows",
      cents(_export_sumifs(True)) == cents(export_summary["income"]))
check("summary year expenses rebuild from the rows",
      cents(_export_sumifs(False)) == cents(export_summary["expenses"]))
check("summary months rebuild from the rows",
      all(cents(_export_sumifs(True, m["month"], True)) == cents(m["income"])
          and cents(_export_sumifs(False, m["month"], True)) == cents(m["expenses"])
          for m in export_summary["months"]))
check("summary carries a live formula next to each stored figure",
      any(isinstance(cell.value, str) and cell.value.startswith("=SUMIFS(")
          for row in export_book["Summary"].iter_rows() for cell in row))
export_path.unlink()

# Deleting an import must take everything it created and nothing else, and must
# not touch a closed month. A surviving anchor asserts a balance for data that is
# no longer in the store, which fails the next chain verification.
anchor_stmt = tmp / "anchor-stmt.csv"
anchor_stmt.write_text(
    "Buchungstag;Wertstellung;Verwendungszweck;Begünstigter / Auftraggeber;IBAN / Kontonummer;Betrag;Währung\n"
    "01.06.2026;01.06.2026;Deletable purchase;SHOP;DE00;-10,00;EUR\n"
    "Kontostand vom 30.06.2026;;;;;1.234,56;EUR\n", encoding="utf-8")
del_stem = "bank1-person1__delete-me"
del_stats = ingest.ingest_upload(anchor_stmt, "bank1-person1", del_stem, original_name="anchor-stmt.csv")
other_anchor_count = len([a for a in anchors.load(2026) if a.get("upload") != del_stem])
check("upload records an anchor tagged with its upload",
      del_stats["anchors"]["added"] == 1
      and any(a.get("upload") == del_stem for a in anchors.load(2026)))
del_view = ingest.upload_contents(del_stem)
check("upload contents are reported before deleting",
      del_view["transactions"] == 1 and del_view["closed_months"] == [])

# a closed month blocks the delete entirely
closed_state = store.months_state(2026)
closed_state["2026-06"] = "closed"
store.save_months_state(2026, closed_state)
check("upload contents flag the closed month", ingest.upload_contents(del_stem)["closed_months"] == ["2026-06"])
try:
    ingest.delete_upload(del_stem)
    check("delete refuses a closed month", False)
except ValueError as exc:
    check("delete refuses a closed month", "closed" in str(exc))
check("refused delete left the transactions in place",
      (tmp / "data" / "2026" / "transactions" / (del_stem + ".jsonl")).exists())
closed_state["2026-06"] = "open"
store.save_months_state(2026, closed_state)

ingest.delete_upload(del_stem)
check("delete removes the upload's transactions",
      not (tmp / "data" / "2026" / "transactions" / (del_stem + ".jsonl")).exists())
check("delete takes the upload's anchor with it",
      not any(a.get("upload") == del_stem for a in anchors.load(2026)))
check("delete leaves anchors it did not create",
      len([a for a in anchors.load(2026) if a.get("upload") != del_stem]) == other_anchor_count)

# Nubank cards export Brazilian amounts (1.234,56) with a comma delimiter, so the
# amounts are quoted. Reading them as dot-decimal multiplies a plain amount by 100
# and divides one carrying a thousands separator by about 1000 — both silent.
nubank_export = tmp / "nubank-card.csv"
nubank_export.write_text(
    "date,title,amount\n"
    "2026-03-01,Padaria,\"26,00\"\n"
    "2026-03-02,Mercado,\"1.234,56\"\n"
    "2026-03-03,Pagamento recebido,\"- 839,46\"\n", encoding="utf-8")
nubank_cfg = formats.detect(nubank_export)
nubank_rows = formats.parse(nubank_export, nubank_cfg)
check("nubank card format is detected", nubank_cfg["name"] == "nubank-card")
check("nubank amounts use Brazilian decimals",
      [row["amount"] for row in nubank_rows] == [-26.0, -1234.56, 839.46],
      str([row["amount"] for row in nubank_rows]))
check("nubank rows are booked in BRL",
      {row["currency"] for row in nubank_rows} == {"BRL"})

# A month with no card spending is a real statement, not a broken parse.
empty_card = tmp / "db-card-empty.csv"
empty_card.write_text(
    "Kreditkartentransaktionen\n"
    "Abrechnungsdatum: 25.4.2026\n"
    "Belegdatum;Eingangstag;Verwendungszweck;Fremdwährung;Betrag;Kurs;Betrag;Währung\n"
    "Saldo:;;;;;;0;EUR\n", encoding="utf-8")
check("a card month with no spending parses as zero transactions",
      formats.parse(empty_card, formats.detect(empty_card)) == [])
# The upload gate must accept that empty statement instead of calling it unsupported.
empty_preview = ingest.preview_file(empty_card)
check("upload preview accepts an empty statement",
      empty_preview["transactions"] == 0 and empty_preview["date_min"] is None
      and empty_preview["format"] == "deutsche-bank-kreditkarte")
try:
    ingest.preview_file(bad_dates)
    check("upload preview still rejects an unreadable file", False)
except ValueError:
    check("upload preview still rejects an unreadable file", True)
# The period line is the only thing dating a statement with no activity, and the
# two shapes below both occur in real files ('25.3.2026' and '2026-04').
check("statement period read from a day-first billing date",
      formats.parse(card_export, card_cfg, with_stats=True)[1]["period"] == "2026-03")
check("statement period read from a year-month billing date",
      formats.parse(empty_card, formats.detect(empty_card), with_stats=True)[1]["period"] == "2026-04")
# Coverage must stop calling a reported-but-empty month a missing statement.
empty_upload_stem = "card1-person1__empty-statement"
ingest.ingest_upload(empty_card, "card1-person1", empty_upload_stem, original_name="empty.csv")
write_json(tmp / "data" / "uploads.json", {"uploads": [
    {"account": "card1-person1", "source_stem": empty_upload_stem, "period": "2026-04", "total": 0}]})
april = next(a for a in coverage.coverage(2026, today=date(2026, 6, 30))["accounts"]
             if a["id"] == "card1-person1")
check("coverage marks a reported empty month",
      april["months"][3] == 0 and april["reported"][3] is True)
check("coverage leaves unreported empty months alone", april["reported"][4] is False)
write_json(tmp / "data" / "uploads.json", {"uploads": []})

# Administrative guards: unique safe rule ids, valid close states, non-zero ratios.
pattern = "Same / quoted ' pattern " + ("x" * 60)
rule_one = server.add_rule(server.RulePayload(pattern=pattern, category="sports/equipment"))["rule"]
rule_two = server.add_rule(server.RulePayload(pattern=pattern, category="sports/equipment"))["rule"]
check("rule ids are safe and unique", rule_one["id"] != rule_two["id"]
      and all(ch.isalnum() or ch == "-" for ch in rule_one["id"]))
store.rules_engine.remove_rule(rule_one["id"])
check("removing one rule leaves colliding peer", any(rule["id"] == rule_two["id"] for rule in store.rules_engine.load_rules()))
store.rules_engine.remove_rule(rule_two["id"])
try:
    server.close_month(server.MonthState(year=2026, month=6, state="colsed"))
    check("invalid month state rejected", False)
except HTTPException as exc:
    check("invalid month state rejected", exc.status_code == 400)
try:
    server.set_ratio_override(server.RatioOverride(year=2026, key="annual", ratio={"person1": 0, "person2": 0}))
    check("zero-sum ratio rejected", False)
except HTTPException as exc:
    check("zero-sum ratio rejected", exc.status_code == 400)

split_parent = {"id": "split-null", "amount_eur": -10.0, "sharing": "shared", "account": "cash-person1"}
split_entries = settle.entries([{**split_parent, "splits": [
    {"amount": -10.0, "category": "sports/equipment", "sharing": None}]}])
check("null split sharing consistently inherits parent", split_entries[0]["sharing"] == "shared")

# --- tolerant, conservative, and cross-year transfer matching ---
def transfer_txn(txn_id, account, day, eur, currency="EUR", counterparty=""):
    return {
        "id": txn_id, "account": account, "date": day, "amount_original": eur,
        "currency": currency, "amount_eur": eur, "fx_rate": None,
        "counterparty": counterparty, "purpose": "", "counterparty_iban": "",
        "force_review": False, "kind": "normal",
        "source": {"file": "transfer-tests.csv", "format": "test"},
    }


store.append_transactions(2027, "transfer-tests", [
    transfer_txn("fx-out", "bank1-person1", "2027-06-01", -500.00),
    transfer_txn("fx-in", "card1-person1", "2027-06-02", 499.87, "BRL"),
    transfer_txn("different-owner-out", "bank1-person1", "2027-09-01", -75.00),
    transfer_txn("different-owner-in", "bank2-person2", "2027-09-02", 75.00),
    transfer_txn("near-out", "bank1-person1", "2027-10-01", -100.00, counterparty="REWE Markt"),
    transfer_txn("near-in", "card1-person1", "2027-10-02", 91.00, "BRL"),
    transfer_txn("salary-far", "bank1-person1", "2027-11-01", 3000.00),
    transfer_txn("rewe-far", "card1-person1", "2027-11-02", -89.40, counterparty="REWE Markt"),
    transfer_txn("boundary-out", "bank1-person1", "2027-12-30", -200.00),
])
store.append_transactions(2028, "transfer-tests", [
    transfer_txn("boundary-in", "card1-person1", "2028-01-02", 200.00),
])
transfers.mark_internal(2027)
transfer_results = {t["id"]: t for y in (2027, 2028) for t in store.effective_year(y)}
check("FX-tolerant pair marked", all(transfer_results[x]["kind"] == "internal-transfer"
      and transfer_results[x]["transfer_reason"] == "pair:fx-tolerant" for x in ("fx-out", "fx-in")))
check("different-owner exact pair untouched", all(transfer_results[x]["kind"] == "normal"
      for x in ("different-owner-out", "different-owner-in")))
check("near-miss pair surfaced as a hint", all(transfer_results[x].get("possible_transfer")
      for x in ("near-out", "near-in")))
check("transfer hint does not suppress merchant rule",
      transfer_results["near-out"]["status"] == "rule-matched"
      and transfer_results["near-out"]["category"] == "core-living/groceries"
      and transfer_results["near-in"]["status"] == "needs_review")
check("far opposite-sign pair is not hinted",
      not transfer_results["salary-far"].get("possible_transfer")
      and not transfer_results["rewe-far"].get("possible_transfer")
      and transfer_results["rewe-far"]["status"] == "rule-matched")
check("cross-year pair marked in both years", all(transfer_results[x]["kind"] == "internal-transfer"
      for x in ("boundary-out", "boundary-in")))

# --- balance anchors: trailer capture, raw chain verification and conflicts ---
anchor_fixture = FIXTURES / "anchors" / "db-giro__2031-01.csv"
anchor_cfg = formats.detect(anchor_fixture)
anchor_rows, anchor_parse_stats = formats.parse(anchor_fixture, anchor_cfg, with_stats=True)
check("balance trailer captured without skipped row",
      len(anchor_rows) == 2 and anchor_parse_stats["skipped"] == 0
      and anchor_parse_stats["anchors"] == [{"date": "2031-01-31", "balance": 1125.0}])

fallback_csv = tmp / "balance-fallback.csv"
fallback_csv.write_text(
    "Buchungstag;Wert;Begünstigter / Auftraggeber;Verwendungszweck;IBAN / Kontonummer;Betrag;Währung\n"
    "05.01.2031;;Test employer;Opening adjustment;;100,00;EUR\n"
    "Kontostand 31.01.2031;1.125,00;;;;;EUR\n")
_, fallback_stats = formats.parse(fallback_csv, anchor_cfg, with_stats=True)
check("balance trailer falls back to one German amount cell",
      fallback_stats["anchors"] == [{"date": "2031-01-31", "balance": 1125.0}]
      and fallback_stats["skipped"] == 0)
unmatched_anchor_csv = tmp / "balance-fail-soft.csv"
unmatched_anchor_csv.write_text(
    "Buchungstag;Wert;Begünstigter / Auftraggeber;Verwendungszweck;IBAN / Kontonummer;Betrag;Währung\n"
    "05.01.2031;;Test employer;Opening adjustment;;100,00;EUR\n"
    "Kontostand without usable details;;;;;;EUR\n")
_, unmatched_anchor_stats = formats.parse(unmatched_anchor_csv, anchor_cfg, with_stats=True)
check("matched but unusable balance trailer fails soft",
      unmatched_anchor_stats["anchors"] == [] and unmatched_anchor_stats["skipped"] == 0)

server.anchor_add(server.AnchorAdd(account="bank1-person1", date="2030-12-31", balance=1000.0))
added, anchor_years, skipped, recorded = ingest._ingest_file(
    anchor_fixture, "bank1-person1", load_accounts()[0])
check("DB giro ingest records trailer anchor",
      added == 2 and anchor_years == [2031] and skipped == 0 and recorded["added"] == 1)
first_span = anchors.verify("bank1-person1", year=2031)
check("manual anchor participates in matching cross-year span",
      len(first_span) == 1 and first_span[0]["ok"] is True
      and first_span[0]["expected_cents"] == 12500 and first_span[0]["actual_cents"] == 12500)

_, _, _, rerecorded = ingest._ingest_file(anchor_fixture, "bank1-person1", load_accounts()[0])
check("re-ingest does not duplicate balance anchor",
      rerecorded["duplicates"] == 1 and len(anchors.load(2031)) == 1)
upload_anchor_stats = ingest.ingest_upload(
    anchor_fixture, "bank1-person1", "uploaded-anchor-retry", original_name="db-giro-upload.csv")
check("staged upload path records balance anchor idempotently",
      upload_anchor_stats["added"] == 0 and upload_anchor_stats["anchors"]["duplicates"] == 1
      and upload_anchor_stats["anchor_message"] == "balance anchor already recorded")

conflict = anchors.record("bank1-person1", [{"date": "2031-01-31", "balance": 1200.0}],
                          "contradicting-export.csv")
check("conflicting balance is recorded but does not overwrite anchor",
      conflict["conflicts"] == 1 and anchors.load(2031)[0]["balance"] == 1125.0
      and anchors.load_conflicts(2031)[0]["incoming_balance"] == 1200.0)

anchor_files = store.load_year_by_file(2031)
anchor_files["db-giro__2031-01.jsonl"] = anchor_files["db-giro__2031-01.jsonl"][:1]
store.rewrite_year(2031, anchor_files)
holed_span = anchors.verify("bank1-person1", year=2031)[0]
check("missing raw transaction causes exact balance mismatch",
      holed_span["ok"] is False and holed_span["actual_cents"] - holed_span["expected_cents"] == -2500)
check("coverage includes balance mismatch summary",
      coverage.coverage(2031)["anchors"]["bank1-person1"]["status"] == "mismatch"
      and "-2500 cents" in coverage.coverage(2031)["anchors"]["bank1-person1"]["detail"])

# Seed representative legacy corruption and prove doctor reports rather than mutates it.
doctor_raw = next(txn for txn in store.load_year_raw(2026)
                  if txn.get("kind") == "normal" and txn["date"].startswith("2026-06"))
doctor_decisions = store.decisions(2026)
doctor_decisions[doctor_raw["id"]] = {
    "category": "missing/category",
    "account": "missing-account",
    "sharing": "personal:ghost",
    "income_owner": "ghost",
    "tax_owner": "ghost",
    "splits": [{"amount": 0, "category": "missing/category", "sharing": "shared"}],
}
doctor_decisions["orphan-doctor-decision"] = {"category": "core-living/groceries"}
store.save_decisions(2026, doctor_decisions)
doctor_months = store.months_state(2026)
doctor_months["2026-06"] = "closed"
store.save_months_state(2026, doctor_months)
write_json(tmp / "rules" / "budgets.json", {"budgets": {"missing/category": 99}})
write_json(tmp / "data" / "uploads.json", {"uploads": [
    {"id": "stale-upload", "source_stem": "missing-source"}]})
duplicate_txn = dict(doctor_raw)
store.append_transactions(2026, "doctor-duplicate", [duplicate_txn])
doctor_marker = transfer_txn("doctor-unpaired-marker", "bank2-person2", "2026-06-29", -333.0)
doctor_marker.update({"kind": "internal-transfer", "transfer_reason": "marker"})
store.append_transactions(2026, "doctor-marker", [doctor_marker])
with open(tmp / "inbox" / "cash.csv", "a", newline="", encoding="utf-8") as cash_file:
    csv.writer(cash_file).writerow(["2026-10-01", "cash-person1", "-3.21", "EUR",
                                    "Doctor desync", "sports/equipment"])

broken_before = tree_snapshot(tmp / "data")
broken_doctor = doctor.run()
broken_after = tree_snapshot(tmp / "data")
broken_checks = {finding["check"] for finding in broken_doctor["findings"]}
check("doctor detects seeded legacy corruption",
      {"orphan-decision", "unknown-category", "split-sum", "unknown-account",
       "duplicate-id", "unknown-sharing", "unknown-owner", "anchor-conflict",
       "anchor-mismatch", "cash-desync", "review-in-closed-month", "unpaired-marker",
       "orphan-budget", "stale-upload-ref"}.issubset(broken_checks), str(broken_checks))
check("doctor broken scan writes nothing", broken_before == broken_after)
broken_cli = subprocess.run([sys.executable, "-m", "pipeline.cli", "doctor"],
                            cwd=str(PROJECT), env=os.environ.copy(), capture_output=True, text=True)
check("doctor CLI exits one when errors exist", broken_cli.returncode == 1,
      broken_cli.stdout + broken_cli.stderr)

# Revert the seeded legacy corruption so the invariants run on a clean store.
store.save_decisions(2026, decs)
(tmp / "data" / "2026" / "transactions" / "doctor-duplicate.jsonl").unlink()
(tmp / "data" / "2026" / "transactions" / "doctor-marker.jsonl").unlink()
(tmp / "rules" / "budgets.json").unlink()
(tmp / "data" / "uploads.json").unlink()

# --- invariants: domain properties over the ENTIRE store, not single examples ---
years_present = {int(p.name) for p in (tmp / "data").iterdir() if p.is_dir() and p.name.isdigit()}
valid_categories = {"%s/%s" % (c["slug"], s["slug"])
                    for c in json.loads((tmp / "rules" / "categories.json").read_text())["categories"]
                    for s in c.get("subs", [])}
valid_categories.add("auto:items")
valid_categories.add("health/doctors")  # tax test maps this without adding it to categories.json

for y in sorted(years_present):
    # 1. The three dashboard scopes partition 'all' exactly (per cents).
    sum_all = settle.year_summary(y, "all")
    sum_parts = [settle.year_summary(y, s) for s in ["shared"] + server._people()]
    check("invariant: scope partition %d income" % y,
          cents(sum_all["income"]) == sum(cents(p["income"]) for p in sum_parts))
    check("invariant: scope partition %d expenses" % y,
          cents(sum_all["expenses"]) == sum(cents(p["expenses"]) for p in sum_parts))
    for cat in sum_all["by_category"]:
        check("invariant: scope partition %d %s" % (y, cat),
              cents(sum_all["by_category"][cat]) == sum(cents(p["by_category"].get(cat, 0)) for p in sum_parts))

    # 2. Settlement balances sum to zero (and settlement never crashes).
    try:
        bal_sum = sum(cents(v) for v in settle.settlement(y)["balances"].values())
        check("invariant: settlement zero-sum %d" % y, abs(bal_sum) <= 1, "sum=%d" % bal_sum)
    except Exception as exc:  # noqa: BLE001
        check("invariant: settlement computes %d" % y, False, str(exc))

    # 3. Every stored split sums; 4. every effective category exists.
    for t in store.effective_year(y):
        if t.get("splits") and not t.get("error"):
            split_sum = sum(cents(sp["amount"]) for sp in t["splits"])
            if split_sum != cents(t["amount_eur"]):
                check("invariant: split integrity %s" % t["id"], False, "sum=%d" % split_sum)
        for cat in [t.get("category")] + [sp.get("category") for sp in (t.get("splits") or [])]:
            if cat and cat not in valid_categories:
                check("invariant: category exists %s" % t["id"], False, cat)
check("invariant: splits + categories scanned", True)

# 5. Idempotency, globally: ingest over an empty inbox must change nothing.
for item in (tmp / "inbox").iterdir():
    if item.is_file() and item.suffix == ".csv":
        item.unlink()
before_idempotency = tree_snapshot(tmp / "data")
ingest.run(verbose=False)
check("invariant: global idempotency over empty inbox",
      before_idempotency == tree_snapshot(tmp / "data"))

# Anti-shrink guard: exact count at implementation time. May only ever be RAISED
# when checks are added — never lowered (see AGENTS.md: never weaken a test).
MIN_CHECKS = 211
check("suite did not shrink", total_checks >= MIN_CHECKS,
      "total_checks=%d < %d" % (total_checks, MIN_CHECKS))

shutil.rmtree(tmp)
print()
if failures:
    print("FAILED: %s" % ", ".join(failures))
    sys.exit(1)
print("All checks passed.")
