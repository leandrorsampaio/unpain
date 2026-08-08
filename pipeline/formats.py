"""Config-driven statement parsing.

Every bank format is a JSON file in pipeline/formats/ describing delimiter,
encoding, date format, decimal style and column names. Adding a bank = adding
a config file, no code. Detection: the config whose 'signature' columns all
appear in one of the first 30 lines of the file wins.
"""
import csv
import io
import re
from datetime import datetime
from pathlib import Path

from .util import CURRENCY_JUNK as _CURRENCY_JUNK_TEXT, MAX_YEAR, MIN_YEAR, cents, parse_amount, read_json

FORMATS_DIR = Path(__file__).parent / "formats"


def load_formats():
    # A schema that describes manifests is not itself a manifest — loading one as a
    # format would give it a signature of None and a vote in every detection.
    return [read_json(p) for p in sorted(FORMATS_DIR.glob("*.json"))
            if not p.name.endswith(".schema.json")]


def _read_lines(path):
    """Return rows as lists of strings for csv/xlsx alike."""
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [["" if c is None else str(c) for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
        return rows, None
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw.decode(enc)
            return None, text
        except UnicodeDecodeError:
            continue
    raise ValueError("Cannot decode %s" % path)


def _rows_for(path, cfg):
    xrows, text = _read_lines(path)
    if xrows is not None:
        return xrows
    delim = cfg.get("delimiter", ";") if cfg else ";"
    return list(csv.reader(io.StringIO(text), delimiter=delim))


def detect(path):
    """The one format that matches this file — or an error naming why not.

    This used to return the first manifest that matched, in directory order. Two
    manifests can both match one file (one signature being a subset of another), and
    then the parser was chosen by filename: adding a format could silently change how
    an existing bank's statements were read, and nothing would say so. Every enabled
    signature is now evaluated, and anything other than exactly one match is refused.
    """
    seen = []
    matched = []
    for cfg in load_formats():
        rows = _rows_for(path, cfg)
        for row in rows[:30]:
            cells = [c.strip() for c in row]
            if all(sig in cells for sig in cfg["signature"]):
                matched.append(cfg)
                break
            if row and len(seen) < 30:
                seen.append(";".join(cells)[:200])
    if len(matched) == 1:
        return matched[0]
    if matched:
        raise ValueError(
            "%s matches %d formats at once (%s), so which parser reads it would depend on "
            "filename order. Narrow one of their signatures in pipeline/formats/ before "
            "importing this file."
            % (path, len(matched), ", ".join(sorted(cfg["name"] for cfg in matched))))
    raise ValueError(
        "No format matched %s.\nFirst lines seen:\n%s\n"
        "Add a config in pipeline/formats/ describing this layout." % (path, "\n".join(seen[:8]))
    )


def parse(path, cfg, with_stats=False):
    """Yield raw transaction dicts: date, amount, currency, counterparty, purpose, counterparty_iban, account(optional)."""
    rows = _rows_for(path, cfg)
    header_idx = None
    for i, row in enumerate(rows):
        cells = [c.strip() for c in row]
        if all(sig in cells for sig in cfg["signature"]):
            header_idx = i
            header = cells
            break
    if header_idx is None:
        raise ValueError("Header row vanished in %s" % path)
    col = {name: idx for idx, name in enumerate(header)}
    cmap = cfg["columns"]
    decimal = cfg.get("decimal", "comma")
    sign = cfg.get("sign", 1)
    out = []
    skipped_dates = []
    bad_amounts = []
    out_of_range_dates = []
    captured_anchors = []
    raw_amounts = []
    stated_total = None
    for row in rows[header_idx + 1:]:
        if len(row) < 2:
            continue

        def get(key):
            name = cmap.get(key)
            # An integer names a column by position. Needed when a header repeats a
            # name (Deutsche Bank credit cards ship two 'Betrag' columns: foreign
            # amount, then the EUR amount) — resolving those by name would depend on
            # which duplicate wins, and picking the wrong one books a foreign amount
            # as euros.
            if isinstance(name, int):
                return str(row[name]).strip() if 0 <= name < len(row) else ""
            if not name or name not in col or col[name] >= len(row):
                return ""
            return str(row[col[name]]).strip()

        rawdate = get("date")
        date = _parse_date(rawdate, cfg.get("date_format", "%d.%m.%Y"))
        if date is None:
            total = _statement_total(row, cfg)
            if total is not None:
                stated_total = total
            matched_balance, anchor = _balance_anchor(row, col, cfg)
            if anchor:
                captured_anchors.append(anchor)
            if matched_balance:
                continue
            if sum(1 for cell in row if str(cell).strip()) >= 2:
                skipped_dates.append(";".join(str(cell).strip() for cell in row)[:200])
            continue  # summary/preamble rows
        if not MIN_YEAR <= int(date[:4]) <= MAX_YEAR:
            out_of_range_dates.append(";".join(str(cell).strip() for cell in row)[:200])
            continue
        if "amount_debit" in cmap or "amount_credit" in cmap:
            written = [value for value in (get("amount_debit"), get("amount_credit")) if value]
            debit = parse_amount(get("amount_debit"), decimal) if get("amount_debit") else None
            credit = parse_amount(get("amount_credit"), decimal) if get("amount_credit") else None
            amount = -abs(debit) if debit is not None else (abs(credit) if credit is not None else None)
            raw_amounts.extend(written)
        else:
            written = [get("amount")] if get("amount") else []
            amount = parse_amount(get("amount"), decimal, sign)
            raw_amounts.extend(written)
        if amount is None and written:
            # The row carries a date and a written amount, so it is a transaction — but
            # the amount does not read as money. Dropping it left a statement that still
            # looked complete while a real line was missing from it, and the skipped-row
            # count could not see it either. An empty amount cell is a different thing
            # (an informational row) and keeps falling through below.
            bad_amounts.append(";".join(str(cell).strip() for cell in row)[:200])
            continue
        if amount is None or amount == 0:
            continue
        txn = {
            "date": date,
            "amount": amount,
            "currency": get("currency") or cfg.get("currency", "EUR"),
            "counterparty": get("counterparty"),
            "purpose": get("purpose"),
            "counterparty_iban": get("iban").replace(" ", ""),
            "force_review": get("force_review").lower() in ("1", "true", "yes"),
        }
        if cfg.get("account_column"):
            txn["account"] = get("account")
        out.append(txn)
    if bad_amounts:
        raise ValueError(
            "%d row%s carry a date but an amount that cannot be read as money, so importing "
            "this file would leave the statement incomplete. Nothing was imported. Samples: %s"
            % (len(bad_amounts), " " if len(bad_amounts) == 1 else "s each ",
               " | ".join(bad_amounts[:3])))
    if out_of_range_dates:
        raise ValueError(
            "%d row%s dated outside %d-%d, which no statement does — the date column is being "
            "read wrongly. Nothing was imported. Samples: %s"
            % (len(out_of_range_dates), " is" if len(out_of_range_dates) == 1 else "s are",
               MIN_YEAR, MAX_YEAR, " | ".join(out_of_range_dates[:3])))
    if skipped_dates and (not out or len(skipped_dates) > max(3, len(out) * .1)):
        raise ValueError("Too many rows have invalid dates (%d skipped, %d parsed). Samples: %s" %
                         (len(skipped_dates), len(out), " | ".join(skipped_dates[:3])))
    _assert_decimal_style(raw_amounts, decimal)
    if stated_total is not None:
        parsed_total = sum(cents(row["amount"]) for row in out)
        if parsed_total != stated_total:
            raise ValueError(
                "This statement says its total is %.2f but the rows read from it add up to %.2f "
                "(%d rows, off by %.2f). Something was mis-read or missed, so nothing was imported."
                % (stated_total / 100.0, parsed_total / 100.0, len(out),
                   (parsed_total - stated_total) / 100.0))
    stats = {"skipped": len(skipped_dates), "skipped_samples": skipped_dates[:3],
             "anchors": captured_anchors, "period": _statement_period(rows[:header_idx], cfg)}
    return (out, stats) if with_stats else out


# A trailing ",dd" is a decimal comma; a trailing ".dd" with no later comma is a
# decimal point. Values like "1234" or "1,234" are ambiguous and are ignored.
_COMMA_DECIMAL = re.compile(r"^[^,]*,\d{1,2}$")
_DOT_DECIMAL = re.compile(r"^[^.]*\.\d{1,2}$")


def _statement_total(row, cfg):
    """The total a statement states for itself, in cents, or None.

    A credit-card statement prints the sum of its own period rather than a running
    balance, so it cannot become a balance anchor — but it is still the bank
    asserting an arithmetic fact about the file, which is exactly what makes a
    mis-parse detectable.
    """
    total_cfg = cfg.get("statement_total") or {}
    match_text = str(total_cfg.get("match") or "")
    if not match_text:
        return None
    cells = [str(cell).strip() for cell in row]
    if not any(match_text.lower() in cell.lower() for cell in cells):
        return None
    index = total_cfg.get("amount_column")
    if not isinstance(index, int) or not 0 <= index < len(cells):
        return None
    # Apply the format's sign so the total is compared on the same footing as the
    # rows it is meant to check.
    value = parse_amount(cells[index], cfg.get("decimal", "comma"), cfg.get("sign", 1))
    return None if value is None else cents(value)


def _assert_decimal_style(raw_amounts, decimal):
    """Refuse a file whose decimal style contradicts the format config.

    This is not a heuristic. Reading '26,00' as dot-decimal yields 2600.00 and
    '1.234,56' yields 1.23 — every amount silently wrong by a factor of a hundred or
    a thousand, with nothing downstream able to notice because the numbers stay
    self-consistent. Only the shape of the source text reveals it, so it is checked
    here, before anything is written.
    """
    comma = dot = 0
    for raw in raw_amounts:
        text = _CURRENCY_JUNK_TEXT.sub("", str(raw)).strip()
        if _COMMA_DECIMAL.match(text):
            comma += 1
        elif _DOT_DECIMAL.match(text):
            dot += 1
    total = comma + dot
    if total < 3:
        return   # too little evidence to accuse a config of being wrong
    if decimal == "comma" and dot > comma * 4:
        raise ValueError(
            "This file writes amounts with a decimal point (%d of %d samples, e.g. %s) but the "
            "'%s' format expects a decimal comma. Reading it this way would misplace every "
            "amount." % (dot, total, _sample(raw_amounts, _DOT_DECIMAL), decimal))
    if decimal != "comma" and comma > dot * 4:
        raise ValueError(
            "This file writes amounts with a decimal comma (%d of %d samples, e.g. %s) but the "
            "format expects a decimal point. Reading it this way would multiply every plain "
            "amount by a hundred." % (comma, total, _sample(raw_amounts, _COMMA_DECIMAL)))


def _sample(raw_amounts, pattern):
    for raw in raw_amounts:
        text = _CURRENCY_JUNK_TEXT.sub("", str(raw)).strip()
        if pattern.match(text):
            return "'%s'" % text
    return "?"


def _statement_period(preamble_rows, cfg):
    """The year-month a statement covers, read from a configured preamble line.

    A statement for a period with no activity has no transactions to date it, so
    this line is the only evidence that the period was reported at all. Returns
    'YYYY-MM' or None; never raises, since it is supplementary information.
    """
    period_cfg = cfg.get("statement_period") or {}
    match_text = str(period_cfg.get("match") or "")
    if not match_text:
        return None
    for row in preamble_rows:
        joined = ";".join(str(cell).strip() for cell in row)
        if match_text.lower() not in joined.lower():
            continue
        month_first = re.search(r"(\d{4})-(\d{1,2})\b", joined)
        if month_first:
            return "%s-%02d" % (month_first.group(1), int(month_first.group(2)))
        day_first = re.search(r"\b(\d{1,2})\.(\d{1,2})\.(\d{4})\b", joined)
        if day_first:
            return "%s-%02d" % (day_first.group(3), int(day_first.group(2)))
    return None


def _parse_date(raw, fmt):
    raw = (raw or "").strip().split(" ")[0]
    if not raw:
        return None
    for f in (fmt, "%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, f).date().isoformat()
        except ValueError:
            continue
    return None


def _balance_anchor(row, columns, cfg):
    """Fail-soft extraction for an optional configured statement balance row."""
    balance_cfg = cfg.get("balance_row")
    if not balance_cfg:
        return False, None
    cells = [str(cell).strip() for cell in row]
    joined = ";".join(cells)
    match_text = str(balance_cfg.get("match") or "")
    if not match_text or not any(match_text.lower() in cell.lower() for cell in cells):
        return False, None
    try:
        date_match = re.search(balance_cfg.get("date_regex") or "", joined)
    except re.error:
        return True, None
    if not date_match:
        return True, None
    raw_date = date_match.group(1) if date_match.groups() else date_match.group(0)
    anchor_date = _parse_date(raw_date, cfg.get("date_format", "%d.%m.%Y"))
    if not anchor_date:
        return True, None

    amount_name = balance_cfg.get("amount_column")
    amount_index = columns.get(amount_name)
    amount_raw = cells[amount_index] if amount_index is not None and amount_index < len(cells) else ""
    if amount_raw:
        balance = parse_amount(amount_raw, cfg.get("decimal", "comma"))
        return True, ({"date": anchor_date, "balance": balance} if balance is not None else None)

    german_amount = re.compile(r"^[+-]?(?:\d{1,3}(?:\.\d{3})*|\d+)(?:,\d{1,2})?\s*(?:EUR|€)?$", re.I)
    candidates = [parse_amount(cell, "comma") for cell in cells if german_amount.match(cell)]
    candidates = [amount for amount in candidates if amount is not None]
    if len(candidates) != 1:
        return True, None
    return True, {"date": anchor_date, "balance": candidates[0]}
