"""CLI entry point: python -m pipeline.cli <command>"""
import argparse
import json
import sys

from . import closings, doctor, export_meta, format_lint, fx, ingest, settle, store
from .mutation_lock import mutation_lock
from .util import ConfigError, load_config


def _export_verify(path):
    """Read a workbook's own Metadata sheet and re-check what it claims.

    An export is a snapshot. This says whether the store still matches the snapshot it
    was taken from — which is the difference between "this spreadsheet is out of date"
    and "this spreadsheet is wrong", and nobody could tell them apart before.
    """
    from openpyxl import load_workbook
    # data_only=False: a formula must come back as the formula that was written. With
    # data_only=True openpyxl returns the cached *result*, which nothing here computes,
    # so every formula cell would read as empty and the digest could never match.
    book = load_workbook(path, data_only=False)
    if export_meta.METADATA_SHEET not in book.sheetnames:
        print("%s carries no Metadata sheet: it predates export provenance and cannot be "
              "verified." % path)
        return 1
    declared = {}
    for row in book[export_meta.METADATA_SHEET].iter_rows(values_only=True):
        if row and row[0] and len(row) > 1 and row[1] is not None:
            declared.setdefault(str(row[0]), str(row[1]))

    print("%s" % path)
    for label in ("Export type", "Reporting year", "Generated at", "App version",
                  "Schema version", "Provenance", "Row count", "Content digest",
                  "Source digest"):
        if label in declared:
            print("  %-16s %s" % (label, declared[label]))

    problems = []
    # Two independent questions, reported separately, because the answers mean different
    # things to whoever is holding the file. "Has this workbook been altered since it was
    # written" is about the file. "Does the store still hold what it was built from" is
    # about the app. Only checking the second is how an edited spreadsheet passed.
    stated_content = declared.get("Content digest", "")
    if not stated_content:
        problems.append("the workbook declares no content digest (exported by an older "
                        "version); its cells cannot be checked")
    else:
        actual = export_meta.content_digest(book)
        if actual == stated_content:
            print("  INTACT:   every cell still hashes to what the file declares.")
        else:
            problems.append("TAMPERED OR CORRUPT: the cells no longer match the workbook's "
                            "own digest (declared %s, computed %s)" % (stated_content, actual))

    year = declared.get("Reporting year")
    if not year or not year.isdigit():
        problems.append("the Metadata sheet does not name a reporting year, so it cannot "
                        "be compared against the store")
    else:
        current = export_meta.source_digest(int(year))
        stated_source = declared.get("Source digest", "")
        if stated_source == current:
            print("  CURRENT:  the store still holds exactly the data this was built from.")
        else:
            problems.append("STALE: the store has changed since this was exported "
                            "(exported from %s, store is now %s). The workbook is not "
                            "wrong; it describes an earlier state." % (stated_source, current))

    if str(declared.get("Provenance", "")).startswith("INCOMPLETE"):
        problems.append(declared["Provenance"])

    if problems:
        for problem in problems:
            print("  %s" % problem)
        return 1
    print("  VERIFIED: the workbook is intact and the store still agrees with it.")
    return 0


def main():
    p = argparse.ArgumentParser(prog="pipeline", description="FamilyAccountability pipeline")
    sub = p.add_subparsers(dest="cmd", required=True)
    sub.add_parser("ingest", help="process all files in inbox/")
    sp = sub.add_parser("status", help="review status per year")
    sp.add_argument("year", nargs="?", type=int)
    sp = sub.add_parser("summary", help="year summary JSON")
    sp.add_argument("year", type=int)
    sp = sub.add_parser("settle", help="settlement for a year (annual, binding)")
    sp.add_argument("year", type=int)
    sp.add_argument("--month", type=int, help="monthly estimate instead")
    sp = sub.add_parser("tax", help="tax evidence pack for a year")
    sp.add_argument("year", type=int)
    sub.add_parser("fx-update", help="refresh ECB rates cache")
    sub.add_parser("formats-lint", help="check every bank format manifest")
    sp = sub.add_parser("export-verify", help="check an exported workbook against the current store")
    sp.add_argument("path")
    sp = sub.add_parser("doctor", help="read-only data integrity check")
    sp.add_argument("year", nargs="?", type=int)
    sp = sub.add_parser("close-baseline",
                        help="adopt current figures as the baseline for already-closed months")
    sp.add_argument("year", nargs="?", type=int)
    args = p.parse_args()
    load_config()  # fail fast with a clear message if config.json is missing/invalid

    if args.cmd == "export-verify":
        return _export_verify(args.path)
    if args.cmd == "formats-lint":
        report = format_lint.lint()
        for problem in report["problems"]:
            print("  %s" % problem)
        if report["best_guess"]:
            print("  note: %s %s not been verified against a real statement."
                  % (", ".join(report["best_guess"]),
                     "has" if len(report["best_guess"]) == 1 else "have"))
        print("%d format manifest(s): %s" %
              (report["manifests"], "all valid" if report["ok"]
               else "%d problem(s)" % len(report["problems"])))
        return 0 if report["ok"] else 1
    if args.cmd == "ingest":
        print("Ingesting inbox/ ...")
        ingest.run()
    elif args.cmd == "status":
        years = [args.year] if args.year else store.years()
        for y in years:
            txns = store.effective_year(y)
            review = sum(1 for t in txns if t["status"] == "needs_review")
            internal = sum(1 for t in txns if t.get("sharing") == "out-of-scope")
            print("%d: %5d transactions | %4d need review | %4d out of scope" % (y, len(txns), review, internal))
    elif args.cmd == "summary":
        json.dump(settle.year_summary(args.year), sys.stdout, indent=2, ensure_ascii=False)
    elif args.cmd == "settle":
        json.dump(settle.settlement(args.year, args.month), sys.stdout, indent=2, ensure_ascii=False)
    elif args.cmd == "tax":
        json.dump(settle.tax_report(args.year), sys.stdout, indent=2, ensure_ascii=False)
    elif args.cmd == "fx-update":
        with mutation_lock():
            fx._download()
        print("ECB rates updated.")
    elif args.cmd == "close-baseline":
        with mutation_lock():
            years = [args.year] if args.year else store.years()
            for year in years:
                adopted = closings.baseline(year)
                print("%d: %d month(s) now watched%s" %
                      (year, len(adopted), (": " + ", ".join(adopted)) if adopted else ""))
        print("These figures are today's, not the figures at the time each month was "
              "closed — that evidence was never recorded. Changes are watched from now on.")
    elif args.cmd == "doctor":
        result = doctor.run(args.year)
        counts = {severity: sum(item["severity"] == severity for item in result["findings"])
                  for severity in ("error", "warning", "info")}
        print("Doctor: %d errors, %d warnings, %d info; %d years, %d transactions, %d decisions checked" %
              (counts["error"], counts["warning"], counts["info"],
               len(result["checked"]["years"]), result["checked"]["transactions"],
               result["checked"]["decisions"]))
        for severity in ("error", "warning", "info"):
            items = [item for item in result["findings"] if item["severity"] == severity]
            if not items:
                continue
            print("%s:" % severity.upper())
            for item in items:
                ids = " [%s]" % ", ".join(str(value) for value in item["ids"]) if item["ids"] else ""
                scope = str(item["year"]) if item["year"] else "all years"
                print("  %s %s: %s%s" % (item["check"], scope, item["message"], ids))
        return 1 if counts["error"] else 0
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except ConfigError as e:
        print("Configuration problem: %s" % e)
        raise SystemExit(2)
