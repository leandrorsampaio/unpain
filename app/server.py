"""Local review + dashboard server.

Run:  .venv/bin/uvicorn app.server:app --reload --port 8765
Then open http://localhost:8765 (reachable on the home network via the Mac's IP).
"""
import contextlib
import csv
import hashlib
import importlib
import io
import math
import os
import secrets
import tempfile
import shutil
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import re

import anyio
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from pipeline import anchors, balances, closings, coverage, doctor, extraction, fx, ingest, networth, overview, recurring, rules_engine, settle, store, transfers  # noqa: E402
from pipeline.util import DATA, ICON_NAME, INBOX, ROOT, RULES, cents, load_config, read_json, write_json, load_accounts  # noqa: E402

app = FastAPI(title="FamilyAccountability")
STATIC = Path(__file__).parent / "static"
# Bundled seed templates ship with the code, so resolve them from the repo dir, not the
# (possibly separate, or empty-on-first-run) data ROOT — mirrors SCRIPTS_DIR below.
EXAMPLES = Path(__file__).resolve().parent.parent / "examples"
FEEDBACK = ROOT / "feedback"
FEEDBACK_FILES = FEEDBACK / "files"
FEEDBACK_ENTRIES = FEEDBACK / "entries.json"


# ---- optional app lock (off by default) ----
# A casual privacy lock, NOT real security: the data files on disk are plain-text and
# readable regardless. It only gates the browser + API. Sessions live in memory, so a
# server restart re-locks; when auto_lock is on they also expire after inactivity.
import hashlib as _hashlib
import hmac as _hmac
import secrets as _secrets
import time as _time

_SESSIONS = {}                       # token -> last-activity epoch seconds (in-memory)
_SCRYPT = dict(n=2 ** 14, r=8, p=1, dklen=32)
_SESSION_COOKIE = "fa_session"
_PUBLIC_API = {"/api/meta", "/api/unlock"}   # reachable while locked


def _hash_password(pw):
    salt = _secrets.token_bytes(16)
    digest = _hashlib.scrypt(pw.encode("utf-8"), salt=salt, **_SCRYPT)
    return "scrypt$%s$%s" % (salt.hex(), digest.hex())


def _verify_password(pw, stored):
    try:
        algo, salt_hex, hash_hex = stored.split("$")
        if algo != "scrypt":
            return False
        digest = _hashlib.scrypt(pw.encode("utf-8"), salt=bytes.fromhex(salt_hex), **_SCRYPT)
        return _hmac.compare_digest(digest.hex(), hash_hex)
    except Exception:
        return False


def _security():
    """The security config block, or None when the lock is off / app unconfigured."""
    if not (ROOT / "config.json").exists():
        return None
    try:
        return load_config().get("security")
    except Exception:
        return None


def _session_valid(token, sec):
    if not token or token not in _SESSIONS:
        return False
    now = _time.time()
    if sec.get("auto_lock") and (now - _SESSIONS[token]) > sec.get("timeout_minutes", 5) * 60:
        _SESSIONS.pop(token, None)
        return False
    _SESSIONS[token] = now             # sliding window: any request counts as activity
    return True


def _access(request):
    """(lock_enabled, unlocked) for this request."""
    sec = _security()
    if not sec or not sec.get("password_hash"):
        return (False, True)
    return (True, _session_valid(request.cookies.get(_SESSION_COOKIE), sec))


def _new_session_response(payload):
    token = _secrets.token_urlsafe(32)
    _SESSIONS[token] = _time.time()
    resp = JSONResponse(payload)
    resp.set_cookie(_SESSION_COOKIE, token, httponly=True, samesite="lax", path="/")
    return resp


# Read-modify-write is the shape of almost every write here: read decisions.json, change
# one entry, write the whole document back. Two of those at once — two browser tabs, or the
# second device that LAN mode exists for — both read the old document, and the later write
# discards the earlier one's decision without a word. These handlers are sync, so FastAPI
# runs them in a threadpool and they really do overlap.
#
# For a local two-person app the complete fix is also the cheap one: one mutation at a time.
# Reads stay parallel. Holding the lock around the whole handler makes each endpoint's
# read-modify-write atomic without every endpoint having to remember to be.
_WRITE_LOCK = anyio.Lock()
_CONCURRENT_METHODS = {"GET", "HEAD", "OPTIONS"}


@app.middleware("http")
async def serialize_writes(request: Request, call_next):
    if request.method in _CONCURRENT_METHODS:
        return await call_next(request)
    async with _WRITE_LOCK:
        return await call_next(request)


@app.middleware("http")
async def lock_gate(request: Request, call_next):
    path = request.url.path
    if path.startswith("/api/") or path.startswith("/receipts/"):
        enabled, unlocked = _access(request)
        request.state.lock_enabled = enabled
        request.state.unlocked = unlocked
        if enabled and not unlocked and path not in _PUBLIC_API:
            return JSONResponse({"detail": "locked"}, status_code=401)
    return await call_next(request)


class UnlockRequest(BaseModel):
    password: str


@app.post("/api/unlock")
def unlock(req: UnlockRequest):
    sec = _security()
    if not sec or not sec.get("password_hash"):
        return {"ok": True}                        # not locked
    if not _verify_password(req.password, sec["password_hash"]):
        _time.sleep(0.4)                            # throttle brute force
        raise HTTPException(401, "Wrong password")
    return _new_session_response({"ok": True})


@app.post("/api/lock")
def lock(request: Request):
    token = request.cookies.get(_SESSION_COOKIE)
    if token:
        _SESSIONS.pop(token, None)
    resp = JSONResponse({"ok": True})
    resp.delete_cookie(_SESSION_COOKIE, path="/")
    return resp


class SetPasswordRequest(BaseModel):
    new_password: str
    current_password: str = ""


@app.post("/api/security/set-password")
def set_password(req: SetPasswordRequest):
    cfg = load_config()
    sec = cfg.get("security") or {}
    if sec.get("password_hash") and not _verify_password(req.current_password, sec["password_hash"]):
        raise HTTPException(403, "Current password is wrong")
    if len(req.new_password) < 4:
        raise HTTPException(400, "Password must be at least 4 characters")
    sec["password_hash"] = _hash_password(req.new_password)
    sec.setdefault("auto_lock", False)
    sec.setdefault("timeout_minutes", 5)
    cfg["security"] = sec
    write_json(ROOT / "config.json", cfg)
    return _new_session_response({"ok": True})       # keep the caller unlocked


class RemovePasswordRequest(BaseModel):
    current_password: str


@app.post("/api/security/remove-password")
def remove_password(req: RemovePasswordRequest):
    cfg = load_config()
    sec = cfg.get("security") or {}
    if not sec.get("password_hash"):
        return {"ok": True}
    if not _verify_password(req.current_password, sec["password_hash"]):
        raise HTTPException(403, "Current password is wrong")
    cfg.pop("security", None)
    write_json(ROOT / "config.json", cfg)
    _SESSIONS.clear()
    return {"ok": True}


class AutoLockRequest(BaseModel):
    auto_lock: bool
    timeout_minutes: int = 5


@app.post("/api/security/auto-lock")
def set_auto_lock(req: AutoLockRequest):
    cfg = load_config()
    sec = cfg.get("security")
    if not sec or not sec.get("password_hash"):
        raise HTTPException(400, "Set a password first")
    if not 1 <= req.timeout_minutes <= 1440:
        raise HTTPException(400, "timeout_minutes must be between 1 and 1440")
    sec["auto_lock"] = bool(req.auto_lock)
    sec["timeout_minutes"] = int(req.timeout_minutes)
    cfg["security"] = sec
    write_json(ROOT / "config.json", cfg)
    return {"ok": True}


# The app shell links its own CSS and JS with a ?v= query so browsers refetch them after a
# change. Hand-written version strings are a footgun: forget one and the user gets new JS with
# the previous stylesheet, which looks exactly like a broken page. Stamp them from the files
# themselves instead, so shipping a change is the only thing anyone has to remember.
# Everything the shell links that this repo edits. Third-party bundles pinned to a real release
# (chart.umd.js?v=4.4.9) are left alone — their version string is already the truth.
VERSIONED_ASSETS = ("app.css", "app.js", "i18n.js", "i18n/de.js",
                    "vendor/theme.css", "vendor/material.js", "favicon.svg")


def _asset_version(name):
    stat = (STATIC / name).stat()
    return hashlib.sha256(f"{stat.st_mtime_ns}:{stat.st_size}".encode()).hexdigest()[:10]


def _index_html():
    html = (STATIC / "index.html").read_text(encoding="utf-8")
    for name in VERSIONED_ASSETS:
        html = re.sub(r"/static/%s(\?v=[^\"']*)?" % re.escape(name),
                      "/static/%s?v=%s" % (name, _asset_version(name)), html)
    return html


@app.get("/")
def index():
    return HTMLResponse(_index_html(), headers={"Cache-Control": "no-cache"})


# ---- product feedback (deliberately separate from accountability data) ----
def _feedback_entries():
    return read_json(FEEDBACK_ENTRIES, default={"entries": []}).get("entries", [])


def _public_feedback(entry):
    item = {key: entry[key] for key in ("id", "title", "description", "created_at")}
    if entry.get("attachment"):
        item["attachment"] = {
            "name": entry["attachment"]["name"],
            "size": entry["attachment"]["size"],
            "url": "/api/feedback-file/%s" % entry["id"],
        }
    return item


def _feedback_attachment_path(attachment):
    path = (FEEDBACK_FILES / attachment["stored"]).resolve()
    try:
        path.relative_to(FEEDBACK_FILES.resolve())
    except ValueError:
        return None
    return path


@app.get("/api/feedback")
def feedback_list():
    entries = sorted(_feedback_entries(), key=lambda item: item.get("created_at", ""), reverse=True)
    return {"items": [_public_feedback(entry) for entry in entries]}


@app.post("/api/feedback")
async def feedback_add(title: str = Form(...), description: str = Form(...),
                       file: Optional[UploadFile] = File(None)):
    title = title.strip()
    description = description.strip()
    if not title or not description:
        raise HTTPException(400, "title and description are required")
    if len(title) > 200 or len(description) > 10000:
        raise HTTPException(400, "title or description is too long")

    entry_id = "%d-%s" % (int(time.time() * 1000), secrets.token_hex(3))
    entry = {
        "id": entry_id,
        "title": title,
        "description": description,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    if file and file.filename:
        raw = await file.read(20 * 1024 * 1024 + 1)
        if len(raw) > 20 * 1024 * 1024:
            raise HTTPException(413, "attachment must be 20 MB or smaller")
        original_name = Path(file.filename).name or "attachment"
        safe = re.sub(r"[^A-Za-z0-9._-]", "_", original_name) or "attachment"
        dest = FEEDBACK_FILES / entry_id
        dest.mkdir(parents=True, exist_ok=False)
        (dest / safe).write_bytes(raw)
        entry["attachment"] = {"name": original_name, "size": len(raw), "stored": "%s/%s" % (entry_id, safe)}

    entries = _feedback_entries()
    entries.append(entry)
    write_json(FEEDBACK_ENTRIES, {"entries": entries})
    return {"ok": True, "entry": _public_feedback(entry)}


@app.get("/api/feedback-file/{entry_id}")
def feedback_file(entry_id: str):
    entry = next((item for item in _feedback_entries() if item.get("id") == entry_id), None)
    attachment = entry and entry.get("attachment")
    if not attachment:
        raise HTTPException(404, "attachment not found")
    path = _feedback_attachment_path(attachment)
    if not path or not path.is_file():
        raise HTTPException(404, "attachment not found")
    return FileResponse(path, filename=attachment["name"])


class FeedbackDelete(BaseModel):
    id: str


@app.post("/api/feedback-delete")
def feedback_delete(item: FeedbackDelete):
    entries = _feedback_entries()
    entry = next((entry for entry in entries if entry.get("id") == item.id), None)
    if not entry:
        raise HTTPException(404, "feedback entry not found")
    attachment = entry.get("attachment")
    if attachment:
        path = _feedback_attachment_path(attachment)
        if path:
            path.unlink(missing_ok=True)
            try:
                path.parent.rmdir()
            except OSError:
                pass
    write_json(FEEDBACK_ENTRIES, {"entries": [entry for entry in entries if entry.get("id") != item.id]})
    return {"ok": True}


@app.get("/api/meta")
def meta(request: Request = None):   # request is None only for direct in-process calls (tests)
    if not is_configured():
        return {"setup_required": True}
    # When locked, reveal only what the lock modal needs (title + language) — no data.
    _st = getattr(request, "state", None)
    if getattr(_st, "lock_enabled", False) and not getattr(_st, "unlocked", True):
        cfg = load_config()
        return {"locked": True, "household_name": cfg.get("household_name") or "Family Accountability",
                "language": cfg.get("language") or "en"}
    accounts, accounts_meta = load_accounts()
    cfg = load_config()
    public_account_fields = ("id", "owner", "bank", "type", "currency", "label", "group", "low_activity")
    return {
        "categories": read_json(RULES / "categories.json")["categories"],
        "tax_buckets": read_json(RULES / "tax-buckets.json")["buckets"],
        "accounts": [{key: account[key] for key in public_account_fields if key in account}
                     for account in accounts.values()],
        "people": cfg["people"],
        "person_labels": cfg.get("person_labels", {}),
        "person_styles": cfg.get("person_styles", {}),
        "shared_style": cfg.get("shared_style", {}),
        "brand_style": cfg.get("brand_style", {}),
        "household_name": cfg.get("household_name") or "Family Accountability",
        "language": cfg.get("language") or "en",
        "root": str(ROOT),  # shown in Settings › Data so the user can copy the folders themselves
        "locked": False,
        "lock_enabled": bool((cfg.get("security") or {}).get("password_hash")),
        "auto_lock": bool((cfg.get("security") or {}).get("auto_lock")),
        "lock_timeout": int((cfg.get("security") or {}).get("timeout_minutes", 5)),
        "currencies": [c.upper() for c in cfg["currencies"]],  # load_config guarantees the key
        "years": store.years(),
    }


# ---- bank account management (labels, grouping, owner) ----
ACCOUNTS_FILE = DATA / "accounts.json"


def is_configured():
    return (ROOT / "config.json").exists() and (DATA / "accounts.json").exists()


def _people():
    return load_config()["people"]


# ---- first-run setup wizard ----
class SetupRequest(BaseModel):
    person1: str
    person2: str
    ratio_person1: int = 50          # reference (monthly-estimate) ratio in percent
    currencies: list[str] = ["EUR"]


def _person_slug(name):
    slug = re.sub(r"[^a-z0-9-]", "", name.strip().lower().replace(" ", "-"))
    if not slug or slug == "couple":
        raise HTTPException(400, "'%s' is not usable as a person name" % name)
    return slug


@app.post("/api/setup")
def setup(req: SetupRequest):
    if is_configured():
        raise HTTPException(409, "Already configured. Use the Settings tab to change values.")
    p1, p2 = _person_slug(req.person1), _person_slug(req.person2)
    if p1 == p2:
        raise HTTPException(400, "The two names must be different")
    if not 1 <= req.ratio_person1 <= 99:
        raise HTTPException(400, "ratio must be between 1 and 99 percent")
    currencies = sorted({c.upper() for c in req.currencies} | {"EUR"})
    config = {
        "_help": "Global settings. reference_ratio is only the monthly ESTIMATE; the binding settlement is recomputed after Dec 31 from actual ratio income (subcategories with ratio_income=true in categories.json). income_owner 'couple' counts 50/50 to each side.",
        "people": [p1, p2],
        "person_labels": {p1: req.person1.strip(), p2: req.person2.strip()},
        "reference_ratio": {p1: req.ratio_person1 / 100, p2: (100 - req.ratio_person1) / 100},
        "items_threshold_eur": 50,
        "transfer_match_window_days": 4,
        "transfer_match_tolerance_cents": 200,
        "base_currency": "EUR",
        "currencies": currencies,
    }
    seeds = ["categories.json", "merchant-rules.json", "tax-buckets.json", "recurring-overrides.json"]
    examples = EXAMPLES
    for name in seeds:
        if not (examples / name).exists():
            raise HTTPException(500, "missing seed file examples/%s" % name)
    RULES.mkdir(parents=True, exist_ok=True)
    for name in seeds:
        if not (RULES / name).exists():                       # never clobber a partial setup
            write_json(RULES / name, read_json(examples / name))
    for d in (DATA, INBOX / "processed", ROOT / "receipts"):
        d.mkdir(parents=True, exist_ok=True)
    write_json(DATA / "accounts.json", {
        "_help": "Register every account. owner must match config.json people, or 'couple' for joint accounts.",
        "owner_names": [req.person1.strip(), req.person2.strip()],
        "accounts": [],
    })
    write_json(ROOT / "config.json", config)                  # config LAST: it completes is_configured()
    return {"ok": True}


# ---- settings (config.json knobs; people slugs stay immutable) ----
# UI languages shipped in app/static/i18n/. Keep in sync with I18N.names there.
SUPPORTED_LANGUAGES = ("en", "de")


class SettingsUpdate(BaseModel):
    person_labels: dict[str, str]
    reference_ratio: dict[str, float]
    items_threshold_eur: float
    transfer_match_window_days: int
    transfer_match_tolerance_cents: int
    currencies: list[str]
    household_name: str = ""          # blank clears the key (footer falls back to default)
    language: str = "en"              # UI language; "en" clears the key (default)
    person_styles: dict = {}          # { slug: {icon?, color?} } custom partner icon/colour
    shared_style: dict = {}           # {icon?, color?} for the shared / together / both option
    brand_style: dict = {}            # {icon?, color?} for the app-bar (header) icon


@app.get("/api/settings")
def settings_get():
    return load_config()


def _icon_name(value, where):
    """An icon is a Material Symbol name, checked in one place for every field that holds one.

    Category, partner, shared and app-bar icons all end up between <md-icon> tags in the
    page. Only the app-bar field used to check what it was storing; the rest accepted any
    string, which made a settings field a way to put markup into every screen that renders
    that icon."""
    name = str(value or "").strip()
    if not ICON_NAME.match(name):
        raise HTTPException(400, "%s must be a Material Symbol name (lowercase letters, digits "
                                 "and underscores)" % where)
    return name


@app.post("/api/settings-update")
def settings_update(s: SettingsUpdate):
    cfg = load_config()
    people = cfg["people"]
    if set(s.reference_ratio) != set(people):
        raise HTTPException(400, "reference_ratio must have exactly the keys %s" % people)
    # Summing to 1 is not enough on its own: {-0.2, 1.2} sums to 1 and then hands one
    # person a negative share of every shared cost. A share is a fraction of a whole.
    if any(not math.isfinite(v) or not 0 <= v <= 1 for v in s.reference_ratio.values()):
        raise HTTPException(400, "each reference_ratio share must be between 0% and 100%")
    if cents(sum(s.reference_ratio.values())) != 100:          # ratio stored as fractions summing to 1
        raise HTTPException(400, "reference_ratio must sum to 100%")
    if not set(s.person_labels) <= set(people) or any(not v.strip() for v in s.person_labels.values()):
        raise HTTPException(400, "person_labels keys must be %s and values non-empty" % people)
    if s.items_threshold_eur <= 0:
        raise HTTPException(400, "items threshold must be positive")
    if s.transfer_match_window_days < 0 or s.transfer_match_tolerance_cents < 0:
        raise HTTPException(400, "transfer matching values must be >= 0")
    currencies = sorted({c.upper() for c in s.currencies} | {"EUR"})
    cfg.update({
        "person_labels": {p: v.strip() for p, v in s.person_labels.items()},
        "reference_ratio": s.reference_ratio,
        "items_threshold_eur": s.items_threshold_eur,
        "transfer_match_window_days": s.transfer_match_window_days,
        "transfer_match_tolerance_cents": s.transfer_match_tolerance_cents,
        "currencies": currencies,
    })
    household = (s.household_name or "").strip()
    if household:
        cfg["household_name"] = household     # stored stripped; load_config validates it
    else:
        cfg.pop("household_name", None)        # blank clears the key -> footer falls back to default
    language = (s.language or "en").strip().lower()
    if language not in SUPPORTED_LANGUAGES:
        raise HTTPException(400, "language must be one of %s" % (SUPPORTED_LANGUAGES,))
    if language != "en":
        cfg["language"] = language            # only non-default is stored (mirrors household_name)
    else:
        cfg.pop("language", None)
    styles = {}
    for slug, st in (s.person_styles or {}).items():
        if slug not in people or not isinstance(st, dict):
            raise HTTPException(400, "person_styles keys must be %s" % people)
        clean = {}
        if st.get("color"):
            if not re.match(r"^#[0-9a-fA-F]{6}$", str(st["color"])):
                raise HTTPException(400, "person_styles color must be a #rrggbb hex")
            clean["color"] = st["color"]
        if st.get("icon"):
            clean["icon"] = _icon_name(st["icon"], "person_styles")
        if clean:
            styles[slug] = clean
    if styles:
        cfg["person_styles"] = styles
    else:
        cfg.pop("person_styles", None)
    shared = {}
    if s.shared_style:
        if s.shared_style.get("color"):
            if not re.match(r"^#[0-9a-fA-F]{6}$", str(s.shared_style["color"])):
                raise HTTPException(400, "shared_style color must be a #rrggbb hex")
            shared["color"] = s.shared_style["color"]
        if s.shared_style.get("icon"):
            shared["icon"] = _icon_name(s.shared_style["icon"], "shared_style")
    if shared:
        cfg["shared_style"] = shared
    else:
        cfg.pop("shared_style", None)
    brand = {}
    if s.brand_style:
        if s.brand_style.get("color"):
            if not re.match(r"^#[0-9a-fA-F]{6}$", str(s.brand_style["color"])):
                raise HTTPException(400, "brand_style color must be a #rrggbb hex")
            brand["color"] = s.brand_style["color"]
        if s.brand_style.get("icon"):
            brand["icon"] = _icon_name(s.brand_style["icon"], "brand_style")
    if brand:
        cfg["brand_style"] = brand
    else:
        cfg.pop("brand_style", None)
    write_json(ROOT / "config.json", cfg)
    return {"ok": True}


@app.get("/api/account-usage")
def account_usage():
    counts = {}
    for y in store.years():
        for t in store.effective_year(y):
            counts[t["account"]] = counts.get(t["account"], 0) + 1
    # inbox/cash.csv is a source of truth that lives OUTSIDE the year folders, so it survives a
    # year deletion. Count rows whose account has no materialized presence — otherwise a cash
    # account looks unused after its years are gone, gets deleted, and the next cash add crashes
    # regeneration on the orphaned rows. (Materialized accounts already counted above, incl.
    # decision-reassigned ones, so this never double-counts a live cash account.)
    materialized = set(counts)
    for row in _cash_rows()[1]:
        acct = row[1] if len(row) >= 2 else ""
        if acct and acct not in materialized:
            counts[acct] = counts.get(acct, 0) + 1
    return {"counts": counts}


class AccountAdd(BaseModel):
    id: str
    owner: str
    bank: str = ""
    type: str = ""
    currency: str = "EUR"
    label: Optional[str] = None
    group: Optional[str] = None
    low_activity: Optional[bool] = None


# Canonical account-type slugs (stored on the account; the UI shows a translated label per
# slug). Machine values stay English. 'cash' is special (manual entries require it). Keep in
# sync with ACCOUNT_TYPES / accountTypeLabel in app/static/app.js.
ACCOUNT_TYPES = ("giro", "savings", "credit-card", "cash", "brokerage", "other")


@app.post("/api/account-add")
def account_add(a: AccountAdd):
    if not re.match(r"^[a-z0-9][a-z0-9-]*$", a.id or ""):
        raise HTTPException(400, "id must be lowercase letters/numbers/dashes")
    if a.owner not in _people() + ["couple"]:
        raise HTTPException(400, "owner must be a person or 'couple'")
    if a.type and a.type not in ACCOUNT_TYPES:
        raise HTTPException(400, "type must be one of %s" % ", ".join(ACCOUNT_TYPES))
    doc = read_json(ACCOUNTS_FILE)
    if any(x["id"] == a.id for x in doc["accounts"]):
        raise HTTPException(409, "account id already exists")
    entry = {"id": a.id, "owner": a.owner, "bank": a.bank, "type": a.type, "currency": a.currency}
    if a.label:
        entry["label"] = a.label
    if a.group:
        entry["group"] = a.group
    if a.low_activity is not None:
        entry["low_activity"] = a.low_activity
    doc["accounts"].append(entry)
    write_json(ACCOUNTS_FILE, doc)
    return {"ok": True}


class AccountUpdate(BaseModel):
    id: str
    label: Optional[str] = None
    group: Optional[str] = None
    owner: Optional[str] = None
    bank: Optional[str] = None
    type: Optional[str] = None
    currency: Optional[str] = None
    low_activity: Optional[bool] = None


@app.post("/api/account-update")
def account_update(a: AccountUpdate):
    doc = read_json(ACCOUNTS_FILE)
    acct = next((x for x in doc["accounts"] if x["id"] == a.id), None)
    if not acct:
        raise HTTPException(404, "unknown account")
    if a.owner is not None:
        if a.owner not in _people() + ["couple"]:
            raise HTTPException(400, "owner must be a person or 'couple'")
        acct["owner"] = a.owner
    if a.type is not None and a.type and a.type not in ACCOUNT_TYPES:
        raise HTTPException(400, "type must be one of %s" % ", ".join(ACCOUNT_TYPES))
    for f in ("bank", "type", "currency"):
        v = getattr(a, f)
        if v is not None:
            acct[f] = v
    if a.low_activity is not None:
        acct["low_activity"] = a.low_activity
    for f in ("label", "group"):  # empty string clears the field
        v = getattr(a, f)
        if v is not None:
            if v:
                acct[f] = v
            else:
                acct.pop(f, None)
    write_json(ACCOUNTS_FILE, doc)
    return {"ok": True}


class AccountDelete(BaseModel):
    id: str


@app.post("/api/account-delete")
def account_delete(a: AccountDelete):
    n = account_usage()["counts"].get(a.id, 0)
    if n:
        raise HTTPException(409, "account has %d transactions — reassign or keep it" % n)
    doc = read_json(ACCOUNTS_FILE)
    before = len(doc["accounts"])
    doc["accounts"] = [x for x in doc["accounts"] if x["id"] != a.id]
    if len(doc["accounts"]) == before:
        raise HTTPException(404, "unknown account")
    write_json(ACCOUNTS_FILE, doc)
    return {"ok": True}


def _check_scope(scope):
    """Dashboard perspective: 'all' | 'shared' | one of the people."""
    if scope not in ["all", "shared"] + _people():
        raise HTTPException(400, "scope must be 'all', 'shared', or one of %s" % _people())
    return scope


@app.get("/api/summary")
def summary(year: int, scope: str = "all"):
    s = settle.year_summary(year, scope=_check_scope(scope))
    s["months_state"] = store.months_state(year)
    # A period whose figures have moved since it was settled is no longer settled, and
    # saying so belongs where a person looks at it — not only in an integrity check
    # they have to remember to run. Detecting drift and never mentioning it is most of
    # the way to not detecting it.
    s["drift"] = {row["month"]: row.get("changes") or []
                  for row in closings.verify(year) if row["status"] == "drifted"}
    return s


class ClosingAccept(BaseModel):
    year: int
    period: str          # "YYYY-MM" or "annual"


@app.post("/api/closing-accept")
def closing_accept(a: ClosingAccept):
    """Adopt a drifted period's current figures as the settled ones.

    The alternative to accepting is reopening, changing things back and closing again.
    Without this, a legitimate correction leaves a permanent complaint, and a check
    that cannot be cleared is one people learn to ignore.
    """
    if a.period == "annual":
        closings.record_year(a.year)
    else:
        try:
            month = int(str(a.period).split("-")[1])
            if not 1 <= month <= 12:
                raise ValueError(a.period)
        except (IndexError, ValueError):
            raise HTTPException(400, "period must be YYYY-MM or 'annual'")
        if store.months_state(a.year).get(a.period) != "closed":
            raise HTTPException(409, "Month %s is not closed." % a.period)
        closings.record(a.year, month)
    return {"ok": True}


@app.get("/api/coverage")
def statement_coverage(year: int):
    return coverage.coverage(year)


@app.get("/api/doctor")
def data_doctor(year: int = None):
    """The whole store, or one year when a period is being checked on its own."""
    return doctor.run(year)


class AnchorAdd(BaseModel):
    account: str
    date: str
    balance: float
    # Correcting a typo has to be possible, but only when the caller says so: recording a
    # different balance for a date that already has one is a contradiction by default (409).
    replace: bool = False


@app.post("/api/anchor")
def anchor_add(item: AnchorAdd):
    accounts, _ = load_accounts()
    if item.account not in accounts:
        raise HTTPException(400, "unknown account")
    try:
        parsed = datetime.strptime(item.date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "date must be YYYY-MM-DD")
    if parsed.isoformat() != item.date:
        raise HTTPException(400, "date must be YYYY-MM-DD")
    if item.replace:
        anchors.remove(item.account, item.date)
    try:
        result = anchors.add_manual(item.account, item.date, item.balance)
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    if result["conflicts"]:
        raise HTTPException(409, "a different balance is already recorded for this account and date")
    return {"ok": True, "result": result}


@app.get("/api/anchors")
def anchors_list(account: str = ""):
    accounts, _ = load_accounts()
    if account:
        if account not in accounts:
            raise HTTPException(400, "unknown account")
        return {"anchors": anchors.list_for(account)}
    return {"anchors": [row for a in accounts for row in anchors.list_for(a)]}


class AnchorDelete(BaseModel):
    account: str
    date: str


@app.post("/api/anchor-delete")
def anchor_delete(item: AnchorDelete):
    try:
        result = anchors.remove(item.account, item.date)
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    if not result["removed"]:
        raise HTTPException(404, "no anchor recorded for that account and date")
    return {"ok": True, "result": result}


@app.get("/api/balances")
def balances_view(year: int):
    """The Settings › Balances grid: every account × every month end of one year."""
    return balances.grid(year)


@app.get("/api/networth")
def networth_view(year: int | None = None):
    return networth.series(year=year)


BUDGETS_FILE = RULES / "budgets.json"


@app.get("/api/budgets")
def get_budgets():
    return read_json(BUDGETS_FILE, default={"_help": "Monthly targets in EUR per category slug.", "budgets": {}})


class Budgets(BaseModel):
    budgets: dict


@app.post("/api/budgets")
def set_budgets(b: Budgets):
    clean = {k: float(v) for k, v in b.budgets.items() if v not in (None, "", 0)}
    write_json(BUDGETS_FILE, {"_help": "Monthly targets in EUR per category slug.", "budgets": clean})
    return {"ok": True}


@app.get("/api/overview")
def overview_view(scope: str = "all"):
    """Every year on record in one payload — the landing page's whole dataset."""
    return overview.series(scope=_check_scope(scope))


@app.get("/api/yoy")
def yoy(scope: str = "all"):
    _check_scope(scope)
    out = []
    for y in store.years():
        s = settle.year_summary(y, scope=scope)
        out.append({"year": y, "income": s["income"], "expenses": s["expenses"], "savings": s["savings"],
                    "savings_rate": round(s["savings"] / s["income"], 4) if s["income"] > 0 else None,
                    "by_category": s["by_category"]})
    return {"years": out}


@app.get("/api/review")
def review(year: int):
    txns = [t for t in store.effective_year(year) if t["status"] == "needs_review"]
    txns.sort(key=lambda t: (t.get("counterparty") or "", t["date"]))
    return {"items": txns}


@app.get("/api/review-count")
def review_count(year: int):
    return {"count": sum(1 for t in store.effective_year(year) if t["status"] == "needs_review")}


@app.get("/api/transfers")
def transfers_list(year: int):
    """Everything detection has excluded from the totals, so it can be reviewed.

    Automatic exclusion is the only operation that removes money from every figure
    without anyone agreeing to it. Listing it is what makes a wrong call visible
    instead of silent.
    """
    months = store.months_state(year)
    rows = transfers.detected(year)
    by_id = {row["id"]: row for row in rows}

    # One movement of money, seen twice. Reviewing it twice would invite the user to
    # answer the same question two ways, which is the one answer the data cannot hold.
    groups, seen = [], set()
    for row in rows:
        if row["id"] in seen:
            continue
        legs = [row]
        partner = by_id.get(row.get("partner_id") or "")
        if partner is None and row.get("partner"):
            # The far leg of a pair booked across New Year lives in another year's
            # file, so it is not among this year's rows. Showing the movement as two
            # unrelated single-leg decisions would invite two different answers to
            # one question — detection already matches across the boundary.
            partner = dict(row["partner"], reason=row["reason"], status=row["status"],
                           partner_id=row["id"], partner=None)
        if partner is not None and partner["id"] != row["id"]:
            legs.append(partner)
        seen.update(leg["id"] for leg in legs)
        for leg in legs:
            leg["month_closed"] = (store.months_state(int(leg["date"][:4]))
                                   if leg["date"][:4] != str(year) else months
                                   ).get(leg["date"][:7]) == "closed"
        groups.append({
            "id": row["id"], "reason": row["reason"], "status": row["status"],
            "amount_eur": abs(row["amount_eur"]),
            "date": min(leg["date"] for leg in legs),
            "month_closed": any(leg["month_closed"] for leg in legs),
            "legs": sorted(legs, key=lambda leg: -leg["amount_eur"]),
        })
    groups.sort(key=lambda g: (g["status"] != "pending", g["date"], g["id"]))
    return {"items": groups,
            "pending": sum(1 for g in groups if g["status"] == "pending"),
            "confirmed": sum(1 for g in groups if g["status"] == "confirmed"),
            "rejected": sum(1 for g in groups if g["status"] == "rejected")}


@app.get("/api/transfers-pending-count")
def transfers_pending_count(year: int):
    return {"count": transfers_list(year)["pending"]}


class TransferVerdict(BaseModel):
    year: int
    id: str
    confirmed: bool


@app.post("/api/transfer-confirm")
def transfer_confirm(v: TransferVerdict):
    """Record a human verdict on a detected transfer. Both legs move together.

    A pair means one movement of money seen twice, so confirming or rejecting one
    side and leaving the other is never coherent.

    The month lock guards figures that have been settled, so it applies to any write
    that changes them — rejecting a transfer puts money back into the totals. Merely
    confirming what is already excluded changes no figure at all, and blocking that
    would make a closed year impossible to audit, which is the opposite of the point.
    """
    # A pair can straddle a year boundary — booked 30 December, landing 2 January —
    # and detection already matches across adjacent years. Writing only the requested
    # year left the far leg excluded while the near one was counted.
    neighbours = [y for y in (v.year - 1, v.year, v.year + 1) if y in store.years()]
    raw = {t["id"]: (y, t) for y in neighbours for t in store.load_year_raw(y)}
    if v.id not in raw:
        raise HTTPException(404, "unknown transaction")
    target_year, target = raw[v.id]
    kind = "internal-transfer" if v.confirmed else "normal"

    legs = [(target_year, target)]
    partner = raw.get(target.get("transfer_partner") or "")
    if partner is not None and partner[1]["id"] != target["id"]:
        legs.append(partner)

    years = sorted({year for year, _ in legs})
    decisions = {year: store.decisions(year) for year in years}
    months = {year: store.months_state(year) for year in years}

    # Validate every leg before writing any of them, so a locked month on one side
    # cannot leave the pair half-answered.
    for year, leg in legs:
        excluded_now = (decisions[year].get(leg["id"], {}).get("kind")
                        or leg.get("kind")) == "internal-transfer"
        if excluded_now == v.confirmed:
            continue  # no figure moves, so the lock has nothing to protect
        key = leg["date"][:7]
        if months[year].get(key) == "closed":
            raise HTTPException(409, "Month %s is closed. Reopen it first." % key)

    for year, leg in legs:
        decisions[year].setdefault(leg["id"], {})["kind"] = kind
    # Every decision is persisted before any reconciliation runs. Saving and
    # reconciling one year at a time meant the first pass saw the far leg as still
    # unanswered, released it as an orphan and dropped its pairing — after which a
    # later verdict from that side could no longer find its partner.
    for year in years:
        store.save_decisions(year, decisions[year])
    for year in years:
        _reconcile_transfers(year)
    return {"ok": True, "updated": [leg["id"] for _, leg in legs]}


@app.get("/api/transactions")
def transactions(year: int, month: int = None):
    txns = store.effective_year(year)
    if month is not None:
        txns = [t for t in txns if int(t["date"][5:7]) == month]
    txns.sort(key=lambda t: t["date"])
    return {"items": txns}


@app.get("/api/settlement")
def settlement(year: int, month: int = None):
    return settle.settlement(year, month)


@app.get("/api/settlement-transfers")
def get_settlement_transfers(year: int):
    return {"transfers": read_json(DATA / str(year) / "settlement-transfers.json", default=[])}


class SettlementTransfer(BaseModel):
    year: int
    sender: str
    receiver: str
    amount: float
    note: str = ""


@app.post("/api/settlement-transfers")
def add_settlement_transfer(t: SettlementTransfer):
    from datetime import date as _date
    import time as _time
    people = _people()
    if t.sender not in people or t.receiver not in people or t.sender == t.receiver:
        raise HTTPException(400, "sender/receiver must be two different people: %s" % people)
    if t.amount <= 0:
        raise HTTPException(400, "amount must be positive")
    path = DATA / str(t.year) / "settlement-transfers.json"
    data = read_json(path, default=[])
    data.append({"id": str(int(_time.time() * 1000)), "date": _date.today().isoformat(),
                 "sender": t.sender, "receiver": t.receiver, "amount": round(float(t.amount), 2),
                 "note": t.note.strip()})
    write_json(path, data)
    return {"ok": True, "transfers": data}


class DeleteSettlementTransfer(BaseModel):
    year: int
    id: str


@app.post("/api/settlement-transfer-delete")
def delete_settlement_transfer(d: DeleteSettlementTransfer):
    path = DATA / str(d.year) / "settlement-transfers.json"
    data = [x for x in read_json(path, default=[]) if x.get("id") != d.id]
    write_json(path, data)
    return {"ok": True, "transfers": data}


@app.get("/api/ratio-override")
def get_ratio_override(year: int):
    return {"overrides": settle.ratio_overrides(year)}


class RatioOverride(BaseModel):
    year: int
    key: str                         # 'annual' or '1'..'12'
    ratio: Optional[dict] = None     # {person: fraction}; None clears the override


@app.post("/api/ratio-override")
def set_ratio_override(o: RatioOverride):
    people = _people()
    # The ratio decides how a shared cost is divided, so changing it moves the amount
    # one person owes the other without touching a single transaction. That made it a
    # way to rewrite a settled period in silence: no transaction changed, no total
    # changed, and the month lock never saw it.
    # An override key is 'annual' or a bare month number ('3'); month state is keyed
    # 'YYYY-MM'. Looking '3' up in {'2026-03': 'closed'} never matched, so the lock
    # this block exists to enforce silently passed every monthly override it saw.
    annual = o.key in (None, "", "annual")
    # settle.ratio_override() looks monthly overrides up as str(month) and the annual one
    # as 'annual'. Any other key would be stored and then read by nothing, so it is an
    # error rather than a no-op that looks like it worked.
    key = "annual" if annual else o.key
    if not annual:
        if not re.fullmatch(r"[1-9]|1[0-2]", o.key):
            raise HTTPException(400, "key must be 'annual' or a month number 1-12")
        month_key = "%d-%02d" % (o.year, int(o.key))
    months = store.months_state(o.year)
    locked = ([month for month, state in months.items() if state == "closed"] if annual else
              [month_key] if months.get(month_key) == "closed" else [])
    if locked:
        raise HTTPException(409, "Month %s is closed. Reopen it first." % sorted(locked)[0])
    path = DATA / str(o.year) / "ratio-overrides.json"
    data = read_json(path, default={})
    if o.ratio is None:
        data.pop(key, None)
    else:
        if not all(p in o.ratio for p in people):
            raise HTTPException(400, "ratio must include all people: %s" % people)
        values = {p: float(o.ratio[p]) for p in people}
        if any(not math.isfinite(value) or value < 0 for value in values.values()) \
                or sum(values.values()) <= 0:
            raise HTTPException(400, "ratio values must be finite, non-negative and sum to more than zero")
        data[key] = values
    write_json(path, data)
    return {"ok": True, "overrides": data}


@app.get("/api/findings")
def findings(year: int, month: int):
    path = DATA / str(year) / "findings" / ("%d-%02d.json" % (year, month))
    if not path.exists():
        return {"exists": False, "findings": []}
    data = read_json(path)
    return {"exists": True, "findings": data.get("findings", []), "summary": data.get("summary", "")}


class FindingDismiss(BaseModel):
    year: int
    month: int
    index: int


@app.post("/api/finding-dismiss")
def finding_dismiss(f: FindingDismiss):
    path = DATA / str(f.year) / "findings" / ("%d-%02d.json" % (f.year, f.month))
    if not path.exists():
        raise HTTPException(404, "no findings file")
    data = read_json(path)
    if not 0 <= f.index < len(data.get("findings", [])):
        raise HTTPException(400, "bad index")
    data["findings"][f.index]["dismissed"] = True
    write_json(path, data)
    return {"ok": True}


RECEIPTS = ROOT / "receipts"


def _attachments_of(dec):
    """Attachment list for a decision, migrating a legacy single `receipt`."""
    atts = list(dec.get("attachments") or [])
    if not atts and dec.get("receipt"):
        atts = [{"file": dec["receipt"], "description": ""}]
    return atts


def _sync_receipt(dec, atts):
    """Keep the legacy `receipt` field mirrored to the first attachment so tax
    evidence / CLI keep working unchanged."""
    dec["attachments"] = atts
    if atts:
        dec["receipt"] = atts[0]["file"]
    else:
        dec.pop("receipt", None)
        dec.pop("attachments", None)


@app.post("/api/attachment-add")
async def attachment_add(year: int = Form(...), txn_id: str = Form(...),
                         description: str = Form(""), file: UploadFile = File(...)):
    _assert_open(year, txn_id)
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", file.filename or "attachment")
    dest = RECEIPTS / str(year)
    dest.mkdir(parents=True, exist_ok=True)
    # unique per upload so multiple files with the same name don't collide
    rel = "%d/%s__%d__%s" % (year, txn_id.replace("#", "-"), int(time.time() * 1000), safe)
    (RECEIPTS / rel).write_bytes(await file.read())
    decs = store.decisions(year)
    dec = decs.setdefault(txn_id, {})
    atts = _attachments_of(dec)
    atts.append({"file": rel, "description": (description or "").strip()})
    _sync_receipt(dec, atts)
    store.save_decisions(year, decs)
    return {"ok": True, "attachments": atts}


class AttachmentDelete(BaseModel):
    year: int
    txn_id: str
    file: str


@app.post("/api/attachment-delete")
def attachment_delete(a: AttachmentDelete):
    _assert_open(a.year, a.txn_id)
    decs = store.decisions(a.year)
    dec = decs.get(a.txn_id)
    if not dec:
        raise HTTPException(404, "no decision")
    atts = [x for x in _attachments_of(dec) if x["file"] != a.file]
    _sync_receipt(dec, atts)
    store.save_decisions(a.year, decs)
    try:
        (RECEIPTS / a.file).unlink(missing_ok=True)
    except OSError:
        pass
    return {"ok": True, "attachments": atts}


@app.get("/api/recurring")
def recurring_view(year: int, scope: str = "all"):
    return recurring.detect(year, scope=_check_scope(scope))


class RecurringOverride(BaseModel):
    key: str
    state: str  # force | never | auto


@app.post("/api/recurring-override")
def recurring_override(o: RecurringOverride):
    if o.state not in ("force", "never", "auto"):
        raise HTTPException(400, "state must be force, never, or auto")
    return {"ok": True, "overrides": recurring.set_override(o.key, o.state)}


@app.get("/api/trend")
def trend(year: int, month: int, n: int = 6, scope: str = "all"):
    """Last `n` months ending at (year, month), crossing into prior years so the
    window is complete early in a year. Only months that exist are returned."""
    _check_scope(scope)
    out = []
    for i in range(n - 1, -1, -1):
        m = month - i
        y = year
        while m < 1:
            m += 12
            y -= 1
        if y not in store.years():
            continue
        s = settle.month_summary(y, m, scope=scope)
        out.append({"ym": "%04d-%02d" % (y, m), "month": m, "year": y,
                    "income": s["income"], "expenses": s["expenses"], "savings": s["savings"]})
    return {"months": out}


@app.get("/api/tax")
def tax(year: int):
    return {"report": settle.tax_report(year)}


# Selectable backup parts → what each maps to under ROOT. Default (no selection) = everything.
BACKUP_PARTS = {
    "config": ("config.json",),   # a file at the root
    "data": ("data",),            # folders (zipped recursively)
    "rules": ("rules",),
    "receipts": ("receipts",),
    "inbox": ("inbox",),
    "feedback": ("feedback",),
}


def _selected_parts(parts: str):
    """No selection means everything; a selection means exactly what it says.

    Dropping unrecognised tokens and then falling back to "all" turned `parts=dta`
    into a request for the whole tree — a typo silently widened a backup or, worse,
    a restore."""
    requested = [p for p in parts.split(",") if p]
    unknown = [p for p in requested if p not in BACKUP_PARTS]
    if unknown:
        raise HTTPException(400, "unknown part(s): %s. Valid parts: %s" %
                            (", ".join(sorted(unknown)), ", ".join(sorted(BACKUP_PARTS))))
    return requested or list(BACKUP_PARTS)


def _write_backup(prefix: str, selected):
    """Zip the selected parts of the current tree into backups/ and return the path."""
    import zipfile
    from datetime import datetime

    root = ROOT
    dest = root / "backups"
    dest.mkdir(exist_ok=True)
    path = dest / ("%s-%s.zip" % (prefix, datetime.now().strftime("%Y%m%d-%H%M%S")))
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for part in selected:
            for entry in BACKUP_PARTS[part]:
                target = root / entry
                if target.is_dir():
                    for f in target.rglob("*"):
                        if f.is_file() and f.suffix != ".tmp":
                            z.write(f, f.relative_to(root))
                elif target.is_file():
                    z.write(target, target.relative_to(root))
    return path


@app.get("/api/backup")
def backup(parts: str = ""):
    path = _write_backup("family-accountability-backup", _selected_parts(parts))
    return FileResponse(path, filename=path.name, media_type="application/zip")


@app.post("/api/restore")
async def restore(file: UploadFile = File(...), mode: str = Form("replace"), parts: str = Form("")):
    """Restore selected parts from an uploaded backup zip. A safety backup of the CURRENT
    tree is written first (so the operation is undoable). 'replace' wipes each restored
    folder before extracting; 'merge' overwrites in place and keeps files not in the zip."""
    import io
    import json as _json
    import shutil
    import zipfile

    if mode not in ("replace", "merge"):
        raise HTTPException(400, "mode must be 'replace' or 'merge'")
    root = ROOT.resolve()
    selected = _selected_parts(parts)
    allowed_tops = {entry.split("/")[0] for p in selected for entry in BACKUP_PARTS[p]}

    raw = await file.read()
    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
    except zipfile.BadZipFile:
        raise HTTPException(400, "The uploaded file is not a valid zip archive.")

    members = [m for m in zf.namelist() if not m.endswith("/")]
    safe = []
    for m in members:
        parts_ = Path(m).parts
        if Path(m).is_absolute() or ".." in parts_ or not parts_:
            raise HTTPException(400, "Unsafe path in archive: %s" % m)          # zip-slip guard
        if parts_[0] not in allowed_tops:
            continue                                                            # outside selected parts
        dest = (root / m).resolve()
        if root != dest and root not in dest.parents:
            raise HTTPException(400, "Archive entry escapes the project folder: %s" % m)
        safe.append(m)
    if not safe:
        raise HTTPException(400, "The archive has nothing to restore for the selected parts.")

    if "config.json" in safe:                                                   # do not brick the app
        try:
            cfg = _json.loads(zf.read("config.json"))
        except ValueError:
            raise HTTPException(400, "config.json in the archive is not valid JSON.")
        if not (isinstance(cfg.get("people"), list) and len(cfg["people"]) == 2):
            raise HTTPException(400, "config.json in the archive is missing a valid 'people' list.")

    safety = _write_backup("pre-restore", list(BACKUP_PARTS))                   # undo point

    present_tops = {Path(m).parts[0] for m in safe}
    if mode == "replace":
        for top in present_tops:
            target = root / top
            if target.is_dir():
                shutil.rmtree(target)
            elif target.is_file():
                target.unlink()
    for m in safe:
        dest = root / m
        dest.parent.mkdir(parents=True, exist_ok=True)
        with zf.open(m) as src, open(dest, "wb") as out:
            shutil.copyfileobj(src, out)

    return {"ok": True, "restored": len(safe), "mode": mode,
            "parts": sorted(present_tops), "safety_backup": safety.name}


def _write_year_backup(year: int):
    """Zip data/<year> + receipts/<year> into backups/ (paths relative to ROOT, so the zip
    is restorable through /api/restore). Returns the path, or None if nothing was written."""
    import zipfile
    from datetime import datetime

    root = ROOT
    dest = root / "backups"
    dest.mkdir(exist_ok=True)
    path = dest / ("deleted-year-%d-%s.zip" % (year, datetime.now().strftime("%Y%m%d-%H%M%S")))
    wrote = False
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for target in (DATA / str(year), RECEIPTS / str(year)):
            if target.is_dir():
                for f in target.rglob("*"):
                    if f.is_file() and f.suffix != ".tmp":
                        z.write(f, f.relative_to(root))
                        wrote = True
    if not wrote:
        path.unlink(missing_ok=True)
        return None
    return path


class DeleteYear(BaseModel):
    year: int
    confirm: str


@app.post("/api/delete-year")
def delete_year(req: DeleteYear):
    """Permanently delete one year: data/<year> (transactions, decisions, months, overrides,
    settlement transfers, findings) and receipts/<year>. Guarded by an exact typed phrase
    (validated here too — never trust the client). A safety backup is written first."""
    year = req.year
    if not (1900 <= year <= 2999):
        raise HTTPException(400, "Invalid year.")
    ydir = DATA / str(year)
    if not ydir.is_dir():
        raise HTTPException(404, "No data stored for %d." % year)
    expected = "delete %d" % year
    if (req.confirm or "").strip() != expected:
        raise HTTPException(400, 'Type exactly "%s" to confirm.' % expected)
    safety = _write_year_backup(year)
    shutil.rmtree(ydir)
    rdir = RECEIPTS / str(year)
    if rdir.is_dir():
        shutil.rmtree(rdir)
    store._EFFECTIVE_CACHE.pop(year, None)   # drop the cached derived view (self-invalidates anyway)
    return {"ok": True, "year": year, "years": store.years(),
            "safety_backup": safety.name if safety else None}


@app.get("/api/tax-export")
def tax_export(year: int):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    report = settle.tax_report(year)
    wb = Workbook()
    wb.remove(wb.active)
    people = sorted({o for b in report for o in b["owners"]}) or ["-"]
    bold = Font(bold=True)
    for person in people:
        ws = wb.create_sheet(title=person.capitalize())
        ws.append(["Tax evidence pack %d — %s" % (year, person.capitalize())])
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([])
        grand = 0.0
        for b in report:
            if person not in b["owners"]:
                continue
            d = b["owners"][person]
            ws.append([b["name"], "", "", round(-d["confirmed_total"], 2), "confirmed total"])
            ws.cell(row=ws.max_row, column=1).font = bold
            ws.cell(row=ws.max_row, column=4).font = bold
            ws.append(["Date", "Counterparty", "Purpose", "Amount EUR", "Status",
                       "Receipt", "Payment proof", "Tax owner"])
            for i in d["items"]:
                status = "candidate" if not i["confirmed"] else ("ready" if i["ready"] else "missing evidence")
                ws.append([i["date"], i["counterparty"], (i["purpose"] or "")[:80], -i["amount"],
                           status, "yes" if i["has_receipt"] else "MISSING",
                           "yes" if i["payment_proof"] else "MISSING", i["tax_owner"]])
            ws.append([])
            grand += -d["confirmed_total"]
        ws.append(["TOTAL", "", "", round(grand, 2)])
        ws.cell(row=ws.max_row, column=1).font = bold
        ws.cell(row=ws.max_row, column=4).font = bold
        for col, width in zip("ABCDEFGH", (12, 40, 60, 12, 18, 12, 16, 16)):
            ws.column_dimensions[col].width = width
    out = DATA / str(year) / ("tax-evidence-%d.xlsx" % year)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return FileResponse(out, filename="tax-evidence-%d.xlsx" % year,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


EXPORT_COLUMNS = [
    ("row_type", "transaction, or split-part for one leg of a split"),
    ("in_expense_math", "TRUE for the rows the dashboards and settlement actually count"),
    ("is_income", "TRUE when the category is an income category. NOT the sign: a refund "
                  "booked to an expense category is a negative expense, never income."),
    ("month", "1-12, so the Summary sheet can pivot on it"),
    ("date", ""), ("account", ""), ("account_label", ""), ("owner", ""),
    ("counterparty", ""), ("purpose", ""), ("note", ""),
    ("category_group", ""), ("category_sub", ""), ("category_slug", ""),
    ("sharing", ""), ("income_owner", ""), ("year_cost", ""),
    ("tax_bucket", ""), ("tax_confirmed", ""), ("tax_owner", ""),
    ("amount_eur", ""), ("amount_original", ""), ("currency", ""), ("fx_rate", ""),
    ("kind", ""), ("status", ""), ("matched_rule", ""),
    ("counterparty_iban", ""),
    ("source_file", "the statement this came from"), ("source_format", ""),
    ("transaction_id", ""), ("attachments", ""),
]


def _category_names():
    """slug -> (group name, sub name), for the two category columns."""
    names = {}
    for group in read_json(RULES / "categories.json")["categories"]:
        for sub in group.get("subs", []):
            names["%s/%s" % (group["slug"], sub["slug"])] = (group["name"], sub["name"])
    return names


def _export_rows(year):
    """One row per money line: split parts replace their parent in the totals.

    A split transaction cannot be one honest row — its legs carry their own
    categories and sharing. Both are emitted, and `in_expense_math` marks exactly
    the rows the app counts, so summing that subset reproduces the dashboards
    instead of double-counting a split or quietly including an internal transfer.
    """
    accounts, _ = load_accounts()
    names = _category_names()
    income_cats = settle.income_categories()
    rows = []

    def line(txn, part=None):
        source = txn.get("source") or {}
        account = accounts.get(txn.get("account"), {})
        # The one definition of split-part inheritance, shared with the totals. A
        # part states only what differs from its parent; resolving that here by hand
        # made the workbook disagree with the app about category, year_cost and
        # tax_bucket — on the very rows an external audit reconciles.
        view = settle.part_view(txn, part)
        category = view["category"]
        group_name, sub_name = names.get(category or "", ("", ""))
        sharing = view["sharing"]
        counted = (txn.get("kind") != "internal-transfer" and sharing != "out-of-scope"
                   and (part is not None or not txn.get("splits")))
        amount = part["amount"] if part else txn.get("amount_eur")
        attachments = [a.get("file", "") for a in (txn.get("attachments") or [])]
        # Mirrors settle._is_income exactly: category first, sign only as a fallback
        # for an uncategorized line.
        is_income = category in income_cats if category else (amount or 0) > 0
        return {
            "row_type": "split-part" if part else "transaction",
            "in_expense_math": bool(counted),
            "is_income": bool(is_income),
            "month": int(str(txn.get("date"))[5:7]),
            "date": txn.get("date"),
            "account": txn.get("account"),
            "account_label": account.get("label") or account.get("bank") or "",
            "owner": account.get("owner") or "",
            "counterparty": txn.get("counterparty") or "",
            "purpose": (part or {}).get("purpose") or txn.get("purpose") or "",
            "note": (part or txn).get("note") or "",
            "category_group": group_name,
            "category_sub": sub_name,
            "category_slug": category or "",
            "sharing": sharing or "",
            "income_owner": txn.get("income_owner") or "",
            "year_cost": view["year_cost"],
            "tax_bucket": view["tax_bucket"] or "",
            "tax_confirmed": bool(txn.get("tax_confirmed")),
            "tax_owner": txn.get("tax_owner") or "",
            "amount_eur": amount,
            "amount_original": None if part else txn.get("amount_original"),
            "currency": "EUR" if part else (txn.get("currency") or "EUR"),
            "fx_rate": None if part else txn.get("fx_rate"),
            "kind": txn.get("kind") or "",
            "status": txn.get("status") or "",
            "matched_rule": txn.get("matched_rule") or "",
            "counterparty_iban": txn.get("counterparty_iban") or "",
            "source_file": source.get("file") or "",
            "source_format": source.get("format") or "",
            "transaction_id": txn.get("id"),
            "attachments": "; ".join(attachments),
        }

    for txn in sorted(store.effective_year(year), key=lambda t: (t["date"], t.get("account") or "")):
        rows.append(line(txn))
        for part in txn.get("splits") or []:
            rows.append(line(txn, part))
    return rows


MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]


def _add_summary_sheet(wb, year, data_sheet, headers, row_count):
    """The dashboard's figures, each beside a formula that rebuilds it from the
    Transactions sheet and a difference that must read 0.

    An audit wants to check the numbers, not take them on faith, so the workbook
    carries its own derivation: the app's stored value, a live SUMIFS over the raw
    rows, and the gap between them.
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    summary = settle.year_summary(year)
    col = {name: get_column_letter(i) for i, name in enumerate(headers, start=1)}
    last = row_count + 1
    quoted = "'%s'" % data_sheet

    def rng(name):
        return "%s!$%s$2:$%s$%d" % (quoted, col[name], col[name], last)

    def sumifs(income, month=None, exclude_year_costs=False):
        parts = ["%s" % rng("amount_eur"),
                 "%s,TRUE" % rng("in_expense_math"),
                 "%s,%s" % (rng("is_income"), "TRUE" if income else "FALSE")]
        if month:
            parts.append("%s,%d" % (rng("month"), month))
        if exclude_year_costs:
            parts.append("%s,FALSE" % rng("year_cost"))
        return "=SUMIFS(%s)" % ",".join(parts)

    ws = wb.create_sheet("Summary")
    bold = Font(bold=True)
    ws.append(["Summary %d" % year])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Year total", "App", "Recomputed from rows", "Difference"])
    for cell in ws[3]:
        cell.font = bold
    totals = [
        ("Money in (income)", summary["income"], sumifs(True)),
        ("Money out (expenses)", summary["expenses"], sumifs(False)),
    ]
    for label, value, formula in totals:
        ws.append([label, value, formula, "=C%d-B%d" % (ws.max_row + 1, ws.max_row + 1)])
    savings_row = ws.max_row + 1
    ws.append(["Savings (in + out)", summary["savings"], "=C%d+C%d" % (savings_row - 2, savings_row - 1),
               "=C%d-B%d" % (savings_row, savings_row)])
    ws.append(["Savings rate", None if not summary["income"] else summary["savings"] / summary["income"],
               "=IF(C%d=0,\"\",C%d/C%d)" % (savings_row - 2, savings_row, savings_row - 2), ""])
    ws.cell(row=ws.max_row, column=2).number_format = "0%"
    ws.cell(row=ws.max_row, column=3).number_format = "0%"
    ws.append(["Transactions counted", summary["transactions"], "", ""])
    ws.append(["Needs review", summary.get("needs_review", 0), "", ""])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(["Month", "Money in", "Money out", "Savings", "Savings rate",
               "Year costs (excluded)", "Transactions",
               "Money in (recomputed)", "Money out (recomputed)", "Difference in", "Difference out"])
    for cell in ws[header_row]:
        cell.font = bold
    for month in summary["months"]:
        index = month["month"]
        row = ws.max_row + 1
        ws.append([
            MONTH_NAMES[index - 1], month["income"], month["expenses"], month["savings"],
            None if not month["income"] else month["savings"] / month["income"],
            month.get("year_costs_excluded", 0), month.get("transactions", 0),
            sumifs(True, month=index, exclude_year_costs=True),
            sumifs(False, month=index, exclude_year_costs=True),
            "=H%d-B%d" % (row, row), "=I%d-C%d" % (row, row),
        ])
        ws.cell(row=row, column=5).number_format = "0%"
    ws.append([])
    ws.append(["Note", "Monthly figures exclude year-cost entries; the year total includes them. "
                       "That is why the twelve months do not add up to the year row."])
    ws.append(["", "Every Difference column must read 0. If one does not, the sheet and the rows "
                   "disagree and neither should be trusted."])
    ws.append(["", "Money out is negative. Income is decided by category, not by sign: a refund "
                   "booked to an expense category reduces expenses rather than adding income."])
    for column, width in (("A", 24), ("B", 16), ("C", 22), ("D", 16), ("E", 13),
                          ("F", 20), ("G", 14), ("H", 20), ("I", 21), ("J", 13), ("K", 13)):
        ws.column_dimensions[column].width = width
    return ws


@app.get("/api/transactions-export")
def transactions_export(year: int):
    """Every transaction of one year as a single flat table."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    rows = _export_rows(year)
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions %d" % year
    headers = [name for name, _ in EXPORT_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(name) for name in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), max(len(rows) + 1, 1))
    for index, name in enumerate(headers, start=1):
        widest = max([len(name)] + [len(str(r.get(name) or "")) for r in rows[:400]])
        ws.column_dimensions[get_column_letter(index)].width = min(max(widest + 2, 10), 52)

    _add_summary_sheet(wb, year, ws.title, headers, len(rows))

    legend = wb.create_sheet("Legend")
    legend.append(["Column", "Meaning"])
    for cell in legend[1]:
        cell.font = Font(bold=True)
    for name, meaning in EXPORT_COLUMNS:
        if meaning:
            legend.append([name, meaning])
    legend.append([])
    legend.append(["Totals", "Filter in_expense_math = TRUE to reproduce the app's figures."])
    legend.append(["", "It excludes internal transfers and out-of-scope rows, which the app "
                       "ignores everywhere, and takes a split's parts instead of its parent."])
    legend.append(["Fairness", "Deliberately absent: the income-proportional split is a yearly "
                               "figure computed from salary, never a per-transaction value. See the "
                               "Settlement tab."])
    legend.column_dimensions["A"].width = 20
    legend.column_dimensions["B"].width = 100

    out = DATA / str(year) / ("transactions-%d.xlsx" % year)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return FileResponse(out, filename="transactions-%d.xlsx" % year,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


class Decision(BaseModel):
    year: int
    id: str
    fields: dict


DECISION_FIELDS = {"category", "sharing", "income_owner", "tax_bucket", "tax_confirmed",
                   "tax_owner", "year_cost", "splits", "note", "receipt", "attachments",
                   "account", "kind", "force_review"}


def _decision_options():
    categories = read_json(RULES / "categories.json")["categories"]
    category_slugs = {"%s/%s" % (group["slug"], sub["slug"])
                      for group in categories for sub in group.get("subs", [])}
    category_slugs.add("auto:items")
    tax_slugs = {bucket["slug"] for bucket in read_json(RULES / "tax-buckets.json")["buckets"]}
    accounts, _ = load_accounts()
    people = _people()
    return category_slugs, tax_slugs, set(accounts), people


def _validate_decision_fields(fields, raw, options=None):
    unknown = sorted(set(fields) - DECISION_FIELDS)
    if unknown:
        raise HTTPException(400, "Unknown decision field(s): %s" % ", ".join(unknown))
    categories, tax_buckets, accounts, people = options or _decision_options()
    sharings = {"shared", "out-of-scope"} | {"personal:" + p for p in people}

    def valid_category(value, label="category"):
        if value is not None and value not in categories:
            raise HTTPException(400, "%s '%s' does not exist" % (label, value))

    def valid_sharing(value, label="sharing"):
        if value not in sharings:
            raise HTTPException(400, "%s must be shared, out-of-scope, or personal:<person>" % label)

    valid_category(fields.get("category"))
    if "sharing" in fields:
        valid_sharing(fields["sharing"])
    if "income_owner" in fields and fields["income_owner"] not in people + ["couple", None]:
        raise HTTPException(400, "income_owner must be one of %s or couple" % people)
    if "tax_owner" in fields and fields["tax_owner"] not in people + [None]:
        raise HTTPException(400, "tax_owner must be one of %s" % people)
    if "account" in fields and fields["account"] not in accounts:
        raise HTTPException(400, "Unknown account '%s'" % fields["account"])
    if "tax_bucket" in fields and fields["tax_bucket"] is not None and fields["tax_bucket"] not in tax_buckets:
        raise HTTPException(400, "Unknown tax bucket '%s'" % fields["tax_bucket"])
    if "kind" in fields and fields["kind"] not in ("normal", "internal-transfer"):
        raise HTTPException(400, "kind must be normal or internal-transfer")
    for flag in ("tax_confirmed", "year_cost", "force_review"):
        if flag in fields and not isinstance(fields[flag], bool):
            raise HTTPException(400, "%s must be true or false" % flag)

    if "splits" in fields and fields["splits"] is not None:
        splits = fields["splits"]
        if not isinstance(splits, list) or not splits:
            raise HTTPException(400, "splits must be a non-empty list")
        total = 0
        for index, part in enumerate(splits, 1):
            if not isinstance(part, dict) or "amount" not in part:
                raise HTTPException(400, "split %d must include an amount" % index)
            try:
                total += cents(part["amount"])
            except (TypeError, ValueError):
                raise HTTPException(400, "split %d has an invalid amount" % index)
            valid_category(part.get("category"), "split %d category" % index)
            sharing = part.get("sharing") or fields.get("sharing") or "shared"
            valid_sharing(sharing, "split %d sharing" % index)
            if not part.get("category") and sharing != "out-of-scope":
                raise HTTPException(400, "split %d needs a category or out-of-scope sharing" % index)
            bucket = part.get("tax_bucket")
            if bucket is not None and bucket not in tax_buckets:
                raise HTTPException(400, "split %d has unknown tax bucket '%s'" % (index, bucket))
        if total != cents(raw["amount_eur"]):
            raise HTTPException(400, "split amounts sum to %.2f, expected %.2f" %
                                (total / 100.0, cents(raw["amount_eur"]) / 100.0))


def _reconcile_transfers(year):
    """Re-run transfer detection after anything that changes its inputs.

    Detection reads categories, splits, kind decisions and the raw rows themselves, so
    every one of those is a reason to ask the question again. Saving a decision without
    doing this left a categorised transaction still excluded as a transfer — the
    category applied to nothing until the next ingest, and the money stayed missing
    from every total in the meantime.

    Closed months are deliberately not exempt. Skipping them would leave one leg of a
    pair straddling a month boundary unmarked, and would make a wrong exclusion inside
    a closed month permanent. Detection runs everywhere; closings.py records what each
    closed month contained so any change it makes is reported rather than silent.
    """
    transfers.mark_internal(year)


@app.post("/api/decision")
def decision(d: Decision):
    raw = _assert_open(d.year, d.id)
    _validate_decision_fields(d.fields, raw)
    decs = store.decisions(d.year)
    current = decs.get(d.id, {})
    current.update(d.fields)
    decs[d.id] = current
    store.save_decisions(d.year, decs)
    _reconcile_transfers(d.year)
    return {"ok": True}


class DecisionsBulk(BaseModel):
    year: int
    items: list


@app.post("/api/decisions-bulk")
def decisions_bulk(batch: DecisionsBulk):
    """Validate an entire review batch, then persist it with one file write."""
    if not batch.items:
        raise HTTPException(400, "items must not be empty")
    raw_by_id = {txn["id"]: txn for txn in store.load_year_raw(batch.year)}
    months = store.months_state(batch.year)
    validated = []
    seen = set()
    options = _decision_options()
    for index, item in enumerate(batch.items, 1):
        if not isinstance(item, dict) or not isinstance(item.get("fields"), dict) or not item.get("id"):
            raise HTTPException(400, "item %d must include id and fields" % index)
        txn_id = item["id"]
        if txn_id in seen:
            raise HTTPException(400, "duplicate transaction id '%s'" % txn_id)
        seen.add(txn_id)
        raw = raw_by_id.get(txn_id)
        if not raw:
            raise HTTPException(404, "unknown transaction '%s'" % txn_id)
        key = raw["date"][:7]
        if months.get(key) == "closed":
            raise HTTPException(409, "Month %s is closed. Reopen it first." % key)
        _validate_decision_fields(item["fields"], raw, options=options)
        validated.append((txn_id, item["fields"]))
    decs = store.decisions(batch.year)
    for txn_id, fields in validated:
        current = decs.get(txn_id, {})
        current.update(fields)
        decs[txn_id] = current
    store.save_decisions(batch.year, decs)
    _reconcile_transfers(batch.year)
    return {"ok": True, "updated": len(validated)}


class RulePayload(BaseModel):
    pattern: str
    field: str = "counterparty"
    category: Optional[str] = None
    sharing: str = "shared"
    tax_bucket: Optional[str] = None
    note: Optional[str] = None
    action: Optional[str] = None
    scope: str = "family"  # family | <person>


RULE_MATCH_FIELDS = ("counterparty", "purpose", "any")
RULE_MIN_PATTERN = 3      # same floor as the UI: shorter patterns over-match the whole history


def _validate_rule(*, pattern=None, field=None, category=None, sharing=None,
                   tax_bucket=None, action=None, scope=None):
    """One definition of a valid rule, shared by create and update.

    A rule is retroactive by nature — it applies to every transaction already in the
    store and every one still to come. Creation used to check `scope` and nothing
    else, so an empty pattern (a substring of every string) reclassified the whole
    history in one POST, and a category that does not exist could be written into a
    rule that every screen then had to cope with. Update checked all of it. The two
    paths write the same file, so they answer to the same rules.

    Every argument is optional because update is a patch: None means "not being set".
    """
    people = _people()
    if scope is not None and scope not in ["family"] + people:
        raise HTTPException(400, "scope must be 'family' or one of %s" % people)
    if pattern is not None and len((pattern or "").strip()) < RULE_MIN_PATTERN:
        raise HTTPException(400, "pattern must contain at least %d characters" % RULE_MIN_PATTERN)
    if field is not None and field not in RULE_MATCH_FIELDS:
        raise HTTPException(400, "field must be counterparty, purpose or any")
    if sharing is not None and sharing not in {"shared", "out-of-scope"} | {"personal:" + p for p in people}:
        raise HTTPException(400, "sharing must be shared, out-of-scope, or personal:<person>")
    if category:
        categories, _, _, _ = _decision_options()
        if category not in categories:
            raise HTTPException(400, "category '%s' does not exist" % category)
    if tax_bucket:
        tax_slugs = {bucket["slug"] for bucket in read_json(RULES / "tax-buckets.json")["buckets"]}
        if tax_bucket not in tax_slugs:
            raise HTTPException(400, "Unknown tax bucket '%s'" % tax_bucket)
    if action is not None and action not in ("review", ""):
        raise HTTPException(400, "action must be 'review' or empty")


@app.post("/api/rule")
def add_rule(r: RulePayload):
    _validate_rule(pattern=r.pattern, field=r.field, sharing=r.sharing, scope=r.scope,
                   action=r.action,
                   # A rule either routes to Review or classifies, never both, so the
                   # classification fields are only validated when they will be stored.
                   category=None if r.action == "review" else r.category,
                   tax_bucket=None if r.action == "review" else r.tax_bucket)
    slug = re.sub(r"[^a-z0-9]+", "-", r.pattern.lower()).strip("-")[:40] or "rule"
    base_id = "ui-%s-%s" % (r.scope, slug)
    existing_ids = {item["id"] for item in rules_engine.load_rules()}
    rule_id, suffix = base_id, 2
    while rule_id in existing_ids:
        rule_id = "%s-%d" % (base_id, suffix)
        suffix += 1
    rule = {
        "id": rule_id,
        "match": {"field": r.field, "contains": r.pattern.strip()},
        "sharing": r.sharing,
        "scope": r.scope,
    }
    if r.action == "review":
        rule["action"] = "review"
    else:
        rule["category"] = r.category
        if r.tax_bucket:
            rule["tax_bucket"] = r.tax_bucket
        if r.note:
            rule["note"] = r.note
    rules_engine.add_rule(rule)
    return {"ok": True, "rule": rule}


@app.get("/api/rules")
def list_rules():
    return read_json(RULES / "merchant-rules.json")


RULE_CONTROLLED_FIELDS = {"category", "sharing", "tax_bucket", "tax_confirmed", "tax_owner"}


def _rule_impact(rule_id, year, apply=False):
    """Preview or clear manual classifications so one rule controls matching rows."""
    rules = rules_engine.load_rules()
    rule = next((item for item in rules if item["id"] == rule_id), None)
    if not rule:
        raise HTTPException(404, "unknown rule")
    accounts, _ = load_accounts()
    config = load_config()
    tax_buckets = read_json(RULES / "tax-buckets.json")["buckets"]
    decisions = store.decisions(year)
    months = store.months_state(year)
    result = {"matched": 0, "already_rule_controlled": 0, "manual_overrides": 0,
              "eligible": 0, "skipped_splits": 0, "skipped_closed": 0, "applied": 0,
              "field_matches": {"counterparty": 0, "purpose": 0, "any": 0}}
    eligible_ids = []
    for txn in store.load_year_raw(year):
        decision = decisions.get(txn["id"], {})
        account = decision.get("account") or txn["account"]
        owner = accounts.get(account, {}).get("owner")
        scope = rule.get("scope", "family")
        if scope == "family" or scope == owner:
            for field in result["field_matches"]:
                probe = dict(rule)
                probe["match"] = dict(rule["match"], field=field)
                if rules_engine.matches(probe, txn):
                    result["field_matches"][field] += 1
        hypothetical = rules_engine.effective(
            txn, None, rules, owner=owner, config=config, tax_buckets=tax_buckets
        )
        if hypothetical.get("matched_rule") != rule_id:
            continue
        result["matched"] += 1
        if not any(key in decision for key in RULE_CONTROLLED_FIELDS):
            result["already_rule_controlled"] += 1
            continue
        result["manual_overrides"] += 1
        if decision.get("splits"):
            result["skipped_splits"] += 1
            continue
        if months.get(txn["date"][:7]) == "closed":
            result["skipped_closed"] += 1
            continue
        result["eligible"] += 1
        eligible_ids.append(txn["id"])
    if apply and eligible_ids:
        for txn_id in eligible_ids:
            decision = decisions[txn_id]
            for key in RULE_CONTROLLED_FIELDS:
                decision.pop(key, None)
            if not decision:
                decisions.pop(txn_id, None)
        store.save_decisions(year, decisions)
        result["applied"] = len(eligible_ids)
    return result


@app.get("/api/rule-impact")
def rule_impact(rule_id: str, year: int):
    return _rule_impact(rule_id, year)


class RuleApply(BaseModel):
    id: str
    year: int


@app.post("/api/rule-apply")
def rule_apply(r: RuleApply):
    return _rule_impact(r.id, r.year, apply=True)


class RuleUpdate(BaseModel):
    """Every field is optional: None means "leave unchanged". For the clearable
    ones (note, category, tax_bucket, action) an empty string means "clear"."""

    id: str
    scope: Optional[str] = None
    note: Optional[str] = None
    field: Optional[str] = None
    pattern: Optional[str] = None
    category: Optional[str] = None
    sharing: Optional[str] = None
    tax_bucket: Optional[str] = None
    action: Optional[str] = None


@app.post("/api/rule-update")
def rule_update(r: RuleUpdate):
    _validate_rule(pattern=r.pattern, field=r.field, category=r.category, sharing=r.sharing,
                   tax_bucket=r.tax_bucket, action=r.action, scope=r.scope)
    fields = {}
    if r.scope is not None:
        fields["scope"] = r.scope
    if r.note is not None:
        fields["note"] = r.note or None  # empty string clears the note
    if r.field is not None:
        fields["field"] = r.field
    if r.pattern is not None:
        fields["contains"] = r.pattern.strip()
    if r.sharing is not None:
        fields["sharing"] = r.sharing
    if r.category is not None:
        fields["category"] = r.category or None
    if r.tax_bucket is not None:
        fields["tax_bucket"] = r.tax_bucket or None
    if r.action is not None:
        fields["action"] = r.action or None
    # A rule either routes to Review or classifies — never both.
    if fields.get("action") == "review":
        fields["category"] = None
        fields["tax_bucket"] = None
    elif fields.get("category"):
        fields["action"] = None
    if not fields:
        return {"ok": True}
    if not rules_engine.update_rule(r.id, fields):
        raise HTTPException(404, "unknown rule id")
    return {"ok": True}


class RuleDelete(BaseModel):
    id: str


@app.post("/api/rule-delete")
def rule_delete(r: RuleDelete):
    if not rules_engine.remove_rule(r.id):
        raise HTTPException(404, "unknown rule id")
    return {"ok": True}


class DecisionClear(BaseModel):
    year: int
    id: str


@app.post("/api/decision-clear")
def decision_clear(d: DecisionClear):
    _assert_open(d.year, d.id)
    decs = store.decisions(d.year)
    if d.id in decs:
        del decs[d.id]
        store.save_decisions(d.year, decs)
        _reconcile_transfers(d.year)
    return {"ok": True}


class DecisionsClearBulk(BaseModel):
    year: int
    ids: list


@app.post("/api/decisions-clear-bulk")
def decisions_clear_bulk(b: DecisionsClearBulk):
    """Clear the manual decision on many transactions in one write.

    Validates the whole selection before touching anything, so a batch that
    includes a closed month fails without half-applying.
    """
    if not b.ids:
        raise HTTPException(400, "ids must not be empty")
    raw_by_id = {txn["id"]: txn for txn in store.load_year_raw(b.year)}
    months = store.months_state(b.year)
    for txn_id in b.ids:
        raw = raw_by_id.get(txn_id)
        if not raw:
            raise HTTPException(404, "unknown transaction '%s'" % txn_id)
        key = raw["date"][:7]
        if months.get(key) == "closed":
            raise HTTPException(409, "Month %s is closed. Reopen it first." % key)
    decs = store.decisions(b.year)
    cleared = 0
    for txn_id in set(b.ids):
        if decs.pop(txn_id, None) is not None:
            cleared += 1
    if cleared:
        store.save_decisions(b.year, decs)
        _reconcile_transfers(b.year)
    return {"ok": True, "cleared": cleared}


@app.post("/api/decision-clear-orphan")
def decision_clear_orphan(d: DecisionClear):
    """Delete a decision whose transaction no longer exists (a doctor
    orphan-decision finding). The normal clear path 404s here because it
    requires the transaction; this one refuses when the transaction is present,
    so it can only ever remove genuine cruft."""
    if any(t["id"] == d.id for t in store.load_year_raw(d.year)):
        raise HTTPException(409, "Transaction exists; use the normal review tools.")
    decs = store.decisions(d.year)
    if d.id in decs:
        del decs[d.id]
        store.save_decisions(d.year, decs)
    return {"ok": True}


class TxnEdit(BaseModel):
    year: int
    id: str
    date: str
    counterparty: str
    amount_eur: float
    account: Optional[str] = None


@app.post("/api/transaction-edit")
def transaction_edit(e: TxnEdit):
    """Manually correct an entry's raw values (date, name, EUR amount, bank account).
    The original values are snapshotted so the edit can be reset; all money math uses
    the corrected values."""
    from datetime import date as _date
    raw = _assert_open(e.year, e.id)   # exists + its (edited) month is open
    try:
        new_date = _date.fromisoformat(e.date)
    except ValueError:
        raise HTTPException(400, "date must be YYYY-MM-DD")
    if new_date.year != e.year:
        raise HTTPException(400, "Date must stay within %d; moving an entry to another year isn't supported." % e.year)
    if store.months_state(e.year).get(e.date[:7]) == "closed":
        raise HTTPException(409, "Month %s is closed. Reopen it first." % e.date[:7])
    if not e.counterparty.strip():
        raise HTTPException(400, "Name is required.")
    if cents(e.amount_eur) == 0:
        raise HTTPException(400, "Amount must not be zero.")
    changes = {"date": e.date, "counterparty": e.counterparty.strip(),
               "amount_eur": round(e.amount_eur, 2)}
    # For EUR entries the original amount mirrors the EUR amount; keep them in sync.
    if (raw.get("currency") or "EUR").upper() == "EUR":
        changes["amount_original"] = round(e.amount_eur, 2)
    if e.account is not None and e.account != raw.get("account"):
        accounts, _ = load_accounts()
        if e.account not in accounts:
            raise HTTPException(400, "Unknown account '%s'" % e.account)
        changes["account"] = e.account
    store.edit_transaction(e.year, e.id, changes)
    _reconcile_transfers(e.year)
    return {"ok": True}


@app.post("/api/transaction-edit-reset")
def transaction_edit_reset(d: DecisionClear):
    raw = _assert_open(d.year, d.id)
    original_month = ((raw.get("original") or {}).get("date") or raw.get("date"))[:7]
    if store.months_state(d.year).get(original_month) == "closed":
        raise HTTPException(409, "Original month %s is closed. Reopen it first." % original_month)
    try:
        store.reset_transaction(d.year, d.id)
    except KeyError:
        raise HTTPException(404, "No manual edit to reset.")
    _reconcile_transfers(d.year)
    return {"ok": True}


class CashEntry(BaseModel):
    date: str
    account: str
    amount: float
    currency: str = "EUR"
    description: str
    category: str = ""


def _validate_cash_values(e):
    from datetime import date as _date

    accounts, _ = load_accounts()
    if e.account not in accounts:
        raise HTTPException(400, "Pick a valid account. Add a cash account on the Accounts tab first." if not e.account
                            else "Unknown account '%s'." % e.account)
    if accounts[e.account].get("type") != "cash":
        raise HTTPException(400, "Manual entries require a cash account")
    try:
        _date.fromisoformat(e.date)
    except ValueError:
        raise HTTPException(400, "date must be YYYY-MM-DD")
    if not e.description.strip():
        raise HTTPException(400, "description is required")
    if cents(e.amount) == 0:
        raise HTTPException(400, "amount must not be zero")
    currency = (e.currency or "EUR").upper()
    allowed = [c.upper() for c in load_config()["currencies"]]  # load_config guarantees the key
    if currency not in allowed:
        raise HTTPException(400, "currency must be one of %s (edit 'currencies' in Settings to add more)" % ", ".join(allowed))
    try:
        eur, rate = (float(e.amount), None) if currency == "EUR" else fx.to_eur(e.amount, currency, e.date)
    except Exception as ex:
        raise HTTPException(400, "cannot convert cash amount: %s" % ex)
    if e.category:
        _validate_decision_fields({"category": e.category}, {"amount_eur": eur})
    return eur, rate


def _assert_cash_month_open(day):
    year = int(day[:4])
    key = day[:7]
    if store.months_state(year).get(key) == "closed":
        raise HTTPException(409, "Month %s is closed. Reopen it first." % key)


def _cash_snapshot():
    return [t for year in store.years() for t in store.load_year_by_file(year).get("cash.jsonl", [])]


def _cash_rows():
    path = INBOX / "cash.csv"
    if not path.exists():
        return ["date", "account", "amount", "currency", "description", "category"], []
    rows = list(csv.reader(open(path, encoding="utf-8")))
    return rows[0], rows[1:]


def _publish_bytes(path, payload):
    """Replace a file in one step, so no reader ever sees it half-written."""
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(dir=str(path.parent), prefix=path.name + ".", suffix=".tmp")
    try:
        with os.fdopen(fd, "wb") as f:
            f.write(payload)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp_name, path)
    except BaseException:
        Path(tmp_name).unlink(missing_ok=True)
        raise


def _write_cash_rows(header, body):
    """Publish cash.csv — the source of truth for every cash entry — atomically.

    Writing in place truncates the file first, so a failure part-way through leaves a
    cash ledger missing whatever had not been written yet."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(header)
    writer.writerows(body)
    _publish_bytes(INBOX / "cash.csv", buffer.getvalue().encode("utf-8"))


@contextlib.contextmanager
def _cash_transaction(header, body):
    """Write cash.csv and rebuild the ledger from it, or change neither.

    cash.csv is the source and the JSONL under data/ is derived from it, so between
    the CSV write and the regeneration that follows they disagree by construction. A
    failure in that window — an unknown currency, a full disk, transfer re-detection
    raising — used to leave the new row in the CSV with no trace of it in the ledger,
    and the next ingest would then adopt it as if a human had approved it. Putting the
    old CSV back makes the two move together.
    """
    path = INBOX / "cash.csv"
    previous = path.read_bytes() if path.exists() else None
    _write_cash_rows(header, body)
    try:
        yield
    except BaseException:
        try:
            if previous is None:
                path.unlink(missing_ok=True)
            else:
                _publish_bytes(path, previous)
            ingest.regenerate_cash(mark_transfers=False)
        except Exception:
            pass    # the failure that got us here is the one worth reporting
        raise


def _assert_cash_ledger_valid():
    """Guard cash writes against a cash.csv that references a removed account (possible from
    accounts deleted before the usage guard counted cash rows). Surface it as an actionable
    409 instead of letting cash regeneration crash with a 500."""
    accounts, _ = load_accounts()
    orphans = sorted({row[1] for row in _cash_rows()[1] if len(row) >= 2 and row[1] and row[1] not in accounts})
    if orphans:
        raise HTTPException(409, "The cash ledger still has entries for removed account(s): %s. "
                            "Re-add them on the Accounts tab (their entries come back), then delete "
                            "the entries you no longer want." % ", ".join(orphans))


def _cash_row_matches(row, target):
    try:
        return (len(row) >= 5 and row[0] == target["date"] and row[1] == target["account"]
                and cents(row[2]) == cents(target["amount_original"])
                and (row[3] or "EUR").upper() == target["currency"]
                and row[4] == (target.get("counterparty") or ""))
    except (TypeError, ValueError):
        return False


def _cash_row_index(body, target):
    wanted = int(target["id"].rsplit("#", 1)[1])
    seen = 0
    for index, row in enumerate(body):
        if _cash_row_matches(row, target):
            seen += 1
            if seen == wanted:
                return index
    raise HTTPException(409, "cash.csv no longer matches the stored entry; ingest it again first")


def _rekey_cash_decisions(old_cash, new_cash, removed_ids=(), explicit=None):
    """Move decisions to regenerated occurrence ids without touching bank decisions."""
    removed_ids = set(removed_ids)
    years = {int(t["date"][:4]) for t in old_cash + new_cash}
    decisions = {year: store.decisions(year) for year in years}
    old_decisions = {txn["id"]: decisions[int(txn["date"][:4])].get(txn["id"])
                     for txn in old_cash}
    for txn in old_cash:
        decisions[int(txn["date"][:4])].pop(txn["id"], None)

    old_groups, new_groups = {}, {}
    for txn in old_cash:
        if txn["id"] not in removed_ids:
            old_groups.setdefault(txn["id"].rsplit("#", 1)[0], []).append(txn)
    for txn in new_cash:
        new_groups.setdefault(txn["id"].rsplit("#", 1)[0], []).append(txn)
    for base, old_group in old_groups.items():
        for old, new in zip(old_group, new_groups.get(base, [])):
            if old_decisions.get(old["id"]):
                decisions[int(new["date"][:4])][new["id"]] = old_decisions[old["id"]]
    if explicit:
        txn, carried = explicit
        if carried:
            decisions[int(txn["date"][:4])][txn["id"]] = carried
    for year, values in decisions.items():
        store.save_decisions(year, values)
    for year in sorted(years):
        transfers.mark_internal(year)


@app.post("/api/cash")
def cash(e: CashEntry):
    _validate_cash_values(e)
    _assert_cash_month_open(e.date)
    _assert_cash_ledger_valid()
    old_cash = _cash_snapshot()
    header, body = _cash_rows()
    body.append([e.date, e.account, e.amount, e.currency.upper(), e.description.strip(), e.category])
    with _cash_transaction(header, body):
        new_cash = ingest.regenerate_cash(mark_transfers=False)
        added = new_cash[-1]
        carried = {"category": e.category, "sharing": "shared"} if e.category else None
        _rekey_cash_decisions(old_cash, new_cash, explicit=(added, carried))
    return {"ok": True, "result": "1 new transaction", "id": added["id"]}


class CashDelete(BaseModel):
    year: int
    id: str


@app.post("/api/cash-delete")
def cash_delete(d: CashDelete):
    old_cash = _cash_snapshot()
    target = next((t for t in old_cash if int(t["date"][:4]) == d.year and t["id"] == d.id), None)
    if not target:
        raise HTTPException(404, "unknown cash entry")
    _assert_cash_month_open(target["date"])
    header, body = _cash_rows()
    del body[_cash_row_index(body, target)]
    with _cash_transaction(header, body):
        new_cash = ingest.regenerate_cash(mark_transfers=False)
        _rekey_cash_decisions(old_cash, new_cash, removed_ids={target["id"]})
    return {"ok": True}


class CashEdit(BaseModel):
    year: int          # the entry's current year (for the delete)
    id: str
    date: str
    account: str
    amount: float
    currency: str = "EUR"
    description: str
    category: str = ""


@app.post("/api/cash-edit")
def cash_edit(e: CashEdit):
    proposed = CashEntry(date=e.date, account=e.account, amount=e.amount,
                         currency=e.currency, description=e.description, category=e.category)
    eur, _ = _validate_cash_values(proposed)
    _assert_cash_ledger_valid()
    old_cash = _cash_snapshot()
    target = next((t for t in old_cash if int(t["date"][:4]) == e.year and t["id"] == e.id), None)
    if not target:
        raise HTTPException(404, "unknown cash entry")
    _assert_cash_month_open(target["date"])
    _assert_cash_month_open(e.date)

    carried = dict(store.decisions(e.year).get(e.id, {}))
    carried["category"] = e.category or None
    carried.pop("account", None)
    _validate_decision_fields(carried, {"amount_eur": eur})

    header, body = _cash_rows()
    row_index = _cash_row_index(body, target)
    body[row_index] = [e.date, e.account, e.amount, e.currency.upper(), e.description.strip(), e.category]
    with _cash_transaction(header, body):
        new_cash = ingest.regenerate_cash(mark_transfers=False)
        edited = new_cash[row_index]
        _rekey_cash_decisions(old_cash, new_cash, removed_ids={target["id"]}, explicit=(edited, carried))
    preserved = sorted(k for k in carried if k not in ("category", "account"))
    return {"ok": True, "id": edited["id"], "carried_over": preserved}


@app.post("/api/ingest")
def run_ingest():
    return {"results": [{"file": n, "result": m} for n, m in ingest.run(verbose=False)]}


# ---- Ingest page: staging -> process -> tracked uploads ----
STAGING = INBOX / "staging"
UPLOADS_FILE = DATA / "uploads.json"
INGESTABLE = {".csv", ".xlsx", ".xls"}
PDF_EXT = {".pdf"}
SCRIPTS_DIR = Path(__file__).resolve().parent.parent / "scripts"


def _discover_pdf_extractors():
    """Load extractor manifests from scripts/*/extractor.json."""
    found = {}
    for manifest_path in sorted(SCRIPTS_DIR.glob("*/extractor.json")):
        manifest = read_json(manifest_path)
        required = ("id", "label", "module", "callable")
        missing = [key for key in required if not manifest.get(key)]
        if missing:
            raise RuntimeError("%s is missing %s" % (manifest_path, ", ".join(missing)))
        extractor_id = manifest["id"]
        if extractor_id in found:
            raise RuntimeError("duplicate PDF extractor id '%s'" % extractor_id)
        module = importlib.import_module(manifest["module"])
        extract = getattr(module, manifest["callable"], None)
        if not callable(extract):
            raise RuntimeError("%s callable %s.%s was not found" %
                               (manifest_path, manifest["module"], manifest["callable"]))
        found[extractor_id] = {
            "label": manifest["label"], "description": manifest.get("description", ""),
            "extract": extract, "account_bank_contains": manifest.get("account_bank_contains"),
            "manifest": str(manifest_path),
        }
    return found


PDF_EXTRACTORS = _discover_pdf_extractors()


def _uploads():
    return read_json(UPLOADS_FILE, default={"uploads": []})["uploads"]


def _save_uploads(items):
    write_json(UPLOADS_FILE, {"uploads": items})


def _staging():
    return read_json(STAGING / "staging.json", default={"files": []})["files"]


def _save_staging(items):
    STAGING.mkdir(parents=True, exist_ok=True)
    write_json(STAGING / "staging.json", {"files": items})


def _migrate_pending_pdfs():
    """Move legacy CLI-era pending PDFs back to the actionable staging list."""
    uploads = _uploads()
    pending = [u for u in uploads if u.get("status") == "pending-extraction" and u.get("inbox_file")]
    if not pending:
        return
    staging = _staging()
    staged_hashes = {e.get("hash") for e in staging}
    migrated_ids = set()
    for old in pending:
        if old.get("file_hash") in staged_hashes:
            migrated_ids.add(old["id"])
            continue
        source = INBOX / old["inbox_file"]
        if not source.exists():
            continue
        STAGING.mkdir(parents=True, exist_ok=True)
        safe = re.sub(r"[^A-Za-z0-9._-]", "_", old.get("original_name") or "statement.pdf")
        stored = "%s__%s" % (old["id"], safe)
        shutil.move(str(source), str(STAGING / stored))
        staging.append({
            "id": old["id"], "original_name": old.get("original_name") or safe,
            "stored": stored, "ext": ".pdf", "kind": "pdf",
            "size": (STAGING / stored).stat().st_size, "hash": old.get("file_hash"),
            "account": old.get("account"), "comment": old.get("comment", ""),
            "extractor": None, "allow_review": False,
        })
        staged_hashes.add(old.get("file_hash"))
        migrated_ids.add(old["id"])
    if migrated_ids:
        _save_staging(staging)
        _save_uploads([u for u in uploads if u.get("id") not in migrated_ids])


@app.get("/api/ingest/staging")
def ingest_staging():
    _migrate_pending_pdfs()
    staging = _staging()
    if len(PDF_EXTRACTORS) == 1:
        default_extractor = next(iter(PDF_EXTRACTORS))
        changed = False
        for entry in staging:
            if entry.get("kind") == "pdf" and not entry.get("extractor"):
                entry["extractor"] = default_extractor
                changed = True
        if changed:
            _save_staging(staging)
    return {"files": _staging(),
            "extractors": [{"id": key, "label": value["label"],
                            "description": value.get("description", ""),
                            "account_bank_contains": value.get("account_bank_contains")}
                           for key, value in PDF_EXTRACTORS.items()]}


@app.get("/api/ingest/uploads")
def ingest_uploads():
    return {"uploads": sorted(_uploads(), key=lambda u: u.get("processed_at", ""), reverse=True)}


@app.post("/api/ingest/upload")
async def ingest_upload_file(file: UploadFile = File(...)):
    STAGING.mkdir(parents=True, exist_ok=True)
    raw = await file.read()
    digest = hashlib.sha256(raw).hexdigest()[:16]
    ext = Path(file.filename or "").suffix.lower()
    if ext not in INGESTABLE and ext not in PDF_EXT:
        raise HTTPException(400, "Unsupported file type '%s' (use CSV, XLSX or PDF)." % ext)
    # file-level dedupe: refuse a file already processed or already staged
    dup = next((u for u in _uploads() if u.get("file_hash") == digest), None)
    if dup:
        raise HTTPException(409, "Already processed on %s (%s)." % (dup.get("processed_at", "?")[:10], dup.get("original_name")))
    staging = _staging()
    if any(s["hash"] == digest for s in staging):
        raise HTTPException(409, "This file is already staged.")
    sid = "s_%d_%s" % (int(time.time() * 1000), digest[:6])
    safe = re.sub(r"[^A-Za-z0-9._-]", "_", file.filename or "file")
    stored = sid + "__" + safe
    staged_path = STAGING / stored
    staged_path.write_bytes(raw)
    kind = "pdf" if ext in PDF_EXT else "table"
    entry = {"id": sid, "original_name": file.filename, "stored": sid + "__" + safe,
             "ext": ext, "kind": kind,
             "size": len(raw), "hash": digest, "account": None, "comment": "",
             "extractor": next(iter(PDF_EXTRACTORS)) if kind == "pdf" and len(PDF_EXTRACTORS) == 1 else None,
             "allow_review": False}
    if entry["kind"] == "table":
        try:
            entry["preview"] = ingest.preview_file(staged_path)
        except Exception as exc:
            staged_path.unlink(missing_ok=True)
            raise HTTPException(400, "Could not validate this file: %s No data was imported." % exc)
    staging.append(entry)
    _save_staging(staging)
    return entry


class StagingUpdate(BaseModel):
    id: str
    account: Optional[str] = None
    comment: Optional[str] = None
    extractor: Optional[str] = None
    allow_review: Optional[bool] = None


@app.post("/api/ingest/staging-update")
def ingest_staging_update(u: StagingUpdate):
    accounts, _ = load_accounts()
    if u.account and u.account not in accounts:
        raise HTTPException(400, "unknown account")
    staging = _staging()
    e = next((s for s in staging if s["id"] == u.id), None)
    if not e:
        raise HTTPException(404, "unknown staged file")
    if u.account is not None:
        e["account"] = u.account or None
    if u.comment is not None:
        e["comment"] = u.comment
    if u.extractor is not None:
        if u.extractor and u.extractor not in PDF_EXTRACTORS:
            raise HTTPException(400, "unknown PDF extractor")
        e["extractor"] = u.extractor or None
        e.pop("error", None)
        e.pop("extraction", None)
    if u.allow_review is not None:
        e["allow_review"] = u.allow_review
    _save_staging(staging)
    return e


class StagingDelete(BaseModel):
    id: str


@app.post("/api/ingest/staging-delete")
def ingest_staging_delete(d: StagingDelete):
    staging = _staging()
    e = next((s for s in staging if s["id"] == d.id), None)
    if not e:
        raise HTTPException(404, "unknown staged file")
    (STAGING / e["stored"]).unlink(missing_ok=True)
    _save_staging([s for s in staging if s["id"] != d.id])
    return {"ok": True}


@app.post("/api/ingest/process")
def ingest_process():
    accounts, _ = load_accounts()
    staging = _staging()
    ready = [s for s in staging if s.get("account") and
             (s.get("kind") != "pdf" or s.get("extractor"))]
    if not ready:
        raise HTTPException(400, "No staged files have an account assigned.")
    uploads = _uploads()
    results = []
    remaining = list(staging)
    for e in ready:
        src = STAGING / e["stored"]
        acct = e["account"]
        owner = accounts.get(acct, {}).get("owner")
        try:
            if e["kind"] == "pdf":
                extractor = PDF_EXTRACTORS[e["extractor"]]
                # A mismatched account bank is only an advisory warning in the UI (the reconciliation
                # gate still protects data integrity), so honor the account the user chose here.
                extracted = STAGING / (e["stored"] + ".extracted.csv")
                report = extractor["extract"](src, extracted, bool(e.get("allow_review")))
                e["extraction"] = extraction.storable(report)
                # The extractor saying "ok" is the extractor's opinion of its own work.
                # extraction.admit re-runs opening + rows == closing here, over the file
                # that is actually about to be imported, so a plugin that reports success
                # it did not earn cannot put a transaction in the store.
                try:
                    e["admission"] = extraction.admit(report, extracted)
                except extraction.ExtractionRejected as rejected:
                    extracted.unlink(missing_ok=True)
                    e["error"] = "Extraction failed reconciliation: %s. No data was imported." % rejected
                    results.append({"file": e["original_name"], "status": "error", "detail": e["error"]})
                    continue
                source_stem = "%s__%s" % (acct, e["id"])
                stats = ingest.ingest_upload(extracted, acct, source_stem,
                                             original_name=e["original_name"], admitted=True)
                extracted.unlink(missing_ok=True)
                # The extractor proved these balances against the statement's own
                # arithmetic; recording them turns each import into a checkpoint the
                # doctor re-verifies against the transactions between them. Dropping
                # them, as we used to, threw away the only evidence that a parser read
                # the file correctly.
                if report.get("balance_anchors"):
                    anchor_result = anchors.record(acct, report["balance_anchors"],
                                                   e["original_name"] or "statement.pdf",
                                                   upload=source_stem)
                    # anchors.record() reports what it did with them; how many there were to
                    # begin with is the caller's to add, and omitting it raised KeyError here —
                    # after the transactions were already stored, so the failure was then
                    # reported as "No data was imported" while the data sat in the store.
                    anchor_result["found"] = len(report["balance_anchors"])
                    stats["anchors"] = anchor_result
                    stats["anchor_message"] = ingest._anchor_message(anchor_result)
                safe = re.sub(r"[^A-Za-z0-9._-]", "_", e["original_name"] or "statement.pdf")
                processed_dir = INBOX / "processed"
                processed_dir.mkdir(parents=True, exist_ok=True)
                dest = processed_dir / ("%s__%s__%s" % (acct, e["id"], safe))
                shutil.move(str(src), str(dest))
                log_path = dest.with_suffix(".extract-log.json")
                write_json(log_path, e["extraction"])
                uploads.append({
                    "id": e["id"], "original_name": e["original_name"], "account": acct,
                    "owner": owner, "comment": e.get("comment", ""), "file_hash": e["hash"],
                    "kind": "pdf", "status": "processed", "source_stem": source_stem,
                    "format": "pdf:%s" % e["extractor"], "extractor": e["extractor"],
                    "extraction": e["extraction"], "processed_pdf": dest.name, "extraction_log": log_path.name,
                    "processed_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    "added": stats["added"], "duplicates": stats["duplicates"], "total": stats["total"],
                    "years": stats["years"], "date_min": stats["date_min"], "date_max": stats["date_max"],
                    "anchors": stats["anchors"],
                })
                results.append({"file": e["original_name"], "status": "processed",
                                "detail": "%d transactions extracted and reconciled; %d new entries%s" %
                                          (report["transactions_extracted"], stats["added"],
                                           "; " + stats["anchor_message"] if stats["anchor_message"] else "")})
            else:
                source_stem = "%s__%s" % (acct, e["id"])
                stats = ingest.ingest_upload(src, acct, source_stem, original_name=e["original_name"])
                src.unlink(missing_ok=True)
                uploads.append({
                    "id": e["id"], "original_name": e["original_name"], "account": acct,
                    "owner": owner, "comment": e.get("comment", ""), "file_hash": e["hash"],
                    "kind": "table", "status": "processed", "source_stem": source_stem,
                    "format": stats["format"], "processed_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    "added": stats["added"], "duplicates": stats["duplicates"], "total": stats["total"],
                    "years": stats["years"], "date_min": stats["date_min"], "date_max": stats["date_max"],
                    "anchors": stats["anchors"],
                    # The period the statement declares it covers. For a statement
                    # with no activity this is the only record that the month was
                    # reported at all, so coverage can stop calling it missing.
                    "period": stats.get("period"),
                })
                if stats["total"]:
                    detail = "%d new (%d duplicates skipped), %s–%s%s" % (
                        stats["added"], stats["duplicates"], stats["date_min"] or "?", stats["date_max"] or "?",
                        "; " + stats["anchor_message"] if stats["anchor_message"] else "")
                else:
                    # A statement covering a period with no activity: a "?–?" date
                    # range and a missing-anchor note would both be noise, not news.
                    detail = "empty statement — no activity in this period"
                results.append({"file": e["original_name"], "status": "processed", "detail": detail})
            remaining = [s for s in remaining if s["id"] != e["id"]]
        except Exception as ex:  # noqa: BLE001
            e["error"] = "%s No data was imported." % ex
            results.append({"file": e["original_name"], "status": "error", "detail": e["error"]})
    _save_staging(remaining)
    _save_uploads(uploads)
    return {"results": results}


class UploadDelete(BaseModel):
    id: str


@app.get("/api/ingest/upload-contents")
def ingest_upload_contents(id: str):
    """Read-only: what deleting this upload would remove, for the confirmation."""
    u = next((x for x in _uploads() if x["id"] == id), None)
    if not u:
        raise HTTPException(404, "unknown upload")
    if not u.get("source_stem"):
        return {"transactions": 0, "decisions": 0, "years": {}, "closed_months": []}
    return ingest.upload_contents(u["source_stem"])


@app.post("/api/ingest/upload-delete")
def ingest_upload_delete(d: UploadDelete):
    uploads = _uploads()
    u = next((x for x in uploads if x["id"] == d.id), None)
    if not u:
        raise HTTPException(404, "unknown upload")
    if u.get("source_stem"):
        try:
            ingest.delete_upload(u["source_stem"])
        except ValueError as exc:   # closed month; same 409 the decision endpoints use
            raise HTTPException(409, str(exc))
    if u.get("kind") == "pdf" and u.get("processed_pdf"):
        (INBOX / "processed" / u["processed_pdf"]).unlink(missing_ok=True)
        if u.get("extraction_log"):
            (INBOX / "processed" / u["extraction_log"]).unlink(missing_ok=True)
    elif u.get("kind") == "pdf" and u.get("inbox_file"):
        (INBOX / u["inbox_file"]).unlink(missing_ok=True)
        (INBOX / "processed" / u["inbox_file"]).unlink(missing_ok=True)
    _save_uploads([x for x in uploads if x["id"] != d.id])
    return {"ok": True}


class MonthState(BaseModel):
    year: int
    month: int
    state: str  # open | closed


@app.post("/api/close")
def close_month(m: MonthState):
    if m.state not in ("open", "closed") or not 1 <= m.month <= 12:
        raise HTTPException(400, "state must be open or closed and month must be 1–12")
    states = store.months_state(m.year)
    states["%d-%02d" % (m.year, m.month)] = m.state
    store.save_months_state(m.year, states)
    # Record what the month contained at the moment it was called settled, so any
    # later change to it — by rule, by detection, by re-ingest — is visible instead
    # of silent. Reopening withdraws the claim, so the baseline goes with it.
    if m.state == "closed":
        closings.record(m.year, m.month)
        if all(states.get("%d-%02d" % (m.year, month)) == "closed" for month in range(1, 13)):
            closings.record_year(m.year)   # this month completed the year
    else:
        closings.drop(m.year, m.month)
        closings.drop_year(m.year)         # the year is no longer fully settled
    return {"ok": True}


class YearState(BaseModel):
    year: int
    state: str  # open | closed


@app.post("/api/close-year")
def close_year(y: YearState):
    """Close/reopen the whole year = all 12 months at once (reuses the month
    lock; year costs live inside those months by date, so they lock too)."""
    if y.state not in ("open", "closed"):
        raise HTTPException(400, "state must be open or closed")
    states = store.months_state(y.year)
    for month in range(1, 13):
        states["%d-%02d" % (y.year, month)] = y.state
    store.save_months_state(y.year, states)
    for month in range(1, 13):
        closings.record(y.year, month) if y.state == "closed" else closings.drop(y.year, month)
    # The annual settlement is the binding figure and does not follow from the twelve
    # months: a ratio override applied to the year moves it while every month stays put.
    closings.record_year(y.year) if y.state == "closed" else closings.drop_year(y.year)
    return {"ok": True}


class NewSubcategory(BaseModel):
    group: str
    name: str


@app.post("/api/category")
def add_subcategory(s: NewSubcategory):
    data = read_json(RULES / "categories.json")
    for c in data["categories"]:
        if c["slug"] == s.group:
            if not c.get("dynamic"):
                raise HTTPException(400, "group '%s' does not accept dynamic subcategories" % s.group)
            slug = s.name.lower().replace(" ", "-")
            if any(x["slug"] == slug for x in c["subs"]):
                raise HTTPException(400, "subcategory exists")
            c["subs"].append({"slug": slug, "name": s.name})
            write_json(RULES / "categories.json", data)
            return {"ok": True, "slug": "%s/%s" % (s.group, slug)}
    raise HTTPException(404, "unknown group")


def _slugify(name):
    return re.sub(r"[^a-z0-9-]", "", name.strip().lower().replace(" ", "-")).strip("-")


def _category_usage():
    """{full_slug: txn_count} across all years, counting split parts too."""
    counts = {}
    for y in store.years():
        for t in store.effective_year(y):
            parts = t.get("splits") or [{"category": t.get("category")}]
            for p in parts:
                cat = p.get("category")
                if cat:
                    counts[cat] = counts.get(cat, 0) + 1
    return counts


def _category_rule_refs():
    counts = {}
    for r in read_json(RULES / "merchant-rules.json").get("rules", []):
        c = r.get("category")
        if c:
            counts[c] = counts.get(c, 0) + 1
    return counts


@app.get("/api/category-usage")
def category_usage():
    return {"usage": _category_usage(), "rules": _category_rule_refs()}


class CategoryAdd(BaseModel):
    parent: Optional[str] = None   # None -> new group; else group slug for a new sub
    name: str
    type: str = "expense"  # only used for a new group: income | expense


@app.post("/api/category-add")
def category_add(c: CategoryAdd):
    data = read_json(RULES / "categories.json")
    slug = _slugify(c.name)
    if not slug:
        raise HTTPException(400, "name must contain letters or digits")
    if c.parent is None:
        if c.type not in ("income", "expense"):
            raise HTTPException(400, "type must be 'income' or 'expense'")
        if any(g["slug"] == slug for g in data["categories"]):
            raise HTTPException(400, "a category with that name already exists")
        data["categories"].append({"slug": slug, "name": c.name.strip(), "type": c.type, "subs": []})
        write_json(RULES / "categories.json", data)
        return {"ok": True, "slug": slug}
    for g in data["categories"]:
        if g["slug"] == c.parent:
            if any(s["slug"] == slug for s in g["subs"]):
                raise HTTPException(400, "a subcategory with that name already exists here")
            g["subs"].append({"slug": slug, "name": c.name.strip()})
            write_json(RULES / "categories.json", data)
            return {"ok": True, "slug": "%s/%s" % (c.parent, slug)}
    raise HTTPException(404, "unknown group")


def _find_category(data, slug):
    """Return (group, sub_or_None) for a 'group' or 'group/sub' slug, or (None, None)."""
    if "/" in slug:
        g_slug, s_slug = slug.split("/", 1)
        for g in data["categories"]:
            if g["slug"] == g_slug:
                for s in g["subs"]:
                    if s["slug"] == s_slug:
                        return g, s
        return None, None
    for g in data["categories"]:
        if g["slug"] == slug:
            return g, None
    return None, None


class CategoryRename(BaseModel):
    slug: str
    name: str


@app.post("/api/category-rename")
def category_rename(r: CategoryRename):
    name = r.name.strip()
    if not name:
        raise HTTPException(400, "name required")
    data = read_json(RULES / "categories.json")
    g, s = _find_category(data, r.slug)
    if g is None:
        raise HTTPException(404, "unknown category")
    (s or g)["name"] = name
    write_json(RULES / "categories.json", data)
    return {"ok": True}


class CategoryArchive(BaseModel):
    slug: str
    archived: bool


@app.post("/api/category-archive")
def category_archive(a: CategoryArchive):
    data = read_json(RULES / "categories.json")
    g, s = _find_category(data, a.slug)
    if g is None:
        raise HTTPException(404, "unknown category")
    target = s or g
    if a.archived:
        if s is not None and s.get("ratio_income"):
            raise HTTPException(400, "This is the settlement income basis and can't be archived.")
        active_income = [x for grp in data["categories"] if grp.get("type") == "income"
                         for x in grp["subs"] if not x.get("archived")]
        if s is not None and s in active_income and len(active_income) == 1:
            raise HTTPException(400, "Can't archive the last active income subcategory.")
        target["archived"] = True
    else:
        target.pop("archived", None)
    write_json(RULES / "categories.json", data)
    return {"ok": True}


class CategoryStyle(BaseModel):
    slug: str
    icon: Optional[str] = None
    color: Optional[str] = None


@app.post("/api/category-style")
def category_style(s: CategoryStyle):
    """Set icon (Material Symbol name) and/or color (#rrggbb) on a main category."""
    data = read_json(RULES / "categories.json")
    g, sub = _find_category(data, s.slug)
    if g is None or sub is not None:
        raise HTTPException(404, "unknown main category")
    if s.color is not None:
        if not re.match(r"^#[0-9a-fA-F]{6}$", s.color):
            raise HTTPException(400, "color must be a #rrggbb hex")
        g["color"] = s.color
    if s.icon is not None:
        g["icon"] = _icon_name(s.icon, "category icon")
    write_json(RULES / "categories.json", data)
    return {"ok": True}


class CategoryWatch(BaseModel):
    slug: str
    watch: bool


@app.post("/api/category-watch")
def category_watch(w: CategoryWatch):
    """Flag a subcategory to watch its spend closely (drives the dashboard
    watch-list charts). Watching is a subcategory-only concern."""
    data = read_json(RULES / "categories.json")
    g, sub = _find_category(data, w.slug)
    if g is None or sub is None:
        raise HTTPException(404, "unknown subcategory")
    if w.watch:
        sub["watch"] = True
    else:
        sub.pop("watch", None)
    write_json(RULES / "categories.json", data)
    return {"ok": True}


class CategoryDelete(BaseModel):
    slug: str


@app.post("/api/category-delete")
def category_delete(d: CategoryDelete):
    data = read_json(RULES / "categories.json")
    g, s = _find_category(data, d.slug)
    if g is None:
        raise HTTPException(404, "unknown category")
    usage = _category_usage()
    rules = _category_rule_refs()
    if s is not None:
        full = d.slug
        if usage.get(full) or rules.get(full):
            raise HTTPException(409, "Used in %d transaction(s) / %d rule(s) — archive it instead."
                                % (usage.get(full, 0), rules.get(full, 0)))
        g["subs"] = [x for x in g["subs"] if x["slug"] != s["slug"]]
    else:
        if g["subs"]:
            raise HTTPException(409, "Delete or move its subcategories first, or archive the group instead.")
        prefix = g["slug"] + "/"
        used = sum(v for k, v in usage.items() if k.startswith(prefix)) + usage.get(g["slug"], 0)
        ruled = sum(v for k, v in rules.items() if k.startswith(prefix)) + rules.get(g["slug"], 0)
        if used or ruled:
            raise HTTPException(409, "Used in %d transaction(s) / %d rule(s) — archive it instead." % (used, ruled))
        data["categories"] = [x for x in data["categories"] if x["slug"] != g["slug"]]
    write_json(RULES / "categories.json", data)
    # clean up an orphaned budget for a deleted slug
    budgets = read_json(BUDGETS_FILE, default={"budgets": {}}).get("budgets", {})
    if d.slug in budgets:
        del budgets[d.slug]
        write_json(BUDGETS_FILE, {"_help": "Monthly targets in EUR per category slug.", "budgets": budgets})
    return {"ok": True}


def _assert_open(year, txn_id):
    for t in store.load_year_raw(year):
        if t["id"] == txn_id:
            key = t["date"][:7]
            if store.months_state(year).get(key) == "closed":
                raise HTTPException(409, "Month %s is closed. Reopen it first." % key)
            return t
    raise HTTPException(404, "unknown transaction")


RECEIPTS.mkdir(exist_ok=True)
FEEDBACK_FILES.mkdir(parents=True, exist_ok=True)
app.mount("/receipts", StaticFiles(directory=RECEIPTS), name="receipts")
app.mount("/static", StaticFiles(directory=STATIC), name="static")
